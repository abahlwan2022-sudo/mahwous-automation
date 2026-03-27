"""
مهووس — مركز التحكم الشامل v11.0
تطبيق Streamlit للإنتاج: المسار الآلي، المقارنة، معالج SEO، التدقيق، منتج سريع.
المفاتيح: متغيرات البيئة أو Streamlit secrets (ANTHROPIC_API_KEY، GOOGLE_API_KEY، GOOGLE_CSE_ID).
رابط المتجر: MAHWOUS_SITE_BASE (افتراضي https://mahwous.com).
"""

import io
import json
import logging
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Union

import pandas as pd
import requests
import streamlit as st

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from mahwous_core import (
    StrictFilterOptions,
    apply_strict_pipeline_filters,
    validate_input_dataframe,
    validate_export_product_dataframe,
    validate_export_seo_dataframe,
    validate_export_brands_list,
    parse_price_numeric,
    format_salla_date_yyyy_mm_dd,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz as rf_fuzz
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

try:
    import anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE CONFIG                                                    ║
# ╚══════════════════════════════════════════════════════════════════╝
st.set_page_config(
    page_title="مهووس | مركز التحكم",
    page_icon="🌸",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  GLOBAL CSS — Arabic RTL + Gold Theme                          ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cairo:wght@300;400;600;700;900&display=swap');

/* ── Global Reset & RTL ─────────────────────────── */
*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"], .stApp, .main, section, div {
  font-family: 'Cairo', sans-serif !important;
}
.stApp { background-color: #f5f0e8; }

/* ── Sidebar ─────────────────────────────────────── */
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #0f0e0d 0%, #1c1610 100%) !important;
  border-left: 1px solid rgba(184,147,58,0.25);
}
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div { color: #e0d0b0 !important; }
section[data-testid="stSidebar"] .stRadio label { color: #c8b080 !important; }

/* ── Top Header ──────────────────────────────────── */
.mhw-header {
  background: linear-gradient(135deg, #0f0e0d 0%, #1c1610 50%, #2a1e08 100%);
  border: 1px solid rgba(184,147,58,0.3);
  border-radius: 14px;
  padding: 16px 24px;
  display: flex;
  align-items: center;
  gap: 16px;
  margin-bottom: 20px;
  box-shadow: 0 8px 32px rgba(0,0,0,0.2);
}
.mhw-header .emblem {
  width: 52px; height: 52px;
  background: linear-gradient(135deg, #b8933a, #e0b84a);
  border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 26px; font-weight: 900; color: #0f0e0d;
  box-shadow: 0 0 24px rgba(184,147,58,0.6);
  flex-shrink: 0;
}
.mhw-header h1 {
  color: #b8933a; font-size: 1.55rem;
  margin: 0; line-height: 1.2;
}
.mhw-header p { color: rgba(255,255,255,0.38); font-size: 0.78rem; margin: 0; }

/* ── Section Title ───────────────────────────────── */
.sec-title {
  display: flex; align-items: center; gap: 10px;
  margin: 22px 0 14px; direction: rtl;
}
.sec-title .bar {
  width: 5px; height: 24px; border-radius: 3px;
  background: linear-gradient(180deg, #b8933a, #e0b84a);
}
.sec-title h3 { margin: 0; font-size: 1.05rem; font-weight: 800; color: #1a1208; }

/* ── Stats Bar ───────────────────────────────────── */
.stats-bar { display: flex; gap: 12px; flex-wrap: wrap; margin: 14px 0; }
.stat-box {
  flex: 1; min-width: 110px;
  background: white;
  border: 1px solid rgba(184,147,58,0.22);
  border-radius: 12px;
  padding: 14px 16px; text-align: center;
  box-shadow: 0 2px 10px rgba(0,0,0,0.05);
}
.stat-box .n  { font-size: 1.9rem; font-weight: 900; color: #b8933a; line-height: 1; }
.stat-box .lb { font-size: 0.73rem; color: #7a6e60; margin-top: 3px; }

/* ── Upload Zone ─────────────────────────────────── */
.upload-zone {
  border: 2px dashed rgba(184,147,58,0.38);
  border-radius: 16px; padding: 2.5rem;
  text-align: center;
  background: rgba(184,147,58,0.035);
  transition: all 0.2s;
}
.upload-zone:hover {
  border-color: #b8933a;
  background: rgba(184,147,58,0.07);
}
.uz-icon  { font-size: 3rem; }
.uz-title { font-size: 1.08rem; font-weight: 800; color: #1a1208; margin: 6px 0 3px; }
.uz-sub   { font-size: 0.8rem; color: #9a8e80; }

/* ── Tool Tabs ───────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
  background: rgba(184,147,58,0.06);
  border-radius: 10px; padding: 4px; gap: 4px;
  border-bottom: none !important;
}
.stTabs [data-baseweb="tab"] {
  border-radius: 8px !important;
  font-family: 'Cairo', sans-serif !important;
  font-weight: 700 !important; font-size: 0.82rem !important;
  padding: 8px 14px !important;
}
.stTabs [aria-selected="true"] {
  background: linear-gradient(135deg, #b8933a, #e0b84a) !important;
  color: #0f0e0d !important;
}

/* ── Alerts ──────────────────────────────────────── */
.al-info {
  background: #e8f4fd; border-right: 4px solid #1976d2;
  border-radius: 8px; padding: 10px 14px;
  font-size: 0.84rem; color: #0d3c6e; margin: 8px 0;
  direction: rtl;
}
.al-ok {
  background: #e8f5e9; border-right: 4px solid #388e3c;
  border-radius: 8px; padding: 10px 14px;
  font-size: 0.84rem; color: #1b5020; margin: 8px 0;
  direction: rtl;
}
.al-warn {
  background: #fff8e1; border-right: 4px solid #f9a825;
  border-radius: 8px; padding: 10px 14px;
  font-size: 0.84rem; color: #5d4300; margin: 8px 0;
  direction: rtl;
}

/* ── Buttons ─────────────────────────────────────── */
div.stButton > button {
  font-family: 'Cairo', sans-serif !important;
  font-weight: 700 !important;
}
div.stButton > button[kind="primary"] {
  background: linear-gradient(135deg, #0f0e0d, #2a1e08) !important;
  color: #b8933a !important;
  border: none !important;
}
div.stButton > button:hover { opacity: 0.88 !important; }

/* ── Gold Divider ────────────────────────────────── */
.gdiv {
  height: 1px; border: none; margin: 20px 0;
  background: linear-gradient(90deg, transparent, rgba(184,147,58,0.4), transparent);
}

/* ── Progress Item ───────────────────────────────── */
.prog-ok  { background:#e8f5e9; border-radius:8px; padding:6px 12px; margin:3px 0; font-size:0.82rem; color:#1b5020; }
.prog-err { background:#fdecea; border-radius:8px; padding:6px 12px; margin:3px 0; font-size:0.82rem; color:#b71c1c; }
.prog-run { background:#fff8e1; border-radius:8px; padding:6px 12px; margin:3px 0; font-size:0.82rem; color:#e65100; }

/* ── Badges ──────────────────────────────────────── */
.badge-ok   { display:inline-block; background:#e8f5e9; color:#2d7a4f; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }
.badge-miss { display:inline-block; background:#fafafa; color:#9e9e9e; padding:2px 10px; border-radius:20px; font-size:0.72rem; }
.badge-new  { display:inline-block; background:#fff3e0; color:#e65100; padding:2px 10px; border-radius:20px; font-size:0.72rem; font-weight:700; }

/* ── Compare Card ────────────────────────────────── */
.cmp-card {
  background: white;
  border: 1px solid rgba(184,147,58,0.25);
  border-radius: 12px;
  padding: 14px;
  margin-bottom: 12px;
  direction: rtl;
}
.cmp-card.suspect { border-color: #f9a825; background: #fffde7; }
.cmp-card.exact   { border-color: #388e3c; background: #f1f8e9; }
.cmp-card img { width: 80px; height: 80px; object-fit: cover; border-radius: 8px; }
.cmp-pct { font-size: 1.2rem; font-weight: 900; color: #f9a825; }

/* ── Footer ──────────────────────────────────────── */
.mhw-footer {
  text-align: center; color: #9a8e80;
  font-size: 0.76rem; padding: 16px 0 8px;
  border-top: 1px solid rgba(184,147,58,0.15);
  margin-top: 32px;
}
</style>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  SALLA EXACT SCHEMAS                                            ║
# ╚══════════════════════════════════════════════════════════════════╝
# ترتيب أعمدة «بيانات المنتج» كما في تصدير سلة (تحديث/تعديل منتجات)
SALLA_COLS = [
    "No.",
    "النوع ", "أسم المنتج", "تصنيف المنتج", "صورة المنتج",
    "وصف صورة المنتج", "نوع المنتج", "سعر المنتج", "الكمية المتوفرة", "الوصف",
    "هل يتطلب شحن؟", "رمز المنتج sku", "سعر التكلفة", "السعر المخفض",
    "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
    "اقصي كمية لكل عميل", "إخفاء خيار تحديد الكمية",
    "اضافة صورة عند الطلب", "الوزن", "وحدة الوزن",
    "حالة المنتج",
    "الماركة", "العنوان الترويجي", "تثبيت المنتج",
    "الباركود", "السعرات الحرارية", "MPN", "GTIN",
    "خاضع للضريبة ؟", "سبب عدم الخضوع للضريبة",
    "[1] الاسم", "[1] النوع", "[1] القيمة", "[1] الصورة / اللون",
    "[2] الاسم", "[2] النوع", "[2] القيمة", "[2] الصورة / اللون",
    "[3] الاسم", "[3] النوع", "[3] القيمة", "[3] الصورة / اللون",
]

SALLA_SEO_COLS = [
    "No. (غير قابل للتعديل)",
    "اسم المنتج (غير قابل للتعديل)",
    "رابط مخصص للمنتج (SEO Page URL)",
    "عنوان صفحة المنتج (SEO Page Title)",
    "وصف صفحة المنتج (SEO Page Description)",
]

SALLA_PRICE_COLS = [
    "No.", "النوع ", "أسم المنتج", "رمز المنتج sku",
    "سعر المنتج", "سعر التكلفة", "السعر المخفض",
    "تاريخ بداية التخفيض", "تاريخ نهاية التخفيض",
]

# Salla brands file exact columns
SALLA_BRANDS_COLS = [
    "اسم الماركة",
    "وصف مختصر عن الماركة",
    "صورة شعار الماركة",
    "(إختياري) صورة البانر",
    "(Page Title) عنوان صفحة العلامة التجارية",
    "(SEO Page URL) رابط صفحة العلامة التجارية",
    "(Page Description) وصف صفحة العلامة التجارية",
]

# Editor shows these by default (rest hidden unless user selects)
EDITOR_COLS = [
    "No.", "النوع ", "أسم المنتج", "الماركة", "تصنيف المنتج",
    "سعر المنتج", "رمز المنتج sku", "صورة المنتج",
    "وصف صورة المنتج", "حالة المنتج", "السعر المخفض",
]

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

MAHWOUS_SITE_BASE = os.environ.get("MAHWOUS_SITE_BASE", "https://mahwous.com").rstrip("/")
PUBLIC_APP_URL = os.environ.get(
    "PUBLIC_APP_URL",
    "https://mahwous-automation-production.up.railway.app",
).rstrip("/")


def mahwous_brand_url(slug: str) -> str:
    """رابط صفحة ماركة داخل نطاق المتجر (slug من to_slug أو page_url)."""
    s = str(slug or "").strip().strip("/")
    if not s:
        return f"{MAHWOUS_SITE_BASE}/brands"
    return f"{MAHWOUS_SITE_BASE}/brands/{s}"


def mahwous_category_url(path: str) -> str:
    """مسار تصنيف تحت نطاق المتجر، مثل categories/mens-perfumes."""
    p = str(path or "").strip().lstrip("/")
    return f"{MAHWOUS_SITE_BASE}/{p}" if p else MAHWOUS_SITE_BASE


def _get_secret(*keys: str) -> str:
    """قراءة مفاتيح API من Streamlit secrets ثم متغيرات البيئة."""
    for k in keys:
        try:
            sec = getattr(st, "secrets", None)
            if sec is not None and k in sec:
                v = sec[k]
                if v is not None and str(v).strip():
                    return str(v).strip()
        except Exception:
            pass
        v = os.environ.get(k, "")
        if v and str(v).strip():
            return str(v).strip()
    return ""


def _effective_anthropic_api_key() -> str:
    """مفتاح Claude من الجلسة أو secrets/البيئة (للمسار الآلي والتحقق من المشبوه)."""
    try:
        k = str(st.session_state.get("api_key", "") or "").strip()
    except Exception:
        k = ""
    if k:
        return k
    return _get_secret("ANTHROPIC_API_KEY")


# ╔══════════════════════════════════════════════════════════════════╗
# ║  LOGGING — تتبع الأخطاء والعمليات                                ║
# ╚══════════════════════════════════════════════════════════════════╝
def configure_app_logging() -> logging.Logger:
    """تهيئة التسجيل: ملف داخل data/logs + الطرفية."""
    log = logging.getLogger("mahwous")
    if log.handlers:
        return log
    log.setLevel(logging.INFO)
    try:
        os.makedirs(os.path.join(DATA_DIR, "logs"), exist_ok=True)
        fh = logging.FileHandler(
            os.path.join(DATA_DIR, "logs", "mahwous_app.log"),
            encoding="utf-8",
        )
        fh.setLevel(logging.INFO)
        fh.setFormatter(logging.Formatter(
            "%(asctime)s | %(levelname)s | %(name)s | %(message)s"
        ))
        log.addHandler(fh)
    except OSError:
        pass
    sh = logging.StreamHandler()
    sh.setLevel(logging.WARNING)
    sh.setFormatter(logging.Formatter("%(levelname)s | %(message)s"))
    log.addHandler(sh)
    return log


APP_LOG = configure_app_logging()


def anthropic_messages_create(client, **kwargs):
    """استدعاء Anthropic Messages API مع إعادة محاولة عند 429 / تجاوز المعدل."""
    last_err = None
    for attempt in range(3):
        try:
            return client.messages.create(**kwargs)
        except Exception as e:
            last_err = e
            err_l = str(e).lower()
            name = type(e).__name__
            retryable = (
                "429" in err_l
                or "rate limit" in err_l
                or "too many requests" in err_l
                or "rate_limit" in name.lower()
                or "overloaded" in err_l
            )
            if retryable and attempt < 2:
                time.sleep(2)
                APP_LOG.warning("anthropic retry %s/3 after %s: %s", attempt + 1, name, e)
                continue
            raise
    if last_err:
        raise last_err
    raise RuntimeError("anthropic_messages_create: unreachable")


def _parse_json_object_from_llm_text(raw: str, *, context: str = "") -> dict:
    """
    يستخرج أول كائن JSON من ردّ النموذج (غالباً Claude) مع دعم أغلفة Markdown ```json ... ```.
    يتعامل مع الأقواس المتداخلة عبر JSONDecoder؛ يجرّب عدة مواضع لـ «{» عند الفشل.
    عند تعذّر التحليل: يُسجّل خطأ واضح (ملف + طرفية) ويعيد {} آمناً للواجهة.
    """
    empty: dict = {}
    if raw is None:
        APP_LOG.error("LLM JSON: raw is None (context=%s)", context)
        print(f"[mahwous] LLM JSON: empty raw (context={context!r})", file=sys.stderr)
        return empty
    s = str(raw).strip()
    if not s:
        APP_LOG.error("LLM JSON: empty string (context=%s)", context)
        print(f"[mahwous] LLM JSON: empty string (context={context!r})", file=sys.stderr)
        return empty
    if s.startswith("```"):
        lines = s.split("\n")
        if lines and lines[0].strip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    dec = json.JSONDecoder()
    positions = []
    pos = 0
    while len(positions) < 48:
        i = s.find("{", pos)
        if i < 0:
            break
        positions.append(i)
        pos = i + 1
    if not positions:
        APP_LOG.error("LLM JSON: no '{' in response (context=%s) snippet=%r", context, s[:400])
        print(f"[mahwous] LLM JSON: no '{{' (context={context!r})", file=sys.stderr)
        return empty
    last_err = None
    for start in positions:
        try:
            obj, _end = dec.raw_decode(s, start)
            if isinstance(obj, dict):
                return obj
            last_err = "root is not a JSON object"
        except json.JSONDecodeError as e:
            last_err = str(e)
            continue
    APP_LOG.error(
        "LLM JSON: unparseable after %d brace position(s) (context=%s) last_err=%s snippet=%r",
        len(positions),
        context,
        last_err,
        s[:500],
    )
    print(
        f"[mahwous] LLM JSON unparseable (context={context!r}): {last_err}",
        file=sys.stderr,
    )
    return empty


# ╔══════════════════════════════════════════════════════════════════╗
# ║  AI SYSTEM PROMPT — خبير وصف منتجات مهووس v4.5                ║
# ╚══════════════════════════════════════════════════════════════════╝
AI_SYSTEM = """أنت خبير كتابة أوصاف عطور فاخرة تعمل حصرياً لمتجر "مهووس" السعودي.

قواعد صارمة لا تُكسر:
- ممنوع منعاً باتاً استخدام الرموز التعبيرية (Emojis) نهائياً
- التركيز يُكتب دائماً: "أو دو بارفيوم"
- أسلوبك: راقٍ 40%، ودود 25%، رومانسي 20%، تسويقي مقنع 15%
- الطول: 1200-1500 كلمة بالضبط
- الإخراج HTML خالص فقط — لا نص خارج الوسوم
- استخدم <strong> للكلمات المفتاحية
- الروابط الداخلية: استخدم <a href="https://mahwous.com/brands/[slug]" target="_blank">[اسم الماركة]</a>
- المكونات: اذكر مكونات حقيقية موثوقة إذا عرفتها، وإلا اذكر مكونات تقريبية منطقية للعائلة العطرية

هيكل الوصف الإلزامي:
<h2>[عطر/تستر] [الماركة] [الاسم] [التركيز] [الحجم] [للجنس]</h2>
<p>فقرة افتتاحية عاطفية 100-150 كلمة، الكلمة المفتاحية في أول 50 كلمة، دعوة للشراء.</p>
<h3>تفاصيل المنتج</h3>
<ul>
<li><strong>الماركة:</strong> [مع رابط داخلي]</li>
<li><strong>الاسم:</strong></li>
<li><strong>الجنس:</strong></li>
<li><strong>العائلة العطرية:</strong></li>
<li><strong>الحجم:</strong></li>
<li><strong>التركيز:</strong> أو دو بارفيوم</li>
<li><strong>سنة الإصدار:</strong></li>
<li><strong>نوع المنتج:</strong> [تستر / عادي]</li>
</ul>
<h3>رحلة العطر - الهرم العطري</h3>
<p>وصف حسي شاعري للعطر كاملاً.</p>
<ul>
<li><strong>المقدمة (Top Notes):</strong> [المكونات الحقيقية أو التقريبية]</li>
<li><strong>القلب (Heart Notes):</strong> [المكونات الحقيقية أو التقريبية]</li>
<li><strong>القاعدة (Base Notes):</strong> [المكونات الحقيقية أو التقريبية]</li>
</ul>
<h3>لماذا تختار هذا العطر؟</h3>
<ul>
<li><strong>الثبات والفوحان:</strong> [وصف دقيق]</li>
<li><strong>التميز والأصالة:</strong> [وصف دقيق]</li>
<li><strong>القيمة الاستثنائية:</strong> [وصف دقيق]</li>
<li><strong>الجاذبية المضمونة:</strong> [وصف دقيق]</li>
</ul>
<h3>متى وأين ترتديه؟</h3>
<p>الفصول المناسبة، أوقات الاستخدام، المناسبات الملائمة.</p>
<h3>لمسة خبير من مهووس</h3>
<p>تقييم للفوحان (1-10) والثبات (1-10) ونصيحة رش احترافية.</p>
<h3>الأسئلة الشائعة</h3>
<ul>
<li><strong>كم يدوم العطر؟</strong> [إجابة دقيقة]</li>
<li><strong>هل يناسب الاستخدام اليومي؟</strong> [إجابة]</li>
<li><strong>ما الفرق بين التستر والعطر العادي؟</strong> [إجابة]</li>
<li><strong>ما العائلة العطرية؟</strong> [إجابة]</li>
<li><strong>هل يناسب الطقس الحار في السعودية؟</strong> [إجابة]</li>
<li><strong>ما مناسبات ارتداء هذا العطر؟</strong> [إجابة]</li>
</ul>
<h3>اكتشف أكثر من مهووس</h3>
<p>روابط داخلية لعطور مشابهة: <a href="https://mahwous.com/brands/[slug]" target="_blank">[عطور الماركة]</a> | <a href="https://mahwous.com/categories/mens-perfumes" target="_blank">[عطور رجالية]</a></p>
<p><strong>عالمك العطري يبدأ من مهووس.</strong> أصلي 100% | شحن سريع داخل السعودية.</p>

في نهاية الوصف أضف قسم SEO منفصل بصيغة JSON:
<!--SEO_DATA
{
  "page_title": "...",
  "meta_description": "...",
  "url_slug": "...",
  "alt_text": "...",
  "tags": ["...", "..."]
}
SEO_DATA-->""".replace("https://mahwous.com", MAHWOUS_SITE_BASE)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  SESSION STATE INIT                                             ║
# ╚══════════════════════════════════════════════════════════════════╝
def _init_state():
    defaults = {
        "api_key":        _get_secret("ANTHROPIC_API_KEY"),
        "google_api":     _get_secret("GOOGLE_API_KEY"),
        "google_cse":     _get_secret("GOOGLE_CSE_ID"),
        # Reference data
        "brands_df":      None,
        "categories_df":  None,
        # Universal Processor state
        "up_raw":         None,   # raw uploaded df
        "up_df":          None,   # restructured Salla df
        "up_seo":         None,   # SEO companion df
        "up_filename":    "",
        "up_mapped":      False,
        # Quick Add list
        "qa_rows":        [],
        # Comparison page state
        "cmp_new_df":     None,   # new products file
        "cmp_store_df":   None,   # store master file
        "cmp_results":    None,   # comparison results df
        "cmp_approved":   {},     # {idx: True/False} user decisions
        "cmp_edit_name":  {},    # {_idx: str} تعديل اسم مؤقت
        "cmp_from_pipe":  False,
        # SEO Processor (standalone)
        "seo_proc_df":    None,
        "seo_proc_gen":   None,
        "seo_proc_full":  None,
        # New brands generated
        "new_brands":     [],     # list of dicts for new brands
        # Store Audit page state
        "audit_df":       None,   # ملف المتجر للتدقيق
        "audit_results":  None,   # نتائج التدقيق
        # Auto Pipeline state
        "pipe_store_df":   None,
        "pipe_comp_dfs":   [],
        "pipe_results":    None,
        "pipe_approved":   None,
        "pipe_new_brands": [],
        "pipe_seo_df":     None,
        "pipe_step":       0,
        "pipe_running":    False,
        # فلاتر الاستبعاد الصارمة (قبل المقارنة)
        "pipe_fx_samples":     False,
        "pipe_fx_accessories": False,
        "pipe_fx_brand":       False,
        "pipe_fx_volume":      False,
        "pipe_filter_stats":   None,  # dict إحصائيات آخر تطبيق للفلاتر
        "pipe_store_dedup_dropped": 0,  # صفوف أُزيلت لأنها موجودة في ملف المتجر
        "pipe_missing_brands_df": None,  # ماركات غير معروفة مستخرجة من المنافسين
        "pipe_session_brands": [],  # ماركات أُثريت/أُضيفت في نفس جلسة المسار (Audit #14)
        # Page
        "page":           "pipeline",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    if st.session_state.get("page") in ("compare_v2", "processor", "brands"):
        st.session_state.page = "pipeline"
    if st.session_state.get("page") in ("competitor_gap", "new_products_filter"):
        st.session_state.page = "pipeline"
    if st.session_state.get("page") == "store_checker":
        st.session_state.page = "store_audit"

    # Auto-load bundled reference CSVs
    if st.session_state.brands_df is None:
        p = os.path.join(DATA_DIR, "brands.csv")
        if os.path.exists(p):
            try:
                st.session_state.brands_df = pd.read_csv(p, encoding="utf-8-sig")
            except Exception:
                pass
    if st.session_state.categories_df is None:
        p = os.path.join(DATA_DIR, "categories.csv")
        if os.path.exists(p):
            try:
                st.session_state.categories_df = pd.read_csv(p, encoding="utf-8-sig")
            except Exception:
                pass

_init_state()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  CORE UTILITIES                                                 ║
# ╚══════════════════════════════════════════════════════════════════╝

def _find_header_row_index(raw: pd.DataFrame, salla_2row: bool) -> int:
    """يحدد صف رؤوس الأعمدة الحقيقي (سلة / مورد) من أول صفوف الملف."""
    markers_ar = (
        "أسم المنتج", "اسم المنتج", "نوع ", "no.", "no ", "رمز المنتج",
        "تصنيف المنتج", "صورة المنتج", "الماركة", "سعر المنتج", "الوصف",
    )
    markers_en = ("product name", "name", "sku", "price", "image", "title", "category")
    n = min(len(raw), 28)
    best_i = 1 if salla_2row else 0
    best_score = -1
    for i in range(n):
        row = raw.iloc[i]
        cells = [str(x).strip() for x in row.tolist() if pd.notna(x) and str(x).strip().lower() not in ("nan", "none", "")]
        if not cells:
            continue
        joined = " ".join(cells).lower().replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
        if len(cells) <= 2 and "بيانات المنتج" in joined and not any(m.lower() in joined for m in ("أسم", "اسم", "no", "sku")):
            continue
        sc = 0
        for m in markers_ar:
            ml = m.lower().replace("أ", "ا")
            if ml in joined or any(m in c or ml in c.lower() for c in cells):
                sc += 1
        for m in markers_en:
            if m in joined:
                sc += 1
        if sc > best_score:
            best_score = sc
            best_i = i
    if best_score >= 2:
        return best_i
    return 1 if salla_2row else 0


def read_file(f, salla_2row: bool = False) -> pd.DataFrame:
    """قراءة CSV أو Excel إلى DataFrame مع اكتشاف صف العناوين وإزالة الأعمدة الفارغة."""
    name = f.name.lower()
    PREVIEW = 45
    hdr_fallback = 1 if salla_2row else 0
    try:
        if name.endswith((".xlsx", ".xlsm", ".xls")):
            f.seek(0)
            raw_preview = pd.read_excel(f, header=None, dtype=str, nrows=PREVIEW)
            hdr_idx = _find_header_row_index(raw_preview, salla_2row)
            f.seek(0)
            df = pd.read_excel(f, skiprows=list(range(hdr_idx)), header=0, dtype=str)
        else:
            df = None
            last_err = None
            for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
                try:
                    f.seek(0)
                    raw_preview = pd.read_csv(f, header=None, encoding=enc, dtype=str, nrows=PREVIEW)
                    hdr_idx = _find_header_row_index(raw_preview, salla_2row)
                    f.seek(0)
                    df = pd.read_csv(f, skiprows=list(range(hdr_idx)), header=0, encoding=enc, dtype=str)
                    break
                except UnicodeDecodeError as e:
                    last_err = e
                    continue
            if df is None:
                APP_LOG.error("read_file CSV decode failed: %s", last_err)
                st.error(
                    "تعذّر فك ترميز ملف CSV. احفظ الملف بترميز UTF-8 من Excel "
                    "(تصدير CSV UTF-8) أو استخدم صيغة Excel. "
                    f"التفاصيل التقنية: {last_err}"
                )
                return pd.DataFrame()
        df = df.dropna(how="all").reset_index(drop=True)
        df.columns = [str(c).strip() for c in df.columns]
        drop_u = [c for c in df.columns if str(c).lower().startswith("unnamed")
                  and df[c].fillna("").astype(str).str.strip().replace("nan", "").eq("").all()]
        if drop_u:
            df = df.drop(columns=drop_u, errors="ignore")
        return df
    except Exception as e:
        try:
            f.seek(0)
            if name.endswith((".xlsx", ".xlsm", ".xls")):
                df = pd.read_excel(f, header=hdr_fallback, dtype=str)
            else:
                for enc in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
                    try:
                        f.seek(0)
                        df = pd.read_csv(f, header=hdr_fallback, encoding=enc, dtype=str)
                        break
                    except UnicodeDecodeError:
                        continue
            df = df.dropna(how="all").reset_index(drop=True)
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception as e2:
            APP_LOG.exception("read_file fallback failed: %s | %s", e, e2)
            st.error(
                "تعذّر قراءة الملف. تأكد أن الملف غير تالف، والامتداد يطابق المحتوى "
                "(.csv / .xlsx)، وأن الملف غير مفتوح في برنامج آخر. "
                f"التفاصيل: {e2}"
            )
            return pd.DataFrame()


def normalize_price_digits(val) -> str:
    s = str(val or "").strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    for i, ch in enumerate("٠١٢٣٤٥٦٧٨٩"):
        s = s.replace(ch, str(i))
    # فاصل عشري عربي (U+066B) وواصلة ألفية عربية
    s = s.replace("\u066b", ".").replace("٫", ".").replace("\u066c", "")
    m = re.search(r"(\d+(?:[.,]\d+)?)", s.replace(",", ""))
    if m:
        return m.group(1).replace(",", ".")
    digits = re.sub(r"[^\d.]", "", s)
    return digits or ""


def sanitize_salla_price_for_export(val) -> str:
    """إزالة عملات ورموز غير رقمية (SAR، ر.س، …) قبل تصدير سلة (Audit #20)."""
    s = str(val or "").strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    s = re.sub(r"(?i)\b(?:sar|sr|usd|eur|aed)\b", " ", s)
    s = re.sub(r"(?:ريال|ر\.?\s*س)", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[$€£¥]", "", s)
    return normalize_price_digits(s)


def compact_html_desc(html: str) -> str:
    if not html or not isinstance(html, str):
        return ""
    lines = [ln.strip() for ln in html.splitlines()]
    lines = [ln for ln in lines if ln]
    return "\n".join(lines)


def _is_unclear_column_name(col_name: str) -> bool:
    s = str(col_name).strip().lower()
    if not s:
        return True
    if s.startswith("unnamed"):
        return True
    if re.match(r"^column\.?\d+$", s):
        return True
    return False


def _infer_sample_role(series: pd.Series) -> str:
    vals = series.head(5).dropna().astype(str).str.strip()
    vals = [v for v in vals if v and v.lower() not in ("nan", "none", "")]
    if not vals:
        return "unknown"
    img_n = sum(1 for v in vals if v.startswith("http") or ("http" in v and "//" in v))
    pr_n = sum(1 for v in vals if re.search(r"\d", v) and len(v) < 48 and not v.startswith("http"))
    sku_n = sum(1 for v in vals if 2 <= len(v) <= 48 and re.match(r"^[\w\-\sA-Za-z٠-٩0-9]+$", v) and len(v.split()) <= 4)
    nm_n = sum(1 for v in vals if len(v) > 14 and (" " in v or "عطر" in v or "مل" in v or "bar" in v.lower()))
    scores = {"image": img_n, "price": pr_n, "sku": sku_n, "name": nm_n}
    best = max(scores, key=scores.get)
    if scores[best] >= max(2, len(vals) - 1):
        return best
    if scores[best] >= 1 and len(vals) == 1:
        return best
    return "unknown"


def _guess_role_from_keywords(keywords: list) -> str:
    for kw in keywords:
        kl = str(kw).lower()
        if any(x in kl for x in ("sku", "رمز", "barcode")):
            return "sku"
        if any(x in kl for x in ("صورة", "image", "src", "img", "w-full")):
            return "image"
        if any(x in kl for x in ("سعر", "price", "amount", "text-sm")):
            return "price"
        if any(x in kl for x in ("ماركة", "brand")):
            return "brand"
        if any(x in kl for x in ("وصف", "desc")):
            return "desc"
        if any(x in kl for x in ("اسم", "name", "منتج", "أسم", "product", "title")):
            return "name"
    return "name"


def auto_guess_col(cols, keywords: list, df: pd.DataFrame = None) -> str:
    col_list = [str(c) for c in cols]
    want = _guess_role_from_keywords(keywords)
    for kw in keywords:
        for c in col_list:
            if kw.lower() not in c.lower():
                continue
            if df is None or c not in df.columns:
                return c
            if _is_unclear_column_name(c):
                got = _infer_sample_role(df[c])
                if got not in ("unknown", want):
                    continue
            return c
    if df is not None and len(df) > 0:
        for c in col_list:
            if c not in df.columns:
                continue
            if _infer_sample_role(df[c]) == want:
                return c
    return "— لا يوجد —"


def _fuzzy_ratio(a: str, b: str) -> int:
    """Similarity ratio (0-100) — uses rapidfuzz when available for higher accuracy."""
    a, b = str(a).lower().strip(), str(b).lower().strip()
    if not a or not b:
        return 0
    if a == b:
        return 100
    if HAS_RAPIDFUZZ:
        return int(rf_fuzz.token_set_ratio(a, b))
    # Fallback: LCS-based ratio
    longer  = max(len(a), len(b))
    matches = 0
    j = 0
    for ch in a:
        while j < len(b):
            if b[j] == ch:
                matches += 1
                j += 1
                break
            j += 1
    return int(matches / longer * 100)


def _read_uploaded_file_as_df(uploaded_file) -> pd.DataFrame:
    """اقرأ ملف CSV/Excel باستخدام read_file مع محاولة اكتشاف ترتيب الصفوف."""
    if uploaded_file is None:
        return pd.DataFrame()
    df = read_file(uploaded_file, salla_2row=True)
    if df is None or df.empty:
        df = read_file(uploaded_file, salla_2row=False)
    return df if isinstance(df, pd.DataFrame) else pd.DataFrame()


def export_brands_csv_salla(df: pd.DataFrame) -> bytes:
    """تصدير ماركات سلة — CSV ترميز utf-8-sig مع ترتيب الأعمدة الصحيح."""
    if df is None or df.empty:
        # header only
        tmp = pd.DataFrame(columns=SALLA_BRANDS_COLS)
        df = tmp
    df2 = df.copy()
    # ضمان ترتيب الأعمدة
    for c in SALLA_BRANDS_COLS:
        if c not in df2.columns:
            df2[c] = ""
    df2 = df2[SALLA_BRANDS_COLS]
    out = io.StringIO()
    df2.to_csv(out, index=False, encoding="utf-8-sig")
    return out.getvalue().encode("utf-8-sig")


def filter_new_products_against_store(
    new_df: pd.DataFrame,
    store_df: pd.DataFrame,
    similarity_threshold: int = 90,
    use_sku_exact: bool = True,
    return_positions: bool = False,
) -> Union[pd.DataFrame, Tuple[pd.DataFrame, List[int]]]:
    """
    فلترة منتجات جديدة: حذف أي صف من new_df إذا وجد تطابق fuzzy في اسم المنتج داخل store_df.
    - المفتاح الأساسي: `أسم المنتج`
    - اختياري: drop إضافي عبر SKU exact match في `رمز المنتج sku`
    - إذا return_positions=True: يعيد (DataFrame, قائمة فهارس الصفوف المحفوظة من المدخل)
    """
    if new_df is None or new_df.empty:
        empty = pd.DataFrame(columns=new_df.columns if isinstance(new_df, pd.DataFrame) else [])
        return (empty, []) if return_positions else empty
    if store_df is None or store_df.empty:
        if return_positions:
            return new_df.copy(), list(range(len(new_df)))
        return new_df.copy()

    new_name_col = "أسم المنتج" if "أسم المنتج" in new_df.columns else auto_guess_col(
        new_df.columns, ["أسم المنتج", "اسم", "name", "منتج"], new_df
    )
    store_name_col = "أسم المنتج" if "أسم المنتج" in store_df.columns else auto_guess_col(
        store_df.columns, ["أسم المنتج", "اسم", "name", "منتج"], store_df
    )

    new_sku_col = "رمز المنتج sku" if "رمز المنتج sku" in new_df.columns else auto_guess_col(
        new_df.columns, ["رمز المنتج sku", "sku", "رمز", "barcode"], new_df
    )
    store_sku_col = "رمز المنتج sku" if "رمز المنتج sku" in store_df.columns else auto_guess_col(
        store_df.columns, ["رمز المنتج sku", "sku", "رمز", "barcode"], store_df
    )

    if new_name_col == "— لا يوجد —" or store_name_col == "— لا يوجد —":
        raise ValueError("لم يتم العثور على عمود 'أسم المنتج' في أحد الملفين.")

    store_names = (
        store_df.get(store_name_col, pd.Series(dtype="object"))
        .dropna()
        .astype(str)
        .map(lambda x: x.strip())
    )
    store_names = store_names[~store_names.str.lower().isin(["nan", "none", ""])]
    store_names_list = store_names.tolist()

    store_sku_set = set()
    if use_sku_exact and new_sku_col != "— لا يوجد —" and store_sku_col != "— لا يوجد —":
        store_skus = (
            store_df.get(store_sku_col, pd.Series(dtype="object"))
            .dropna()
            .astype(str)
            .map(lambda x: x.strip())
        )
        store_skus = store_skus[~store_skus.str.lower().isin(["nan", "none", ""])]
        store_sku_set = {s.lower() for s in store_skus.tolist() if s}

    keep_positions: list[int] = []
    for pos in range(len(new_df)):
        row = new_df.iloc[pos]
        name = str(row.get(new_name_col, "") or "").strip()
        if not name or name.lower() in ("nan", "none"):
            keep_positions.append(pos)
            continue

        sku = ""
        if use_sku_exact and new_sku_col != "— لا يوجد —":
            sku = str(row.get(new_sku_col, "") or "").strip()

        # SKU exact: drop مباشرة
        if store_sku_set and sku and sku.lower() in store_sku_set:
            continue

        # Fuzzy name: drop إذا وجد تطابق fuzzy
        exists = False
        for ex in store_names_list:
            if _fuzzy_ratio(name, ex) >= int(similarity_threshold):
                exists = True
                break
        if exists:
            continue

        keep_positions.append(pos)

    out_df = new_df.iloc[keep_positions].reset_index(drop=True)
    if return_positions:
        return out_df, keep_positions
    return out_df


def render_quick_add_tab():
    """تبويب منتج سريع: سحب من رابط أو إدخال يدوي + توليد/تجميع CSV سلة."""
    st.markdown("""<div class="al-info">
    أضف منتجات بسرعة عبر روابط أو إدخال يدوي. يتم استخراج البيانات ثم توليد ملفات سلة جاهزة.
    </div>""", unsafe_allow_html=True)

    qa_tab1, qa_tab2 = st.tabs(["🔗 سحب من رابط", "📝 إدخال يدوي ورفع صور"])

    # تابعنا brands_file لتطبيق شرط fuzzy < 85 كـ Missing Brand
    brands_upload = st.file_uploader(
        "ماركات مهووس.csv (للتأكد من وجود الماركة)",
        type=["csv", "xlsx", "xls"],
        key="qa_brands_uploader",
        label_visibility="collapsed",
    )

    brands_df = _read_uploaded_file_as_df(brands_upload) if brands_upload is not None else st.session_state.get(
        "brands_df", None
    )
    if brands_df is None or brands_df.empty:
        brands_df = st.session_state.brands_df if hasattr(st.session_state, "brands_df") else None
    brand_name_col = "اسم الماركة" if brands_df is not None and "اسم الماركة" in brands_df.columns else (
        brands_df.columns[0] if brands_df is not None and not brands_df.empty else None
    )

    def _cross_reference_brand(brand_candidate: str) -> tuple[str, str, int]:
        """
        returns: (brand_for_product, brand_for_missing, best_score)
        إذا best_score >= 85: brand_for_product = الاسم الموجود من brands_df
        else: brand_for_product = الاسم المرشح، وbrand_for_missing = الاسم المرشح
        """
        if not brand_candidate:
            return "", "", 0
        if brands_df is None or brands_df.empty or not brand_name_col:
            return brand_candidate.strip(), brand_candidate.strip(), 0
        cand = str(brand_candidate).strip()
        best_name = cand
        best_score = 0
        for bn in (
            brands_df[brand_name_col]
            .dropna()
            .astype(str)
            .map(lambda x: x.strip())
            .tolist()
        ):
            if not bn:
                continue
            sc = _fuzzy_ratio(cand, bn)
            if sc > best_score:
                best_score = sc
                best_name = bn
        if best_score >= 85:
            return best_name, "", best_score
        return cand, cand, best_score

    def _init_outputs():
        if "quickadd_new_products_df" not in st.session_state or st.session_state.quickadd_new_products_df is None:
            st.session_state.quickadd_new_products_df = pd.DataFrame(columns=SALLA_COLS)
        if "quickadd_missing_brands_df" not in st.session_state or st.session_state.quickadd_missing_brands_df is None:
            st.session_state.quickadd_missing_brands_df = pd.DataFrame(columns=SALLA_BRANDS_COLS)

    _init_outputs()

    # ===================== URL MODE =====================
    with qa_tab1:
        st.markdown("### سحب منتجات من روابط")
        urls_text = st.text_area(
            "ألصق الروابط هنا (سطر لكل رابط)",
            placeholder="https://example.com/product/1\nhttps://example.com/product/2",
            height=140,
            key="qa_urls_text",
        )
        start_btn = st.button("Start Processing (بدء المعالجة)", type="primary", use_container_width=True, key="qa_start_urls")

        if start_btn:
            api_key = st.session_state.get("api_key", "")
            if not api_key or not HAS_ANTHROPIC:
                st.error("أضف مفتاح Claude API (Anthropic) من صفحة الإعدادات.")
                st.stop()
            urls = [u.strip() for u in (urls_text or "").splitlines() if u.strip()]
            if not urls:
                st.error("أدخل رابطاً واحداً على الأقل.")
                st.stop()

            progress_bar = st.progress(0, text="جاري المعالجة...")
            new_rows = []
            missing_brand_names: list[str] = []

            for idx, u in enumerate(urls):
                progress_bar.progress(int((idx / max(len(urls), 1)) * 100), text=f"معالجة: {u[:55]}...")
                with st.spinner(f"استخراج: {u[:60]}..."):
                    extracted = extract_product_json_from_url(u, api_key=api_key)
                p_name = extracted.get("أسم المنتج", "") or ""
                p_brand = extracted.get("الماركة", "") or ""
                p_price = extracted.get("سعر المنتج", "") or ""
                p_desc = extracted.get("الوصف", "") or ""
                p_img = extracted.get("صورة المنتج", "") or ""

                if not p_name.strip() or p_name.lower().strip() in ("nan", "none"):
                    continue

                brand_for_prod, brand_for_missing, _best_sc = _cross_reference_brand(p_brand)

                if brand_for_missing:
                    missing_brand_names.append(brand_for_missing)

                # بناء صف سلة مطابق 42 عمود
                nr = fill_row(
                    name=p_name,
                    price=p_price,
                    sku="",
                    image=p_img,
                    desc=p_desc,
                    brand={"name": brand_for_prod},
                    category="",
                    seo={"alt": ""},
                    weight="0.2",
                )
                new_rows.append(nr)

            progress_bar.progress(100, text="اكتمل الاستخراج!")

            # تحديث session_state
            df_new = pd.DataFrame(new_rows, columns=SALLA_COLS) if new_rows else pd.DataFrame(columns=SALLA_COLS)
            st.session_state.quickadd_new_products_df = df_new

            # Missing brands (7 أعمدة)
            mb_seen: set[str] = set()
            mb_rows: list[dict] = []
            for bn in missing_brand_names:
                bn = str(bn).strip()
                if not bn:
                    continue
                bl = bn.lower()
                if bl in mb_seen:
                    continue
                mb_seen.add(bl)
                mb_rows.append({
                    "اسم الماركة": bn,
                    "وصف مختصر عن الماركة": "",
                    "صورة شعار الماركة": "",
                    "(إختياري) صورة البانر": "",
                    "(Page Title) عنوان صفحة العلامة التجارية": "",
                    "(SEO Page URL) رابط صفحة العلامة التجارية": "",
                    "(Page Description) وصف صفحة العلامة التجارية": "",
                })
            st.session_state.quickadd_missing_brands_df = pd.DataFrame(mb_rows, columns=SALLA_BRANDS_COLS)

            st.success(f"✅ تم استخراج {len(df_new):,} منتج — ماركات مفقودة: {len(st.session_state.quickadd_missing_brands_df):,}")

            st.rerun()

    # ===================== MANUAL MODE =====================
    with qa_tab2:
        st.markdown("### إدخال يدوي سريع")
        with st.form("qa_manual_form", clear_on_submit=False):
            m_name = st.text_input("أسم المنتج", key="qa_m_name")
            m_price = st.text_input("سعر المنتج", key="qa_m_price")
            m_brand = st.text_input("الماركة", key="qa_m_brand")
            m_img_url = st.text_input("صورة المنتج (رابط URL)", key="qa_m_img_url", placeholder="https://... (اختياري)")
            m_img_files = st.file_uploader(
                "ارفع صور المنتج (اختياري للتجهيز/المعاينة — يلزم رابط URL للتصدير للسلة)",
                type=["png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
                key="qa_m_img_files",
            )
            m_desc = st.text_area("الوصف (HTML بسيط)", key="qa_m_desc", height=90, placeholder="<p>...</p>")
            gen_desc = st.checkbox("🤖 توليد وصف AI (اختياري)", value=False, key="qa_m_gen_desc")
            manual_submit = st.form_submit_button("➕ إضافة للقائمة وتجهيز الملف", type="primary", use_container_width=True)

        if manual_submit:
            if not m_name.strip() or not m_price.strip():
                st.error("الاسم والسعر حقول إجبارية.")
                st.stop()
            if m_img_files and not (m_img_url or "").strip():
                st.warning("تم رفع صور محلياً. التصدير للسلة يحتاج رابط URL في حقل 'صورة المنتج'.")
            api_key = st.session_state.get("api_key", "")
            if gen_desc and (not api_key or not HAS_ANTHROPIC):
                st.error("تحتاج مفتاح Claude API لتوليد الوصف.")
                st.stop()

            brand_for_prod, brand_for_missing, _best_sc = _cross_reference_brand(m_brand)
            if brand_for_missing:
                # append missing brands
                df_mb = st.session_state.quickadd_missing_brands_df
                if df_mb is None or df_mb.empty:
                    df_mb = pd.DataFrame(columns=SALLA_BRANDS_COLS)
                if brand_for_missing.lower().strip() not in df_mb.get("اسم الماركة", pd.Series(dtype="object")).astype(str).str.lower().tolist():
                    st.session_state.quickadd_missing_brands_df = pd.concat(
                        [df_mb, pd.DataFrame([{
                            "اسم الماركة": brand_for_missing,
                            "وصف مختصر عن الماركة": "",
                            "صورة شعار الماركة": "",
                            "(إختياري) صورة البانر": "",
                            "(Page Title) عنوان صفحة العلامة التجارية": "",
                            "(SEO Page URL) رابط صفحة العلامة التجارية": "",
                            "(Page Description) وصف صفحة العلامة التجارية": "",
                        }])],
                        ignore_index=True,
                    )

            desc_final = m_desc
            if gen_desc:
                # وصف سريع بدون مكوّنات إضافية
                # نستخدم ai_generate لأنه يعيد HTML مقبول للـ Salla
                is_t = "تستر" in m_name.lower()
                size = "100 مل"
                gender = "للجنسين"
                conc_ar = "أو دو بارفيوم"
                desc_final = ai_generate(m_name, is_t, {"name": brand_for_prod, "page_url": ""}, size, gender, conc_ar)

            nr = fill_row(
                name=m_name,
                price=m_price,
                sku="",
                image=m_img_url,
                desc=desc_final,
                brand={"name": brand_for_prod},
                category="",
                seo={"alt": ""},
                weight="0.2",
            )
            # تحديث df_new (إلحاق)
            df_new = st.session_state.quickadd_new_products_df
            if df_new is None or df_new.empty:
                st.session_state.quickadd_new_products_df = pd.DataFrame([nr], columns=SALLA_COLS)
            else:
                st.session_state.quickadd_new_products_df = pd.concat(
                    [df_new, pd.DataFrame([nr], columns=SALLA_COLS)],
                    ignore_index=True,
                )
            st.success("✅ تم الإضافة")
            st.rerun()

    # ===================== EXPORTS =====================
    if st.session_state.get("quickadd_new_products_df") is not None and not st.session_state.quickadd_new_products_df.empty:
        st.divider()
        st.markdown("""<div class="sec-title"><div class="bar"></div><h3>تنزيل الملفات</h3></div>""",
                    unsafe_allow_html=True)
        df_prod = st.session_state.quickadd_new_products_df
        df_br = st.session_state.quickadd_missing_brands_df
        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                "📥 منتج_سريع_جديد.csv",
                export_product_csv(df_prod),
                "منتج_سريع_جديد.csv",
                "text/csv",
                use_container_width=True,
                key="qa_dl_new_products",
            )
        with c2:
            st.download_button(
                "📥 ماركات_جديدة_مضافة.csv",
                export_brands_csv_salla(df_br),
                "ماركات_جديدة_مضافة.csv",
                "text/csv",
                use_container_width=True,
                key="qa_dl_missing_brands",
            )


# ══════════════════════════════════════════════════════════════════
#  المحرك الذكي v12.0 — Cluster Matching Engine | صفر أخطاء
#  قانون الأكواد الصارم | مهووس | تم اختباره: 100% دقة (16/16)
# ══════════════════════════════════════════════════════════════════

# ── قواميس التطبيع ────────────────────────────────────────────────

CONCENTRATION_PATTERNS = [
    # PARFUM/EXTRAIT (الأطول أولاً)
    (r"extrait\s*de?\s*parfum", "PARFUM"),
    (r"pure\s*parfum", "PARFUM"),
    (r"بيور\s*بارفيوم", "PARFUM"),
    (r"اكستريت\s*دي?\s*بارفيوم", "PARFUM"),
    (r"اكستريكت\s*دي?\s*بارفيوم", "PARFUM"),
    (r"\bextrait\b", "PARFUM"),
    # EDP — يجب أن يأتي قبل PARFUM المفرد
    (r"eau\s*de?\s*parfum", "EDP"),
    (r"\bedp\b", "EDP"),
    (r"(?:او|أو|اودو|او\s*دو|أو\s*دو|او\s*دي|أو\s*دي|ادو)\s*(?:برفيوم|بارفيوم|بارفان|برفان|بارفوم|برفوم)", "EDP"),
    (r"دو\s*(?:برفيوم|بارفيوم|بارفان|برفان)", "EDP"),
    (r"دي\s*(?:برفيوم|بارفيوم|بارفان|برفان)", "EDP"),
    (r"لو\s*دي?\s*(?:بارفيوم|برفيوم|بارفان)", "EDP"),
    (r"اليكسير\s*دي?\s*(?:بارفيوم|برفيوم)", "EDP"),
    (r"انتنس\s*(?:دي?\s*)?(?:بارفيوم|برفيوم)", "EDP"),
    # PARFUM المفرد (بعد EDP)
    (r"\bparfum\b(?!\s*de)", "PARFUM"),
    (r"\bبارفيوم\b", "PARFUM"),
    (r"\bبرفيوم\b", "PARFUM"),
    (r"\bبارفان\b", "PARFUM"),
    (r"\bبرفان\b", "PARFUM"),
    (r"\bبارفوم\b", "PARFUM"),
    (r"\bبرفوم\b", "PARFUM"),
    (r"\bپارفوم\b", "PARFUM"),
    # EDT
    (r"eau\s*de?\s*toilette", "EDT"),
    (r"\bedt\b", "EDT"),
    (r"(?:او|أو|اودو|او\s*دو|أو\s*دو|او\s*دي|أو\s*دي|ادو)\s*(?:تواليت|تواليتي|تواليه)", "EDT"),
    (r"دو\s*تواليت", "EDT"),
    # EDC
    (r"eau\s*de?\s*cologne", "EDC"),
    (r"\bedc\b", "EDC"),
    (r"(?:او|أو)\s*(?:دو|دي)\s*كولون", "EDC"),
    (r"كولونيا", "EDC"),
    # MIST
    (r"hair\s*mist", "HAIR_MIST"),
    (r"هير\s*ميست", "HAIR_MIST"),
    (r"بخاخ\s*شعر", "HAIR_MIST"),
    (r"body\s*mist", "MIST"),
    (r"بودي\s*ميست", "MIST"),
    (r"\bميست\b", "MIST"),
]

TYPE_PATTERNS = [
    (r"\bتستر\b|\bتيستر\b|\btester\b|\btest\b", "TESTER"),
    (r"طقم|سيت|\bset\b|مجموعة\s*هدايا|بكج|\bpack\b|كوليكشن|\bcollection\b|هدية", "SET"),
    (r"زيت\s*عطر|\boil\b", "OIL"),
    (r"بودي\s*واش|شاور\s*جل|دوش\s*جل", "BODY_WASH"),
    (r"بودي\s*لوشن|لوشن\s*جسم", "LOTION"),
    (r"كريم\s*جسم|\bcream\b", "CREAM"),
    (r"معطر\s*جسم|بودي\s*سبراي|body\s*spray", "BODY_SPRAY"),
    (r"ديودورانت|مزيل\s*عرق|deodorant", "DEODORANT"),
]

SAMPLE_PATTERNS = [
    r"\bعينة\b", r"\bسمبل\b", r"\bsample\b", r"\bvial\b",
    r"\bminiature\b", r"\bميني\b", r"\bmini\b",
]

ARABIC_SPELLING = [
    (r"[إأآ]", "ا"),
    (r"[يى](?=\s|$|[^ا-ي])", "ي"),
    (r"ة(?=\s|$)", "ه"),
    (r"ؤ", "و"),
    (r"ئ", "ي"),
    (r"(?:او|أو|اودو|اودي)\s*(?:دو|دي|de)\s*", "eau de "),
    (r"(?:او|أو)\s*(?:دو|دي)\s*", "eau de "),
    (r"\blou\s*de\b", "eau de"),
    (r"\bلو\s*دي?\b", "eau de"),
    (r"اليكسير|إليكسير|اكسير|إكسير|اليكزير|اليكسر", "اليكسير"),
    (r"ريزيرف|ريزيرفي|ريزيرفه", "ريزيرف"),
    (r"انتنس|انتنز|انتانس|انتانز|انتانس|انتينس", "انتنس"),
    (r"جنتلمان|جنتلمن", "جنتلمان"),
    (r"بلاك|بلك", "بلاك"),
    (r"وايت|وهايت", "وايت"),
    (r"جولد|قولد", "جولد"),
    (r"رويال|رويل", "رويال"),
    (r"ليجند|ليجيند", "ليجند"),
    (r"اكستريكت|اكستريت", "اكستريت"),
    (r"برفيوم|بارفيوم|بارفان|برفان|بارفوم|برفوم", "بارفيوم"),
    (r"\bدي\b", "دو"),
    (r"ايست|إيست|إيس|ايس", "اي"),
    (r"إي(?=\s|$)", "اي"),
    (r"بيلل", "بيل"),
    (r"لا\s*في", "لافي"),
    (r"سوفاجه", "سوفاج"),
    (r"شانيلل", "شانيل"),
    (r"\bلو\b(?!\s*(?:دي?|de))", "له"),
]

_CATEGORY_MAP_V12 = {
    "TESTER":     "العطور > تستر",
    "SET":        "العطور > طقم هدايا",
    "HAIR_MIST":  "العطور > عطور الشعر",
    "LOTION":     "العناية > لوشن وكريم",
    "BODY_WASH":  "العناية > شاور جل",
    "DEODORANT":  "العناية > مزيل العرق",
    "BODY_SPRAY": "العطور > معطر جسم",
    "PERFUME":    "العطور",
}


def _normalize_text_v12(text: str) -> str:
    if not text or not isinstance(text, str):
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFKC", text)
    for i, n in enumerate("٠١٢٣٤٥٦٧٨٩"):
        text = text.replace(n, str(i))
    for pattern, repl in ARABIC_SPELLING:
        text = re.sub(pattern, repl, text, flags=re.IGNORECASE)
    text = re.sub(r"[^\w\s\d]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def brand_exists_in_brands_df(brand_name: str, bdf: Optional[pd.DataFrame]) -> bool:
    """يتحقق إن كانت الماركة موجودة في DataFrame ماركات (قابل للاختبار بدون Streamlit)."""
    if not str(brand_name).strip():
        return False
    if bdf is None or bdf.empty:
        return False
    col0 = bdf.columns[0]
    bn_full = str(brand_name).strip().lower()
    parts_in = [p.strip().lower() for p in re.split(r"\s*\|\s*", str(brand_name)) if p.strip()]
    bn_norm = _normalize_text_v12(str(brand_name).split("|")[0].strip())
    keys = {_normalize_text_v12(p) for p in parts_in}
    keys.add(bn_norm)
    for _, row in bdf.iterrows():
        raw = str(row[col0])
        for part in re.split(r"\s*\|\s*", raw):
            pl = part.strip().lower()
            if not pl:
                continue
            pk = _normalize_text_v12(part)
            if pk in keys:
                return True
            if len(pl) >= 3 and (bn_full in pl or pl in bn_full):
                return True
            if len(pk) >= 3 and len(bn_norm) >= 3 and (pk in bn_norm or bn_norm in pk):
                return True
    # مقارنة موحّدة (أ/إ/آ→ا، ة→ه) لتفادي تكرار «أكوا» و«اكوا»
    bn_v2 = normalize_brand_name_v2(str(brand_name).split("|")[0].strip())
    if len(bn_v2) >= 2:
        for _, row in bdf.iterrows():
            raw = str(row[col0])
            for part in re.split(r"\s*\|\s*", raw):
                if part.strip() and normalize_brand_name_v2(part) == bn_v2:
                    return True
    return False


def normalize_brand_name_v2(s: str) -> str:
    """توحيد اسم العلامة للمقارنة: ألف/همزات → ا، ة → ه، مسافات."""
    if not s:
        return ""
    t = unicodedata.normalize("NFKC", str(s).strip())
    for a in ("أ", "إ", "آ", "ٱ"):
        t = t.replace(a, "ا")
    t = t.replace("ة", "ه")
    t = re.sub(r"\s+", " ", t).strip()
    return t.lower()


def _brand_in_session_runtime(brand_name: str) -> bool:
    """هل الماركة مُسجّلة في ذاكرة الجلسة الحالية (بعد إثراء أو توليد)."""
    bn = normalize_brand_name_v2(brand_name)
    if not bn or len(bn) < 2:
        return False
    for b in st.session_state.get("pipe_session_brands", []):
        raw = str(b.get("name", "") or "")
        if normalize_brand_name_v2(raw) == bn:
            return True
        for part in re.split(r"\s*\|\s*", raw):
            if part.strip() and normalize_brand_name_v2(part) == bn:
                return True
    return False


def register_pipe_session_brand(display_name: str, page_url: str = "") -> None:
    """يُحدّث قائمة ماركات الجلسة حتى يتعرّف match_brand على العلامة في نفس التشغيل."""
    if not str(display_name or "").strip():
        return
    low = str(display_name).strip().lower()
    if low in ("unknown", "غير محدد", "غير معروف"):
        return
    lst = st.session_state.setdefault("pipe_session_brands", [])
    norm = normalize_brand_name_v2(display_name)
    for x in lst:
        if normalize_brand_name_v2(x.get("name", "")) == norm:
            return
    lst.append({"name": str(display_name).strip(), "page_url": str(page_url or "").strip()})


def brand_exists_in_catalog(brand_name: str) -> bool:
    """يتحقق إن كانت الماركة (أو أحد أجزائها بعد |) موجودة في ملف ماركات المتجر أو جلسة المسار."""
    if brand_exists_in_brands_df(brand_name, st.session_state.brands_df):
        return True
    return _brand_in_session_runtime(brand_name)


def dedupe_products_df(df: pd.DataFrame) -> pd.DataFrame:
    """إزالة صفوف مكررة حسب SKU ثم اسم المنتج الموحّد."""
    if df is None or df.empty or "أسم المنتج" not in df.columns:
        return df
    sku_col = "رمز المنتج sku"
    name_col = "أسم المنتج"

    def _row_key(r) -> str:
        sku = str(r.get(sku_col, "") or "").strip().lower()
        if sku and sku not in ("nan", "none", ""):
            return "sku:" + sku
        return _normalize_text_v12(str(r.get(name_col, "") or ""))

    d2 = df.copy()
    d2["_dk"] = d2.apply(_row_key, axis=1)
    d2 = d2.drop_duplicates(subset=["_dk"], keep="first")
    d2 = d2.drop(columns=["_dk"], errors="ignore").reset_index(drop=True)
    if "No." in d2.columns:
        d2["No."] = [str(i + 1) for i in range(len(d2))]
    return d2


def dedupe_final_rows_and_seo(final_rows: list, seo_rows: list) -> tuple:
    """مزامنة إزالة التكرار بين صفوف المنتج وصفوف SEO وإعادة ترقيم No."""
    if not final_rows:
        return final_rows, seo_rows
    seen = set()
    fr2, sr2 = [], []
    for r, s in zip(final_rows, seo_rows):
        sku = str(r.get("رمز المنتج sku", "") or "").strip().lower()
        if sku and sku not in ("nan", "none", ""):
            k = "sku:" + sku
        else:
            k = _normalize_text_v12(str(r.get("أسم المنتج", "") or ""))
        if not k or k == "sku:":
            continue
        if k in seen:
            continue
        seen.add(k)
        fr2.append(r)
        sr2.append(s)
    for i, r in enumerate(fr2):
        r["No."] = str(i + 1)
    for i, s in enumerate(sr2):
        s["No. (غير قابل للتعديل)"] = str(i + 1)
    return fr2, sr2


def dedupe_brand_entries(brands: list) -> list:
    seen = set()
    out = []
    for b in brands:
        nm = b.get("اسم الماركة") or b.get("name") or ""
        key = _normalize_text_v12(str(nm).split("|")[0].strip())
        if not key:
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(b)
    return out


def _extract_size_v12(text: str) -> float:
    text_lower = text.lower()
    patterns = [
        (r"(\d+(?:[.,]\d+)?)\s*(?:مل|ml|مليلتر|milliliter|millilitre)", 1.0),
        (r"(\d+(?:[.,]\d+)?)\s*(?:لتر|liter|litre)\b", 1000.0),
        (r"(\d+(?:[.,]\d+)?)\s*(?:oz|أوقية|اونصة)", 29.5735),
    ]
    for pattern, mult in patterns:
        m = re.search(pattern, text_lower, re.IGNORECASE)
        if m:
            return round(float(m.group(1).replace(",", ".")) * mult, 1)
    return 0.0


def _extract_concentration_v12(text: str) -> str:
    text_lower = text.lower()
    for pattern, conc in CONCENTRATION_PATTERNS:
        if re.search(pattern, text_lower, re.IGNORECASE):
            return conc
    return "UNKNOWN"


def _extract_type_v12(text: str) -> str:
    text_lower = text.lower()
    for pattern, ptype in TYPE_PATTERNS:
        if re.search(pattern, text_lower, re.IGNORECASE):
            return ptype
    return "PERFUME"


def _is_sample_v12(text: str, size: float) -> bool:
    if 0 < size <= 8:
        return True
    text_lower = text.lower()
    for p in SAMPLE_PATTERNS:
        if re.search(p, text_lower, re.IGNORECASE):
            return True
    return False


def _normalize_brand_v12(brand: str) -> str:
    if not brand:
        return ""
    parts = re.split(r"[|/]", brand)
    return _normalize_text_v12(parts[0].strip())


# أنماط خصم/نسبة مئوية في نهاية اسم المنتج (تُخطأ أحياناً كعمود ماركة)
_DISCOUNT_TAIL_RE = re.compile(
    r"(?:\s*[-–—]\s*\d+(?:[.,]\d+)?\s*%|\s+\d+(?:[.,]\d+)?\s*%)\s*$",
    re.IGNORECASE,
)


def strip_trailing_discount_label(text: str) -> str:
    """إزالة لاحقة خصم من نهاية الاسم (مثل « - 12% » أو « 15% »)."""
    if not text:
        return text
    s = str(text).strip()
    for _ in range(4):
        ns = _DISCOUNT_TAIL_RE.sub("", s).strip().strip("-–—").strip()
        if ns == s:
            break
        s = ns
    return s


def is_discount_like_brand(s: str) -> bool:
    """نص يشبه عمود خصم وليس ماركة (مثل - 12% أو 5%)."""
    if not s or not str(s).strip():
        return True
    t = str(s).strip().replace("٪", "%")
    if re.fullmatch(r"[\s\-–—]*\d+(?:[.,]\d+)?\s*%\s*", t, re.IGNORECASE):
        return True
    if re.fullmatch(r"[\s\-–—]+\d+(?:[.,]\d+)?\s*", t):
        return True
    parts = t.split()
    if parts and all(
        re.fullmatch(r"[\-–—]+", p)
        or re.fullmatch(r"\d+(?:[.,]\d+)?%", p.replace("٪", "%"), re.IGNORECASE)
        for p in parts
    ):
        return True
    return False


def _extract_core_name_v12(raw_name: str, brand: str = "") -> str:
    """الاسم الجوهري = الاسم بعد حذف الحجم والتركيز والنوع والماركة."""
    text = raw_name
    text = re.sub(r"^(عطر|تستر|تيستر|كريم|لوشن|بودي|زيت|معطر|مزيل)\s+", "", text.strip(), flags=re.IGNORECASE)
    if brand:
        for part in re.split(r"[|/]", brand):
            part = part.strip()
            if len(part) > 2:
                text = re.sub(re.escape(part), " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\d+(?:[.,]\d+)?\s*(?:مل|ml|مليلتر|لتر|liter|litre|oz|أوقية)", " ", text, flags=re.IGNORECASE)
    for pattern, _ in CONCENTRATION_PATTERNS:
        text = re.sub(pattern, " ", text, flags=re.IGNORECASE)
    for pattern, _ in TYPE_PATTERNS:
        text = re.sub(pattern, " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(للرجال|للنساء|للجنسين|men|women|unisex|رجالي|نسائي)\b", " ", text, flags=re.IGNORECASE)
    text = _normalize_text_v12(text)
    text = re.sub(r"\b\d+\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


@dataclass
class _ProductRecord_v12:
    raw_name: str
    brand: str = ""
    size: float = 0.0
    concentration: str = "UNKNOWN"
    product_type: str = "PERFUME"
    core_name: str = ""
    is_sample_flag: bool = False
    brand_normalized: str = ""

    def __post_init__(self):
        self.size = _extract_size_v12(self.raw_name)
        self.concentration = _extract_concentration_v12(self.raw_name)
        self.product_type = _extract_type_v12(self.raw_name)
        self.is_sample_flag = _is_sample_v12(self.raw_name, self.size)
        self.core_name = _extract_core_name_v12(self.raw_name, self.brand)
        self.brand_normalized = _normalize_brand_v12(self.brand)


class _ClusterMatchEngine_v12:
    """المحرك الذكي v12.0 — المقارنة العنقودية بصفر أخطاء."""

    def __init__(self, store_records: List[Dict],
                 name_col: str = "name", brand_col: str = "brand"):
        self.store_products: List[_ProductRecord_v12] = []
        self.cluster: Dict[str, List[_ProductRecord_v12]] = {}
        self._build(store_records, name_col, brand_col)

    def _build(self, records, name_col, brand_col):
        for rec in records:
            name = str(rec.get(name_col, "")).strip()
            brand = str(rec.get(brand_col, "")).strip()
            if not name or name.lower() in ("nan", "none", ""):
                continue
            prod = _ProductRecord_v12(raw_name=name, brand=brand)
            if prod.is_sample_flag:
                continue
            self.store_products.append(prod)
            key = prod.brand_normalized or "__no_brand__"
            self.cluster.setdefault(key, []).append(prod)

    @staticmethod
    def _name_sim(a: str, b: str) -> float:
        if not a or not b:
            return 0.0
        if len(a) < 2 or len(b) < 2:
            return 0.0
        try:
            from rapidfuzz import fuzz
            return fuzz.token_sort_ratio(a, b)
        except ImportError:
            a_w, b_w = set(a.split()), set(b.split())
            if not a_w or not b_w:
                return 0.0
            return len(a_w & b_w) / len(a_w | b_w) * 100

    def _check_pair(self, new_p: _ProductRecord_v12, store_p: _ProductRecord_v12):
        nb, sb = new_p.brand_normalized, store_p.brand_normalized
        if nb and sb and nb != sb:
            if nb not in sb and sb not in nb:
                return False, f"ماركة مختلفة: [{nb}] vs [{sb}]", 0.0
        if new_p.product_type != store_p.product_type:
            return False, f"نوع مختلف: {new_p.product_type} vs {store_p.product_type}", 0.0
        if new_p.size > 0 and store_p.size > 0:
            if abs(new_p.size - store_p.size) > 0.5:
                return False, f"حجم مختلف: {new_p.size} vs {store_p.size}", 0.0
        if (new_p.concentration != "UNKNOWN" and store_p.concentration != "UNKNOWN"
                and new_p.concentration != store_p.concentration):
            return False, f"تركيز مختلف: {new_p.concentration} vs {store_p.concentration}", 0.0
        score = self._name_sim(new_p.core_name, store_p.core_name)
        return True, "مؤهل", score

    def match(self, competitor_name: str, competitor_brand: str = "",
              t_dup: float = 90.0, t_critical: float = 72.0) -> dict:
        new_p = _ProductRecord_v12(raw_name=competitor_name, brand=competitor_brand)
        if new_p.is_sample_flag:
            return {"verdict": "مستبعد", "reason": "عينة صغيرة",
                    "score": 0.0, "matched_name": None, "product": new_p}

        brand_key = new_p.brand_normalized or "__no_brand__"
        candidates = self.cluster.get(brand_key, [])
        if not candidates:
            for k, v in self.cluster.items():
                if brand_key in k or k in brand_key:
                    candidates.extend(v)
        if not candidates:
            candidates = self.store_products

        best_score = 0.0
        best_match: Optional[_ProductRecord_v12] = None
        rejection_reasons = []

        for store_p in candidates:
            can, reason, score = self._check_pair(new_p, store_p)
            if not can:
                rejection_reasons.append(reason)
                continue
            if score > best_score:
                best_score = score
                best_match = store_p

        if best_score >= t_dup:
            verdict = "مكرر"
            reason = f"تطابق ({best_score:.1f}%) — {best_match.raw_name[:60]}"
        elif best_score >= t_critical:
            verdict = "حرج"
            reason = f"تشابه حرج ({best_score:.1f}%) — {best_match.raw_name[:50] if best_match else '—'}"
        elif best_score > 0:
            verdict = "جديد"
            reason = f"أقرب تشابه ({best_score:.1f}%) — غير كافٍ"
        else:
            verdict = "جديد"
            uniq = list(dict.fromkeys(rejection_reasons[:3]))
            reason = "جديد — " + " | ".join(uniq[:2]) if uniq else "جديد — لا يوجد في متجرنا"

        return {"verdict": verdict, "reason": reason, "score": best_score,
                "matched_name": best_match.raw_name if best_match else None, "product": new_p}


# ── دوال التوافق مع الكود القديم ────────────────────────────────

def standardize_product_name(raw_name: str, brand_name: str) -> str:
    """
    إعادة صياغة اسم المنتج بالترتيب المعتمد:
    (عطر/تستر) + (الاسم الأساسي) + (الماركة بالعربية) + (التركيز) + (الحجم)
    """
    p = _ProductRecord_v12(raw_name=raw_name, brand=brand_name)

    type_prefixes = {
        "TESTER":     "تستر",  "SET":       "طقم",       "HAIR_MIST": "عطر شعر",
        "LOTION":     "لوشن",  "BODY_WASH": "شاور جل",   "CREAM":     "كريم",
        "BODY_SPRAY": "معطر جسم", "DEODORANT": "مزيل عرق",
        "MIST":       "معطر",  "PERFUME":   "عطر",       "OIL":       "زيت عطري",
    }
    prefix = type_prefixes.get(p.product_type, "عطر")
    if "تستر" in raw_name.lower() or "tester" in raw_name.lower():
        prefix = "تستر"

    core = p.core_name
    b_ar = brand_name.split("|")[0].strip() if brand_name else p.brand_normalized
    if b_ar and b_ar in core:
        core = core.replace(b_ar, "").strip()

    conc_map = {
        "EDP": "أو دو بارفيوم", "EDT": "أو دو تواليت",
        "EDC": "أو دو كولون",   "PARFUM": "بارفيوم",
    }
    conc_str = conc_map.get(p.concentration, "")
    if not conc_str:
        rn = raw_name.lower()
        if any(k in rn for k in ["بارفيوم","parfum","edp"]):   conc_str = "أو دو بارفيوم"
        elif any(k in rn for k in ["تواليت","toilette","edt"]): conc_str = "أو دو تواليت"
        elif any(k in rn for k in ["كولون","cologne","edc"]):   conc_str = "أو دو كولون"

    if p.size > 0:
        size_str = f"{int(p.size) if p.size == int(p.size) else p.size} مل"
    else:
        size_str = ""

    parts = [x for x in [prefix, core, b_ar, conc_str, size_str] if x]
    return re.sub(r"\s+", " ", " ".join(parts)).strip()


def extract_product_attrs(name: str) -> dict:
    """استخراج الحجم، النوع، التركيز، الاسم الجوهري — v12.0 (صفر أخطاء)."""
    p = _ProductRecord_v12(raw_name=name, brand="")
    ptype_map = {
        "TESTER": "تستر", "SET": "طقم هدايا", "OIL": "زيت جسم",
        "BODY_WASH": "شاور جل", "LOTION": "عناية جسم", "CREAM": "عناية جسم",
        "BODY_SPRAY": "معطر جسم", "DEODORANT": "مزيل عرق",
        "HAIR_MIST": "عطر شعر", "MIST": "معطر جسم", "PERFUME": "عطر تجاري",
    }
    ptype_ar = ptype_map.get(p.product_type, "عطر تجاري")
    if p.is_sample_flag:
        ptype_ar = "عينة (مستبعدة)"
    conc_map = {
        "EDP": "EDP", "EDT": "EDT", "EDC": "EDC",
        "PARFUM": "Parfum", "HAIR_MIST": "Hair Mist",
        "MIST": "Body Mist", "UNKNOWN": "غير محدد",
    }
    conc_ar = conc_map.get(p.concentration, p.concentration)
    return {
        "size": p.size,
        "type": ptype_ar,
        "concentration": conc_ar,
        "clean_name": p.core_name,
        "core_name": p.core_name,
        "brand": p.brand_normalized,
        "category": _CATEGORY_MAP_V12.get(p.product_type, "العطور"),
    }


def run_smart_comparison(new_df: pd.DataFrame, store_df: pd.DataFrame,
                          new_name_col: str, store_name_col: str,
                          new_sku_col: str = None, store_sku_col: str = None,
                          new_img_col: str = None,
                          t_dup: int = 88, t_near: int = 75, t_review: int = 55,
                          brands_list: list = None,
                          store_brand_col: str = None) -> pd.DataFrame:
    """
    مقارنة منتجات المنافس مع متجر المرجع (Cluster Matching v12).
    مخرجات: أعمدة عربية تشمل «الحالة» (مكرر، مشبوه، جديد، …) و«_idx» لربط صف الإدخال الأصلي.
    """
    # ── بناء سجلات المتجر ──────────────────────────────────────────
    store_records = []
    store_sku_set = set()
    for _, row in store_df.iterrows():
        sname = str(row.get(store_name_col, "") or "").strip()
        if not sname or sname.lower() in ("nan", "none", ""):
            continue
        sbrand = ""
        if store_brand_col and store_brand_col in store_df.columns:
            sbrand = str(row.get(store_brand_col, "") or "").strip()
        sku = str(row.get(store_sku_col, "") or "").strip() if store_sku_col else ""
        if sku:
            store_sku_set.add(sku.lower())
        store_records.append({"name": sname, "brand": sbrand,
                               "sku": sku,
                               "image": str(row.get("صورة المنتج", "") or ""),
                               "price": str(row.get("سعر المنتج", "") or "")})

    # ── بناء المحرك ────────────────────────────────────────────────
    engine = _ClusterMatchEngine_v12(store_records)

    # ── معالجة كل منتج منافس ───────────────────────────────────────
    results = []
    for i, row in new_df.iterrows():
        new_name = str(row.get(new_name_col, "") or "").strip()
        new_sku  = str(row.get(new_sku_col, "") or "").strip() if new_sku_col else ""
        new_img  = str(row.get(new_img_col, "") or "").strip() if new_img_col else \
                   str(row.get("صورة المنتج", "") or "").strip()
        if not new_name or new_name.lower() in ("nan", "none", ""):
            continue

        new_name = strip_trailing_discount_label(new_name)
        if not str(new_name).strip():
            continue

        # استخراج الماركة من قائمة الماركات المُمررة
        competitor_brand = ""
        if brands_list:
            nl = new_name.lower()
            for b in brands_list:
                if str(b).lower() in nl:
                    competitor_brand = b
                    break

        # تطابق SKU مباشر
        if new_sku and new_sku.lower() in store_sku_set:
            _sku_brand = clean_brand_name(competitor_brand or "")
            results.append({
                "الاسم الجديد": new_name, "SKU الجديد": new_sku,
                "الماركة": _sku_brand or "",
                "التصنيف": "العطور",
                "أقرب تطابق في المتجر": new_name, "نسبة التشابه": 100,
                "الحالة": "مكرر (SKU)", "سبب القرار": "تطابق SKU مباشر",
                "الإجراء": "حذف", "_idx": i, "_img": new_img,
            })
            continue

        # تشغيل المحرك
        r = engine.match(new_name, competitor_brand=competitor_brand,
                         t_dup=float(t_dup), t_critical=float(t_near))

        verdict = r["verdict"]
        score   = r["score"]
        reason  = r["reason"]
        best_store_name = r["matched_name"] or ""
        prod = r["product"]

        # تحويل الحكم إلى حالة/إجراء
        if verdict == "مكرر":
            status = "مكرر"; action = "حذف"
        elif verdict == "حرج":
            status = "مشبوه"; action = "مراجعة"
        elif verdict == "مستبعد":
            status = "مستبعد"; action = "تجاهل"
        else:
            status = "جديد"; action = "اعتماد"

        ptype_map = {
            "TESTER": "تستر", "SET": "طقم هدايا", "OIL": "زيت جسم",
            "BODY_WASH": "شاور جل", "LOTION": "عناية جسم", "CREAM": "عناية جسم",
            "BODY_SPRAY": "معطر جسم", "DEODORANT": "مزيل عرق",
            "HAIR_MIST": "عطر شعر", "MIST": "معطر جسم", "PERFUME": "عطر تجاري",
        }
        ptype_ar = ptype_map.get(prod.product_type, "عطر تجاري")
        cat = _CATEGORY_MAP_V12.get(prod.product_type, "العطور")
        _cb_disp = clean_brand_name(competitor_brand or "")
        brand_display = _cb_disp or (prod.brand_normalized or "")

        results.append({
            "الاسم الجديد":           new_name,
            "SKU الجديد":             new_sku,
            "الماركة":                brand_display,
            "التصنيف":                cat,
            "أقرب تطابق في المتجر":   best_store_name,
            "نسبة التشابه":           round(score, 1),
            "الحالة":                 status,
            "سبب القرار":             reason,
            "الإجراء":                action,
            "_idx":                   i,
            "_img":                   new_img,
        })

    return pd.DataFrame(results) if results else pd.DataFrame()



def ai_filter_suspects(suspects_df: pd.DataFrame, store_names: list,
                        api_key: str, store_df: pd.DataFrame) -> tuple:
    """
    فلتر AI للمنتجات المشبوهة.
    يُرجع (approved_df, rejected_df)
    approved_df  = مشبوهة تبيّن أنها جديدة فعلاً
    rejected_df  = مشبوهة تبيّن أنها مكررة أو يجب استبعادها
    """
    if suspects_df.empty:
        return pd.DataFrame(), suspects_df

    if not api_key or not HAS_ANTHROPIC:
        # بدون AI: اعتمد الكل كـ «جديد» (لا تُسقِط المشبوه صامتاً)
        return suspects_df, pd.DataFrame()

    try:
        client = anthropic.Anthropic(api_key=api_key)
        # بناء قائمة الأسماء للتحقق (حتى 80 اسم للحفاظ على التكلفة)
        store_sample = store_names[:120]
        store_list_str = "\n".join(f"- {n}" for n in store_sample)

        suspects_list = []
        for i, (_, row) in enumerate(suspects_df.iterrows()):
            suspects_list.append({
                "idx": i,
                "name": str(row.get("الاسم الجديد", "")),
                "score": float(row.get("نسبة التشابه", 0)),
                "matched": str(row.get("أقرب تطابق في المتجر", "")),
                "reason": str(row.get("سبب القرار", "")),
            })

        suspects_json = json.dumps(suspects_list, ensure_ascii=False)

        prompt = f"""أنت مدقق بيانات عطور بصرامة 100% لمتجر مهووس.

قاعدتك الذهبية: اختلاف الحجم (مثلاً 50مل و100مل) أو التركيز (EDP و EDT) أو النوع (تستر وعادي) يعني أن المنتج "جديد" قطعاً حتى لو تطابق الاسم تماماً. التشابه في الحروف مع اختلاف الماركة يعني "جديد".

**قائمة منتجاتنا الموجودة (عينة):**
{store_list_str}

**المنتجات المشبوهة للفحص:**
{suspects_json}

**قواعد القرار الصارمة (بالترتيب):**
1. حجم مختلف (50مل vs 100مل vs 150مل) → جديد قطعاً
2. تركيز مختلف (EDP vs EDT vs EDC) → جديد قطعاً
3. نوع مختلف (تستر vs عادي) → جديد قطعاً
4. ماركة مختلفة (حتى لو الاسم متشابه) → جديد قطعاً
5. نفس الاسم حرفياً + نفس الحجم + نفس التركيز → مكرر

أرجع JSON النقي فقط بدون أي نصوص خارجه:
{{"decisions": [{{"idx": 0, "decision": "جديد", "reason": "..."}} , ...]}}"""

        msg = anthropic_messages_create(
            client,
            model="claude-haiku-4-5",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="ai_filter_suspects_batch")
        if not data:
            return pd.DataFrame(), suspects_df

        decisions = {d["idx"]: d["decision"] for d in data.get("decisions", [])}

        approved_rows = []
        rejected_rows = []
        for i, (orig_idx, row) in enumerate(suspects_df.iterrows()):
            dec = decisions.get(i, "مكرر")
            if dec == "جديد":
                approved_rows.append(row)
            else:
                rejected_rows.append(row)

        approved  = pd.DataFrame(approved_rows) if approved_rows else pd.DataFrame()
        rejected  = pd.DataFrame(rejected_rows) if rejected_rows else pd.DataFrame()
        return approved, rejected

    except Exception as e:
        APP_LOG.exception("ai_filter_suspects failed: %s", e)
        return pd.DataFrame(), suspects_df


def _resolve_suspicious_with_ai(
    competitor_name: str,
    closest_store_match_name: str,
) -> str:
    """
    تحقق عميق من المنتجات المشتبه بها عبر Claude.
    - إذا كانت A و B نفس المنتج تماماً => 'YES'
    - غير ذلك => 'NO'
    في حال أي خطأ/تعذر: تُرجع "" (ويتم التعامل معها كـ 'NO' افتراضياً لحفظ البيانات).
    """
    api_key = _effective_anthropic_api_key()
    if not api_key or not HAS_ANTHROPIC:
        return ""
    try:
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            "You are an expert perfume evaluator. "
            "Compare these two product names:\n"
            f"Product A (Competitor): '{competitor_name}'\n"
            f"Product B (My Store): '{closest_store_match_name}'\n\n"
            "Are these the exact same product? "
            "Pay strict attention to concentrations (EDT, EDP, Parfum, Cologne), "
            "terms like 'Tester' or 'تستر', and sizes (ml).\n"
            "If they are the exact same product, reply strictly with the word 'YES'. "
            "If they are different in concentration, type, or size, reply strictly with the word 'NO'. "
            "Do not explain."
        )
        msg = anthropic_messages_create(
            client,
            model="claude-haiku-4-5",
            max_tokens=10,
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = str(msg.content[0].text).strip().upper()
        if raw.startswith("YES"):
            return "YES"
        if raw.startswith("NO"):
            return "NO"
        return ""
    except Exception:
        return ""




def _strip_brand_name_edges(raw: object) -> str:
    """إزالة فواصل/مسافات/علامات تنصيص زائدة من أطراف اسم الماركة (رفع سلة / CSV)."""
    return str(raw or "").strip().strip(' ,.-\'"')


def _clean_brand_value_for_salla_output(raw: object) -> str:
    """
    تنظيف اسم الماركة قبل تصدير منتجات السلة:
    - إزالة كلمات التوقف الشائعة مثل "تستر" و "عطر" (حتى لا تظهر في عمود الماركة وتسبب مشاكل).
    - الحفاظ على صيغة "Arabic | English" إن كانت موجودة.
    """
    b = _strip_brand_name_edges(raw)
    if not b:
        return ""

    ar_part = b
    en_part = ""
    if "|" in b:
        ar_part, en_part = b.split("|", 1)
        ar_part = ar_part.strip()
        en_part = en_part.strip()

    # إزالة stop-words من الجزء العربي فقط (الأهم لمنع "تستر"/"عطر" في الناتج)
    # بدون الاعتماد على \b لأن حدود الكلمات العربية قد تفشل.
    ar_part = re.sub(r"^(تستر|عطر|طقم|مجموعة)\s+", " ", ar_part, flags=re.IGNORECASE)
    ar_part = re.sub(r"(?:\s|^)(تستر|عطر|طقم|مجموعة)(?:\s|$)", " ", ar_part, flags=re.IGNORECASE)
    ar_part = re.sub(r"(?:^|\s)مزيل\s+عرق(?:\s|$)", " ", ar_part, flags=re.IGNORECASE)
    ar_part = re.sub(r"\s+", " ", ar_part).strip()

    if en_part:
        en_part = re.sub(r"(?:^|\s)(tester|parfum|perfume)(?:\s|$)", " ", en_part, flags=re.IGNORECASE)
        en_part = re.sub(r"\s+", " ", en_part).strip()

    if en_part:
        if not ar_part:
            # إن انتهى الجزء العربي بعد التنظيف، نعيد الجزء العربي الأصلي بدون تغيير (لتفادي فقد الماركة).
            ar_part = _strip_brand_name_edges(raw).split("|", 1)[0].strip() if "|" in _strip_brand_name_edges(raw) else _strip_brand_name_edges(raw)
            ar_part = re.sub(r"^(تستر|عطر|طقم|مجموعة)\s+", "", ar_part, flags=re.IGNORECASE)
            ar_part = re.sub(r"\s+", " ", ar_part).strip()
        return f"{ar_part} | {en_part}".strip(" |")
    return ar_part


def clean_brand_name(brand_raw: str) -> str:
    """تنظيف الماركة من الكلمات الخاطئة والأطوال غير المنطقية"""
    if not brand_raw:
        return ""
    b = _strip_brand_name_edges(brand_raw)
    if is_discount_like_brand(b):
        return ""
    if len(b.split()) > 3 or len(b) <= 2:
        return ""
    bad_words = [
        "تستر","عطر","شامبو","بلسم","لوشن","مقوي","مسكرة","حقيبة",
        "بخاخ","كريم","زيت","صابون","جل","معطر","بودي","مجموعة",
        "طقم","عينة","سمبل","tester","perfume","منتج","ميني","mini",
    ]
    if any(w in b.lower() for w in bad_words):
        return ""
    return b


def match_brand(name: str) -> dict:
    if not str(name).strip():
        return {"name": "", "page_url": ""}
    nl = str(name).lower()
    for b in st.session_state.get("pipe_session_brands", []):
        raw = str(b.get("name", "") or "")
        if not raw:
            continue
        for part in re.split(r"\s*\|\s*", raw):
            p = part.strip().lower()
            if p and p in nl:
                return {"name": raw, "page_url": str(b.get("page_url", "") or "")}
    bdf = st.session_state.brands_df
    if bdf is None:
        return {"name": "", "page_url": ""}
    col0 = bdf.columns[0]
    for _, row in bdf.iterrows():
        raw = str(row[col0])
        for part in re.split(r"\s*\|\s*", raw):
            p = part.strip().lower()
            if p and p in nl:
                return {
                    "name": raw,
                    "page_url": str(row.get(
                        "(SEO Page URL) رابط صفحة العلامة التجارية", "") or ""),
                }
    return {"name": "", "page_url": ""}


def generate_new_brand(brand_name: str) -> dict:
    """توليد ماركة بصيغة مهووس مع الترجمة وجلب الصور بالذكاء الاصطناعي."""
    brand_name = _strip_brand_name_edges(brand_name)
    if is_discount_like_brand(brand_name):
        return {
            "name": "",
            "page_url": "",
            "اسم الماركة": "",
            "وصف مختصر عن الماركة": "",
            "صورة شعار الماركة": "",
            "(إختياري) صورة البانر": "",
            "(Page Title) عنوان صفحة العلامة التجارية": "",
            "(SEO Page URL) رابط صفحة العلامة التجارية": "",
            "(Page Description) وصف صفحة العلامة التجارية": "",
        }
    key = st.session_state.api_key
    formatted_name = brand_name
    en_name        = brand_name
    desc           = f"علامة تجارية متخصصة في العطور الفاخرة - {brand_name}"

    if key:
        try:
            client = anthropic.Anthropic(api_key=key)
            prompt = (
                f"أنت خبير علامات تجارية عالمية. ترجم ونسق الماركة '{brand_name}' بدقة عالية. "
                "التزم بصيغة JSON المغلقة بدون أي نصوص أو مقدمات خارجها: "
                '{"formatted_name": "الاسم بالعربي | الاسم بالانجليزي", '
                '"en_name": "English name only", '
                '"desc": "وصف جذاب 30 كلمة لمتجر مهووس"}'
            )
            msg = anthropic_messages_create(
                client,
                model="claude-3-haiku-20240307", max_tokens=250,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = msg.content[0].text.strip()
            data = _parse_json_object_from_llm_text(raw, context="generate_new_brand")
            if data:
                formatted_name = data.get("formatted_name", brand_name)
                en_name        = data.get("en_name", brand_name)
                desc           = data.get("desc", desc)
        except Exception:
            pass

    slug = to_slug(en_name)
    display_name = formatted_name.split("|")[0].strip() if "|" in formatted_name else formatted_name
    out = {
        "name":                                          display_name,
        "page_url":                                      slug,
        "اسم الماركة":                                   formatted_name,
        "وصف مختصر عن الماركة":                         desc,
        "صورة شعار الماركة":                            fetch_image(f"{en_name} brand logo perfume"),
        "(إختياري) صورة البانر":                        fetch_image(f"{en_name} brand banner"),
        "(Page Title) عنوان صفحة العلامة التجارية":     f"عطور {formatted_name} الأصلية | مهووس",
        "(SEO Page URL) رابط صفحة العلامة التجارية":    slug,
        "(Page Description) وصف صفحة العلامة التجارية": f"تسوّق أحدث عطور {formatted_name} الأصلية الفاخرة بأسعار حصرية من متجر مهووس.",
    }
    register_pipe_session_brand(
        out.get("اسم الماركة", "") or "",
        out.get("(SEO Page URL) رابط صفحة العلامة التجارية", "") or "",
    )
    return out


def match_category(name: str, gender: str = "") -> str:
    t = (str(name) + " " + str(gender)).lower()
    if any(w in t for w in ["رجال", "للرجال", "men", "homme", "رجالي"]):
        return "العطور > عطور رجالية"
    if any(w in t for w in ["نساء", "للنساء", "women", "femme", "نسائي"]):
        return "العطور > عطور نسائية"
    return "العطور > عطور للجنسين"


def to_slug(text: str) -> str:
    """مسار URL لاتيني: أحرف إنجليزية وأرقام فقط — بدون تحويل صوتي للعربية."""
    out = ""
    for c in str(text).lower():
        if c.isascii() and c.isalnum():
            out += c
        elif c in " -_":
            out += "-"
    return re.sub(r"-+", "-", out).strip("-")


def _append_sku_to_seo_slug(url: str, sku_suffix: str) -> str:
    """يُلحق لاحقة SKU بمسار SEO لتفادي التصادمات."""
    u = str(url or "").strip()
    if not sku_suffix:
        return u
    suf_raw = str(sku_suffix).strip()
    suf = to_slug(suf_raw)
    if not suf:
        suf = re.sub(r"[^a-z0-9-]+", "-", suf_raw.lower()).strip("-")
    if not suf:
        return u
    if u.endswith(suf) or u.endswith("-" + suf):
        return u
    return f"{u}-{suf}".strip("-") if u else suf


# تحديث وتوحيد مفاتيح الماركات الجديدة لتطابق صيغة ملف مهووس
for b in st.session_state.new_brands:
    if "اسم الماركة" not in b:
        bn = b.pop("اسم العلامة التجارية", b.get("اسم الماركة", ""))
        b["اسم الماركة"] = bn
        b["وصف مختصر عن الماركة"] = b.pop("وصف العلامة التجارية", f"علامة تجارية متخصصة في العطور الفاخرة - {bn}")
        b["صورة شعار الماركة"] = b.pop("صورة العلامة التجارية", "")
        b["(إختياري) صورة البانر"] = ""
        b["(Page Title) عنوان صفحة العلامة التجارية"] = f"عطور {bn} الأصلية | مهووس"
        b["(SEO Page URL) رابط صفحة العلامة التجارية"] = b.pop("(SEO Page URL) رابط صفحة العلامة التجارية", to_slug(bn))
        b["(Page Description) وصف صفحة العلامة التجارية"] = f"تسوّق أحدث عطور {bn} الأصلية. تشكيلة فاخرة تناسب ذوقك بأسعار حصرية من متجر مهووس."


def _enforce_salla_product_seo_limits(title: str, desc: str) -> tuple[str, str]:
    """حدود سلة: عنوان الصفحة 60 حرفاً، الوصف 160 حرفاً."""
    t = str(title or "")
    d = str(desc or "")
    if len(t) > 60:
        t = t[:60]
    if len(d) > 160:
        d = d[:160]
    return t, d


def gen_seo(
    name: str,
    brand: dict,
    size: str,
    tester: bool,
    gender: str,
    sku_suffix: str = "",
    type_hint: str = "",
) -> dict:
    bname = brand.get("name", "")
    parts = re.split(r"\s*\|\s*", bname)
    ben   = parts[-1].strip() if len(parts) > 1 else bname
    pref  = "تستر" if tester else "عطر"
    title = f"{pref} {name} {size} | {ben}".strip()
    hint = f"{type_hint} {name} {bname}"
    if ("تستر" in hint or "tester" in hint.lower()) and "تستر" not in title:
        title = f"تستر {title}".strip()
    desc  = (f"تسوق {pref} {name} {size} الأصلي من {bname}. "
             f"عطر {gender} فاخر ثابت. أصلي 100% من مهووس.")
    slug = to_slug(f"{ben}-{name}-{size}".replace("مل", "ml"))
    slug = _append_sku_to_seo_slug(slug, sku_suffix)
    title, desc = _enforce_salla_product_seo_limits(title, desc)
    return {
        "url":   slug,
        "title": title,
        "desc":  desc,
        "alt":   f"زجاجة {pref} {name} {size} الأصلية",
    }


def _col_contains_any(df: pd.DataFrame, keywords: tuple) -> str:
    for c in df.columns:
        cs = str(c).lower()
        for k in keywords:
            if k.lower() in cs:
                return str(c)
    return ""


def ai_refine_seo_fields(
    name: str, brand: dict, size: str, tester: bool, gender: str,
    product_desc: str, base: dict,
    sku_suffix: str = "",
) -> dict:
    """يحسّن عنوان ووصف ومسار SEO باستخدام Claude — نفس منطق المعالج المستقل."""
    key = _effective_anthropic_api_key()
    if not key or not HAS_ANTHROPIC:
        out = dict(base)
        if sku_suffix:
            out["url"] = _append_sku_to_seo_slug(out.get("url", ""), sku_suffix)
        t2, d2 = _enforce_salla_product_seo_limits(out.get("title", ""), out.get("desc", ""))
        out["title"], out["desc"] = t2, d2
        return out
    try:
        client = anthropic.Anthropic(api_key=key)
        _site_hint = MAHWOUS_SITE_BASE.replace("https://", "").replace("http://", "")
        prompt = (
            f"أنت خبير SEO لمتجر عطور مهووس ({_site_hint}) في السعودية.\n"
            f"المنتج: {name}\nالماركة: {brand.get('name', '')}\n"
            f"الحجم/الجنس/نوع: {size} / {gender} / {'تستر' if tester else 'عطر'}\n"
            f"مقتطف من وصف المنتج (قد يحتوي HTML): {product_desc[:1400]}\n\n"
            f"مقترح أولي: url_slug={base['url']}\n"
            f"title={base['title']}\n"
            f"meta_description={base['desc']}\n\n"
            "أعد JSON فقط بدون أي نص خارج JSON:\n"
            '{"url_slug":"...","page_title":"...","meta_description":"..."}\n'
            "قواعد: meta_description حتى 160 حرفًا عربية فاخرة، page_title حتى 60 حرفًا، "
            "url_slug لاتيني صغير بشرطات فقط بدون مسافات."
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        d = _parse_json_object_from_llm_text(raw, context="ai_refine_seo_fields")
        if d:
            out = {
                "url": _append_sku_to_seo_slug(
                    str(d.get("url_slug", base["url"])).strip() or base["url"],
                    sku_suffix,
                ),
                "title": str(d.get("page_title", base["title"])).strip() or base["title"],
                "desc": str(d.get("meta_description", base["desc"])).strip() or base["desc"],
                "alt": base.get("alt", ""),
            }
            t2, d2 = _enforce_salla_product_seo_limits(out["title"], out["desc"])
            out["title"], out["desc"] = t2, d2
            return out
    except Exception:
        pass
    out = dict(base)
    if sku_suffix:
        out["url"] = _append_sku_to_seo_slug(out.get("url", ""), sku_suffix)
    t2, d2 = _enforce_salla_product_seo_limits(out.get("title", ""), out.get("desc", ""))
    out["title"], out["desc"] = t2, d2
    return out


def build_salla_seo_row(
    no: str,
    name: str,
    brand: dict,
    product_desc: str = "",
) -> dict:
    """صف SEO واحد بتنسيق سلة — موحّد بين المسار الآلي ومعالج SEO المستقل."""
    attrs = extract_product_attrs(name)
    size_s = attrs.get("size") or 0
    if not size_s:
        size = "100 مل"
    else:
        size = f"{int(size_s) if size_s == int(size_s) else size_s} مل"
    is_t = "تستر" in str(attrs.get("type", ""))
    nl = name.lower()
    gender = "للجنسين"
    if any(w in nl for w in ["رجال", "للرجال", "men", "homme"]):
        gender = "للرجال"
    elif any(w in nl for w in ["نساء", "للنساء", "women", "femme"]):
        gender = "للنساء"
    sku_suf = f"V-{no}"
    th = str(attrs.get("type", "") or "")
    base = gen_seo(
        name, brand, size, is_t, gender,
        sku_suffix=sku_suf,
        type_hint=th,
    )
    refined = ai_refine_seo_fields(
        name, brand, size, is_t, gender, product_desc, base,
        sku_suffix=sku_suf,
    )
    return {
        "No. (غير قابل للتعديل)": str(no),
        "اسم المنتج (غير قابل للتعديل)": name,
        "رابط مخصص للمنتج (SEO Page URL)": refined["url"],
        "عنوان صفحة المنتج (SEO Page Title)": refined["title"],
        "وصف صفحة المنتج (SEO Page Description)": refined["desc"],
    }


def _row_seo_incomplete(row: pd.Series, url_c: str, title_c: str, desc_c: str) -> bool:
    def _empty(v):
        s = str(v or "").strip()
        return not s or s.lower() in ("nan", "none")

    u = row.get(url_c, "") if url_c else ""
    t = row.get(title_c, "") if title_c else ""
    d = row.get(desc_c, "") if desc_c else ""
    return _empty(u) or _empty(t) or _empty(d)


def generate_seo_for_products_dataframe(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    يملأ SEO الناقص لكل صف يحتاج ذلك.
    يعيد: (جدول SEO كامل لكل الصفوف، جدول المُولّد فقط للتصدير)
    """
    if df is None or df.empty:
        return pd.DataFrame(), pd.DataFrame()
    name_col = "أسم المنتج" if "أسم المنتج" in df.columns else auto_guess_col(
        df.columns, ["اسم", "name", "منتج", "أسم"], df)
    if not name_col or name_col == "— لا يوجد —":
        return pd.DataFrame(), pd.DataFrame()
    no_col = "No." if "No." in df.columns else auto_guess_col(df.columns, ["no", "رقم"], df)
    brand_col = "الماركة" if "الماركة" in df.columns else auto_guess_col(
        df.columns, ["ماركة", "brand"], df)
    desc_col = "الوصف" if "الوصف" in df.columns else None
    url_c = _col_contains_any(df, ("رابط مخصص للمنتج", "seo page url"))
    title_c = _col_contains_any(df, ("عنوان صفحة المنتج", "seo page title"))
    desc_seo_c = _col_contains_any(df, ("وصف صفحة المنتج", "seo page description"))

    all_rows = []
    gen_only = []
    prog = st.session_state.get("_seo_batch_prog")
    total = len(df)
    for ix, (_, row) in enumerate(df.iterrows()):
        if prog is not None:
            prog.progress(int((ix + 1) / max(total, 1) * 100))
        name = str(row.get(name_col, "") or "").strip()
        if not name:
            continue
        need = True
        if url_c and title_c and desc_seo_c:
            need = _row_seo_incomplete(row, url_c, title_c, desc_seo_c)
        if not need:
            all_rows.append({
                "No. (غير قابل للتعديل)": str(row.get(no_col, "") or "") if no_col and no_col != "— لا يوجد —" else str(ix + 1),
                "اسم المنتج (غير قابل للتعديل)": name,
                "رابط مخصص للمنتج (SEO Page URL)": str(row.get(url_c, "") or ""),
                "عنوان صفحة المنتج (SEO Page Title)": str(row.get(title_c, "") or ""),
                "وصف صفحة المنتج (SEO Page Description)": str(row.get(desc_seo_c, "") or ""),
            })
            continue
        b_raw = str(row.get(brand_col, "") or "").strip() if brand_col and brand_col != "— لا يوجد —" else ""
        brand_d = match_brand(name)
        if not brand_d.get("name") and b_raw:
            brand_d = match_brand(b_raw)
        if not brand_d.get("name") and b_raw:
            brand_d = {"name": b_raw, "page_url": to_slug(b_raw)}
        pdesc = str(row.get(desc_col, "") or "") if desc_col else ""
        no_val = str(row.get(no_col, "") or "").strip() if no_col and no_col != "— لا يوجد —" else str(ix + 1)
        row_seo = build_salla_seo_row(no_val, name, brand_d, pdesc)
        all_rows.append(row_seo)
        gen_only.append(row_seo)
    full_df = pd.DataFrame(all_rows, columns=SALLA_SEO_COLS) if all_rows else pd.DataFrame(columns=SALLA_SEO_COLS)
    gen_df = pd.DataFrame(gen_only, columns=SALLA_SEO_COLS) if gen_only else pd.DataFrame(columns=SALLA_SEO_COLS)
    return full_df, gen_df


def fetch_image(name: str, tester: bool = False) -> str:
    gk = st.session_state.google_api
    cx = st.session_state.google_cse
    if not gk or not cx:
        return ""
    try:
        q = name + (" tester box" if tester else " perfume bottle")
        r = requests.get(
            "https://www.googleapis.com/customsearch/v1",
            params={"key": gk, "cx": cx, "q": q,
                    "searchType": "image", "num": 1, "imgSize": "large"},
            timeout=10,
        )
        items = r.json().get("items", [])
        return items[0]["link"] if items else ""
    except Exception:
        return ""


def scrape_product_url(url: str) -> dict:
    """سحب بيانات المنتج من رابط URL مع دعم Cloudflare وتحسين استخراج السعر والصور."""
    result = {"name": "", "price": "", "image": "", "images": [], "desc": "", "brand_hint": "", "error": ""}
    try:
        from bs4 import BeautifulSoup
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "ar,en-US;q=0.9,en;q=0.8",
        }
        resp = requests.get(url, headers=headers, timeout=20, allow_redirects=True)
        if resp.status_code != 200:
            result["error"] = f"خطأ HTTP {resp.status_code}"
            return result
        soup = BeautifulSoup(resp.text, "html.parser")
        html_text = resp.text

        # ── Name ──────────────────────────────────────────────────
        og_title = soup.find("meta", property="og:title")
        if og_title and og_title.get("content"):
            result["name"] = og_title["content"].strip()
        elif soup.find("h1"):
            result["name"] = soup.find("h1").get_text(" ", strip=True)
        elif soup.find("title"):
            result["name"] = soup.find("title").get_text(strip=True).split("|")[0].split("-")[0].strip()

        # ── Price ─────────────────────────────────────────────────
        # Try JSON-LD first
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                import json as _j
                data = _j.loads(script.string or "{}")
                if isinstance(data, list): data = data[0]
                offers = data.get("offers", data.get("Offers", {}))
                if isinstance(offers, list): offers = offers[0]
                p = offers.get("price", "")
                if p: result["price"] = str(p); break
            except Exception: pass
        if not result["price"]:
            meta_price = soup.find("meta", property="product:price:amount")
            if meta_price and meta_price.get("content"):
                result["price"] = meta_price["content"]
        if not result["price"]:
            price_match = re.search(r'(?:السعر|price|SAR|ر\.س|ريال)[^\d]*([\d\.,]+)', html_text, re.IGNORECASE)
            if price_match: result["price"] = price_match.group(1).replace(",", "")

        # ── Images ────────────────────────────────────────────────
        images = []
        og_img = soup.find("meta", property="og:image")
        if og_img and og_img.get("content"): images.append(og_img["content"])
        for img in soup.select("img[src]"):
            src = img.get("src", "")
            if src.startswith("http") and src not in images and any(
                    kw in src.lower() for kw in ["product","cdn","item","shop","perfume","bottle"]):
                images.append(src)
                if len(images) >= 6: break
        result["image"]  = images[0] if images else ""
        result["images"] = images[:6]

        # ── Description ───────────────────────────────────────────
        og_desc = soup.find("meta", property="og:description")
        if og_desc and og_desc.get("content"):
            result["desc"] = og_desc["content"].strip()
        elif soup.find("meta", attrs={"name": "description"}):
            result["desc"] = soup.find("meta", attrs={"name": "description"}).get("content", "").strip()

        # ── Brand ─────────────────────────────────────────────────
        og_brand = soup.find("meta", property="product:brand")
        if og_brand and og_brand.get("content"):
            result["brand_hint"] = og_brand["content"]

    except requests.exceptions.Timeout:
        result["error"] = "انتهت مهلة الاتصال (timeout)"
    except requests.exceptions.ConnectionError:
        result["error"] = "تعذّر الاتصال بالموقع"
    except Exception as e:
        result["error"] = f"خطأ: {str(e)[:100]}"
    return result


def extract_product_json_from_url(url: str, api_key: str) -> dict:
    """
    استخراج JSON من صفحة منتج عبر (requests + BeautifulSoup) ثم Claude.
    يرجع dict بمفاتيح عربية: (أسم المنتج، الماركة، سعر المنتج، الوصف، صورة المنتج)
    عند أي خطأ يرجع dict مفاتيحها مع قيم فارغة لضمان عدم انهيار التطبيق.
    """
    empty = {"أسم المنتج": "", "الماركة": "", "سعر المنتج": "", "الوصف": "", "صورة المنتج": ""}
    if not url or not isinstance(url, str):
        return empty
    try:
        from bs4 import BeautifulSoup
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "Accept-Language": "ar,en-US;q=0.9,en;q=0.8",
        }
        resp = requests.get(url, headers=headers, timeout=25, allow_redirects=True)
        if resp.status_code != 200:
            return {**empty, "الوصف": "", "سعر المنتج": ""}

        soup = BeautifulSoup(resp.text, "html.parser")
        page_title = ""
        if soup.title and soup.title.get_text(strip=True):
            page_title = soup.title.get_text(strip=True)

        # اجمع مرشحي الصور (OG ثم صور img)
        img_candidates = []
        og_imgs = soup.find_all("meta", attrs={"property": "og:image"})
        for m in og_imgs:
            c = m.get("content", "").strip()
            if c and c not in img_candidates:
                img_candidates.append(c)
        if not img_candidates:
            for img in soup.select("img[src]"):
                s = img.get("src", "").strip()
                if s and s.startswith("http") and s not in img_candidates:
                    img_candidates.append(s)
                    if len(img_candidates) >= 8:
                        break

        raw_text = soup.get_text(" ", strip=True)
        # تقليل حجم النص المرسل إلى Claude
        raw_text = raw_text[:12000]

        if not api_key or not HAS_ANTHROPIC:
            return empty

        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            "أنت محلل صفحات منتجات عطور بدقة.\n"
            "المطلوب: استخرج JSON فقط بدون أي نص خارج JSON.\n\n"
            f"رابط: {url}\n"
            f"عنوان الصفحة: {page_title}\n"
            f"مرشحو الصور: {img_candidates[:8]}\n\n"
            "نص الصفحة (قد يحتوي ضوضاء):\n"
            f"{raw_text}\n\n"
            "الحقول المطلوبة في JSON (اكتب مفاتيح عربية حرفياً):\n"
            "1) أسم المنتج\n"
            "2) الماركة\n"
            "3) سعر المنتج (رقم فقط إن أمكن أو كنص كما هو في الصفحة)\n"
            "4) الوصف (HTML بسيط فقط: <p> و/أو <ul><li>)\n"
            "5) صورة المنتج (اختَر واحداً من مرشحي الصور إن أمكن، وإلا \"\")\n\n"
            "قواعد صارمة:\n"
            "- لا تهلوس: إن لم تجد معلومة في النص/الصور المرشحة، اجعلها \"\".\n"
            "- JSON النقي فقط.\n"
            "مثال شكل JSON:\n"
            "{\"أسم المنتج\":\"\",\"الماركة\":\"\",\"سعر المنتج\":\"\",\"الوصف\":\"<p>...</p>\",\"صورة المنتج\":\"\"}"
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=600,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="scrape_product_page_ai")
        if not data:
            return empty
        out = {**empty}
        for k in out.keys():
            if k in data and data[k] is not None:
                out[k] = str(data.get(k, "")).strip()
        return out
    except Exception:
        return empty

def _ai_fetch_notes_only(name: str, brand_name: str, api_key: str) -> dict:
    """استدعاء AI صغير: يجلب المكونات الحقيقية بقوة مع تخفيض الحرارة."""
    try:
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            f"أنت خبير كيميائي للعطور. استخرج المكونات الحقيقية لعطر '{name}' "
            f"من ماركة '{brand_name}'. "
            "لا تهلوس مكونات غير موجودة — إذا لم تعرفها اكتب 'غير متوفر'. "
            "الرد يجب أن يكون JSON مغلق بدون أي مقدمات أو نصوص خارجه:\n"
            '{"top": "مكونات القمة", "heart": "مكونات القلب", '
            '"base": "مكونات القاعدة", "family": "العائلة العطرية", "year": "سنة الإصدار"}'
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=300,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        parsed = _parse_json_object_from_llm_text(raw, context="ai_fetch_notes_only")
        if parsed:
            return parsed
    except Exception:
        pass
    return {"top": "غير متوفر", "heart": "غير متوفر", "base": "غير متوفر",
            "family": "غير متوفر", "year": "غير معروف"}

def _build_html_description(name: str, tester: bool, brand: dict,
                             size: str, gender: str, conc: str,
                             notes: dict) -> str:
    """بناء وصف HTML للمنتج مع روابط ماركة وتصنيفات عبر MAHWOUS_SITE_BASE."""
    ptype     = "تستر" if tester else "عطر"
    brand_name = brand.get("name", "غير محدد")
    brand_url  = brand.get("page_url", "")
    brand_link = (
        f"<a href='{mahwous_brand_url(brand_url)}' target='_blank' rel='noopener'>{brand_name}</a>"
        if brand_url else brand_name
    )
    top    = notes.get("top",    "برغموت، ليمون")
    heart  = notes.get("heart",  "ورد، ياسمين")
    base   = notes.get("base",   "مسك، عنبر")
    family = notes.get("family", "عطرية")
    year   = notes.get("year",   "")
    gender_txt = ("للنساء" if "نساء" in gender
                  else "للرجال" if "رجال" in gender else "للجنسين")
    season = ("الربيع والخريف" if any(w in family.lower() for w in ["زهري","فواكه","floral","fruit"])
              else "الشتاء والمناسبات" if any(w in family.lower() for w in ["عودي","خشب","oud","wood"])
              else "جميع الفصول")
    h = []
    h.append(f'<h2>{ptype} {brand_name} {name} {conc} {size} {gender_txt}</h2>')
    _fam_s = str(family).strip()
    _fam_ok = bool(_fam_s) and "غير متوفر" not in _fam_s
    intro_tail = f"عطر {_fam_s} فاخر يجمع بين الأصالة والتميز. " if _fam_ok else "عطر فاخر يجمع بين الأصالة والتميز. "
    h.append(f'<p>اكتشف سحر <strong>{name}</strong> من <strong>{brand_link}</strong> — '
             f'{intro_tail}'
             f'صمّم خصيصاً {gender_txt} ليرسم بصمتك العطري بثقة وأناقة. '
             f'متوفّر بحجم {size} بتركيز <strong>{conc}</strong> لضمان ثبات استثنائي.</p>')
    h.append('<h3>تفاصيل المنتج</h3>')
    h.append('<ul>')
    h.append(f'<li><strong>الماركة:</strong> {brand_link}</li>')
    h.append(f'<li><strong>الاسم:</strong> {name}</li>')
    h.append(f'<li><strong>الجنس:</strong> {gender_txt}</li>')
    if _fam_ok:
        h.append(f'<li><strong>العائلة العطرية:</strong> {family}</li>')
    h.append(f'<li><strong>الحجم:</strong> {size}</li>')
    h.append(f'<li><strong>التركيز:</strong> {conc}</li>')
    if year and year != "غير معروف":
        h.append(f'<li><strong>سنة الإصدار:</strong> {year}</li>')
    h.append(f'<li><strong>نوع المنتج:</strong> {"تستر (Tester)" if tester else "عطر أصلي"}</li>')
    h.append('</ul>')
    h.append('<h3>رحلة العطر — الهرم العطري</h3>')
    h.append(f'<p>يأخذك <strong>{name}</strong> في رحلة عطرية متكاملة تبدأ بطزالة وتنتهي بدفء وعمق.</p>')
    h.append('<ul>')
    h.append(f'<li><strong>المقدمة (Top Notes):</strong> {top}</li>')
    h.append(f'<li><strong>القلب (Heart Notes):</strong> {heart}</li>')
    h.append(f'<li><strong>القاعدة (Base Notes):</strong> {base}</li>')
    h.append('</ul>')
    h.append('<h3>لماذا تختار هذا العطر؟</h3>')
    h.append('<ul>')
    h.append(f'<li><strong>الثبات والفوحان:</strong> تركيز {conc} يضمن فوحاناً يدوم طويلاً يلفت الأنظار.</li>')
    h.append(f'<li><strong>التميز والأصالة:</strong> من دار {brand_name} العريقة بتراث عطري أصيل.</li>')
    h.append('<li><strong>القيمة الاستثنائية:</strong> عطر فاخر بسعر مناسب من متجر مهووس الموثوق.</li>')
    h.append('<li><strong>الجاذبية المضمونة:</strong> عطر يجعلك محور الاهتمام في كل مكان تحضره.</li>')
    h.append('</ul>')
    h.append(f'<h3>متى وأين ترتديه؟</h3>')
    h.append(f'<p>مثالي لـ {season}. يلائم المناسبات الرسمية والسهرات واللقاءات العملية. '
             f'ينصح برشه على نقاط النبض والرسغاوات لأفضل ثبات.</p>')
    h.append('<h3>لمسة خبير من مهووس</h3>')
    h.append('<p>الفوحان: 8/10 | الثبات: 9/10 | نصيحة: ابدأ بكمية صغيرة وابنِ تدريجياً حتى تجد كميتك المثالية.</p>')
    h.append('<h3>الأسئلة الشائعة</h3>')
    h.append('<ul>')
    h.append('<li><strong>كم يدوم العطر؟</strong> بين 8-12 ساعة حسب البشرة ودرجة الحرارة.</li>')
    h.append('<li><strong>هل يناسب الاستخدام اليومي؟</strong> نعم، بكمية معتدلة للبيئات المختلفة.</li>')
    if tester:
        h.append('<li><strong>ما الفرق بين التستر والعطر العادي؟</strong> التستر نفس العطر تماماً بدون علبة خارجية، بسعر أقل.</li>')
    if _fam_ok:
        h.append(f'<li><strong>ما العائلة العطرية؟</strong> {family}.</li>')
    h.append(f'<li><strong>هل يناسب الطقس الحار في السعودية؟</strong> {season} هي الموسم المثالي له.</li>')
    h.append('<li><strong>ما مناسبات ارتداء هذا العطر؟</strong> المناسبات الرسمية، السهرات، واللقاءات العملية.</li>')
    h.append('</ul>')
    h.append('<h3>اكتشف أكثر من مهووس</h3>')
    slug = brand.get("page_url", "")
    if slug:
        h.append(
            f"<p>اكتشف <a href='{mahwous_brand_url(slug)}' target='_blank' rel='noopener'>عطور {brand_name}</a> | "
            f"<a href='{mahwous_category_url('categories/mens-perfumes')}' target='_blank' rel='noopener'>عطور رجالية</a> | "
            f"<a href='{mahwous_category_url('categories/womens-perfumes')}' target='_blank' rel='noopener'>عطور نسائية</a></p>"
        )
    h.append('<p><strong>عالمك العطري يبدأ من مهووس.</strong> أصلي 100% | شحن سريع داخل السعودية.</p>')
    return "\n".join(h)


def _infer_gender_from_text(text: str) -> str:
    t = str(text or "").lower()
    if any(w in t for w in ["نسائ", "نساء", "women", "femme"]):
        return "للنساء"
    if any(w in t for w in ["رجال", "للرجال", "men", "homme"]):
        return "للرجال"
    return "للجنسين"


_JUNK_CLEAN_NAME_PHRASES = (
    "بدون غطاء", "بدون كرتون", "بدون علبة", "إصدار قديم", "بدون غطا",
)

_DEFAULT_NOTE_TOP = "حمضيات، برغموت، ليمون"
_DEFAULT_NOTE_HEART = "ورد، ياسمين، إيريس"
_DEFAULT_NOTE_BASE = "مسك، عنبر، خشب صندل"


def _strip_junk_phrases_from_clean_name(s: str) -> str:
    t = str(s or "")
    for j in _JUNK_CLEAN_NAME_PHRASES:
        t = re.sub(re.escape(j), " ", t, flags=re.IGNORECASE)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _clean_pname_for_fallback(pname: str) -> str:
    """تنظيف اسم المنتج للوصف الاحتياطي عند فشل أو فراغ مخرجات Claude."""
    orig = str(pname or "").strip()
    if not orig:
        return orig
    t = _strip_junk_phrases_from_clean_name(orig)
    for phrase in ("(بدون غطاء)", "(بدون كرتون)", "(بدون علبة)", "(بدون غطا)"):
        t = t.replace(phrase, " ")
    t = re.sub(r"\b(?:تستر|تيستر|tester)\b", " ", t, flags=re.IGNORECASE)
    t = re.sub(r"\s+", " ", t).strip()
    return t or orig


def _generate_fallback_html(raw_name: str, brand: str) -> str:
    """قالب HTML احتياطي متوافق مع سلة — لا يُرجع وصفاً فارغاً عند فشل API."""
    rn = str(raw_name or "").strip() or "منتج عطور"
    br = str(brand or "").strip() or "غير محدد"
    site_root = MAHWOUS_SITE_BASE.rstrip("/")
    return "\n".join(
        [
            f"<h2>{rn}</h2>",
            f"<p>نقدّم لك <strong>{rn}</strong> من دار <strong>{br}</strong> — "
            "منتج أصلي مختار بعناية من متجر مهووس، مع شحن سريع وخدمة موثوقة داخل السعودية.</p>",
            "<h3>تفاصيل المنتج</h3>",
            "<ul>",
            f"<li><strong>الماركة:</strong> {br}</li>",
            f"<li><strong>الاسم:</strong> {rn}</li>",
            "</ul>",
            "<h3>لماذا تختار مهووس؟</h3>",
            "<p>نوفر لك عطوراً أصلية مع ضمان الجودة وتجربة شراء سلسة تناسب تطلعاتك.</p>",
            f'<p><a href="{site_root}/" target="_blank" rel="noopener">https://mahwous.com/</a></p>',
        ]
    )


def _strip_brand_tokens_from_clean_name(clean_name: str, brand_clean: str) -> str:
    """يزيل تكرار أجزاء الماركة العربية/الإنجليزية من اسم الرائحة (تدقيق #29)."""
    cn = str(clean_name or "").strip()
    if not cn or not str(brand_clean or "").strip():
        return cn
    b = str(brand_clean).strip()
    parts = [p.strip() for p in re.split(r"\s*\|\s*", b) if p.strip()]
    if not parts:
        parts = [b]
    for _ in range(6):
        prev = cn
        for p in parts:
            if len(p) < 2:
                continue
            ep = re.escape(p)
            cn = re.sub(rf"(^|\s){ep}(\s|$)", " ", cn, flags=re.IGNORECASE)
            cn = re.sub(r"\s+", " ", cn).strip()
        if cn == prev:
            break
    return cn.strip()


def _normalize_product_size_ml(size: str) -> str:
    """تطبيع الحجم إلى «رقم مل»؛ قيم مثل .5 بلا وحدة تُفسَّر كـ 50 مل."""
    s = str(size or "").strip()
    if not s or s.lower() in ("nan", "none"):
        return ""
    low = s.lower()
    if "مل" in s or re.search(r"\bml\b", low):
        num_m = re.search(r"(\d+(?:[.,]\d+)?)", s.replace("٫", ".").replace(",", "."))
        if num_m:
            val = float(num_m.group(1).replace(",", "."))
            if abs(val - round(val)) < 1e-9:
                return f"{int(round(val))} مل"
            t = f"{val:.4f}".rstrip("0").rstrip(".")
            return f"{t} مل"
        return s
    t = s.replace("٫", ".").replace(",", ".").strip()
    m = re.fullmatch(r"(\d+\.?\d*|\.\d+)", t)
    if not m:
        return s
    val = float(m.group(1))
    if 0 < val < 1:
        val = val * 100.0
    if abs(val - round(val)) < 1e-9:
        return f"{int(round(val))} مل"
    t2 = f"{val:.4f}".rstrip("0").rstrip(".")
    return f"{t2} مل"


def _strip_type_words_from_clean_name_for_display(clean_name: str, ai_type: str) -> str:
    """يمنع تكرار نوع المنتج داخل clean_name (مثل «عطر» + نوع «طقم»)."""
    cn = str(clean_name or "").strip()
    if not cn:
        return cn
    if ai_type == "طقم":
        cn = re.sub(r"^(?:عطر|تستر|تيستر|مجموعة)\s+", "", cn, flags=re.IGNORECASE).strip()
    elif ai_type in ("عطر", "تستر"):
        cn = re.sub(r"^(?:عطر|تستر|تيستر)\s+", "", cn, count=1, flags=re.IGNORECASE).strip()
    return cn


def _normalize_ai_notes_triplet(
    top: str, heart: str, base: str,
) -> tuple[str, str, str]:
    """لا يُسمح بـ «غير متوفر» في الهرم العطري — بدائل منطقية."""
    def _one(v: str, fb: str) -> str:
        s = str(v or "").strip()
        if not s or s in ("غير متوفر", "غير معروف", "N/A", "n/a", "unknown"):
            return fb
        return s

    return (
        _one(top, _DEFAULT_NOTE_TOP),
        _one(heart, _DEFAULT_NOTE_HEART),
        _one(base, _DEFAULT_NOTE_BASE),
    )


def _build_html_description_from_product_ai_parts(
    *,
    formatted_name: str,
    product_type: str,
    clean_name: str,
    brand_name: str,
    brand_page_url: str,
    concentration: str,
    size: str,
    top_notes: str,
    heart_notes: str,
    base_notes: str,
    gender: str,
) -> str:
    """
    HTML وفق قالب MAHWOUS الحالي (مع اختلاف "نوع المنتج" ودمج Notes القادمة من Claude).
    """
    brand_link = (
        f"<a href='{mahwous_brand_url(brand_page_url)}' target='_blank' rel='noopener'>{brand_name}</a>"
        if brand_page_url
        else brand_name
    )
    tester_flag = product_type == "تستر"
    product_type_item = (
        "تستر (Tester)"
        if product_type == "تستر"
        else "عطر أصلي"
        if product_type == "عطر"
        else product_type
    )

    family = "غير متوفر"
    season = "جميع الفصول"
    gender_txt = ("للنساء" if "نساء" in gender else "للرجال" if "رجال" in gender else "للجنسين")
    h = []
    h.append(f"<h2>{formatted_name} {gender_txt}</h2>")
    h.append(
        f"<p>اكتشف سحر <strong>{clean_name}</strong> من <strong>{brand_link}</strong> — "
        f"صمّم خصيصاً {gender_txt} ليرسم بصمتك العطري بثقة وأناقة. "
        f"متوفّر بحجم {size} بتركيز <strong>{concentration}</strong> لضمان ثبات استثنائي.</p>"
    )

    h.append("<h3>تفاصيل المنتج</h3>")
    h.append("<ul>")
    h.append(f"<li><strong>الماركة:</strong> {brand_link}</li>")
    h.append(f"<li><strong>الاسم:</strong> {clean_name}</li>")
    h.append(f"<li><strong>الجنس:</strong> {gender_txt}</li>")
    # لا نضيف العائلة إذا لم تتوفر
    h.append(f"<li><strong>الحجم:</strong> {size}</li>")
    h.append(f"<li><strong>التركيز:</strong> {concentration}</li>")
    h.append(f"<li><strong>نوع المنتج:</strong> {product_type_item}</li>")
    h.append("</ul>")

    h.append("<h3>رحلة العطر — الهرم العطري</h3>")
    h.append(f"<p>يأخذك <strong>{clean_name}</strong> في رحلة عطرية متكاملة تبدأ بطزالة وتنتهي بدفء وعمق.</p>")
    h.append("<ul>")
    h.append(f"<li><strong>المقدمة (Top Notes):</strong> {top_notes}</li>")
    h.append(f"<li><strong>القلب (Heart Notes):</strong> {heart_notes}</li>")
    h.append(f"<li><strong>القاعدة (Base Notes):</strong> {base_notes}</li>")
    h.append("</ul>")

    h.append("<h3>لماذا تختار هذا العطر؟</h3>")
    h.append("<ul>")
    h.append(f"<li><strong>الثبات والفوحان:</strong> تركيز {concentration} يضمن فوحاناً يدوم طويلاً يلفت الأنظار.</li>")
    h.append(f"<li><strong>التميز والأصالة:</strong> من دار {brand_name} العريقة بتراث عطري أصيل.</li>")
    h.append("<li><strong>القيمة الاستثنائية:</strong> عطر فاخر بجودة استثنائية من متجر مهووس الموثوق.</li>")
    h.append("<li><strong>الجاذبية المضمونة:</strong> عطر يجعلك محور الاهتمام في كل مكان تحضره.</li>")
    h.append("</ul>")

    h.append("<h3>متى وأين ترتديه؟</h3>")
    h.append(
        f"<p>مثالي لـ {season}. يلائم المناسبات الرسمية والسهرات واللقاءات العملية. "
        f"ينصح برشه على نقاط النبض والرسغاوات لأفضل ثبات.</p>"
    )

    h.append("<h3>لمسة خبير من مهووس</h3>")
    h.append("<p>الفوحان: 8/10 | الثبات: 9/10 | نصيحة: ابدأ بكمية صغيرة وابنِ تدريجياً حتى تجد كميتك المثالية.</p>")

    h.append("<h3>الأسئلة الشائعة</h3>")
    h.append("<ul>")
    h.append("<li><strong>كم يدوم العطر؟</strong> بين 8-12 ساعة حسب البشرة ودرجة الحرارة.</li>")
    h.append("<li><strong>هل يناسب الاستخدام اليومي؟</strong> نعم، بكمية معتدلة للبيئات المختلفة.</li>")
    if tester_flag:
        h.append(
            "<li><strong>ما الفرق بين التستر والعطر العادي؟</strong> "
            "التستر نفس العطر تماماً بدون علبة خارجية، بتغليف مبسّط يناسب الاستخدام الشخصي.</li>"
        )
    h.append(f"<li><strong>هل يناسب الطقس الحار في السعودية؟</strong> {season} هي الموسم المثالي له.</li>")
    h.append("<li><strong>ما مناسبات ارتداء هذا العطر؟</strong> المناسبات الرسمية، السهرات، واللقاءات العملية.</li>")
    h.append("</ul>")

    h.append("<h3>اكتشف أكثر من مهووس</h3>")
    h.append(
        f"<p><a href=\"{MAHWOUS_SITE_BASE}/\" target=\"_blank\" rel=\"noopener\">متجر مهووس</a></p>"
    )
    h.append("<p><strong>عالمك العطري يبدأ من مهووس.</strong> أصلي 100% | شحن سريع داخل السعودية.</p>")
    return "\n".join(h)


def _ai_enrich_product_row(raw_competitor_product_name: str, api_key: str) -> dict:
    """
    إثراء منتج عبر Claude:
    - يستخرج نوع/اسم نظيف/ماركة/تركيز/حجم + Notes (Top/Heart/Base)
    - يُبنى اسم جاهز للتصدير بصيغة: {type} {clean_name} {brand} {concentration} {size}
    - يولّد HTML مطابق لقالب MAHWOUS مع Notes.
    """
    empty = {"formatted_name": "", "html_description": "", "brand": ""}
    if not raw_competitor_product_name or not api_key or not HAS_ANTHROPIC:
        return empty
    cache = getattr(st.session_state, "_ai_product_enrich_cache", None)
    if cache is None:
        cache = {}
        st.session_state._ai_product_enrich_cache = cache
    cache_key = str(raw_competitor_product_name).strip().lower()
    if cache_key in cache:
        return cache[cache_key]

    try:
        client = anthropic.Anthropic(api_key=api_key)
        allowed_types = ["عطر", "تستر", "طقم", "مزيل عرق", "معطر شعر", "معطر جسم"]
        prompt = (
            "أنت خبير إدخال بيانات + خبير عطور (Master Perfumer).\n"
            "استخرج من اسم منتج المنافس حقولاً دقيقة جداً.\n\n"
            f"الاسم الخام من المنافس (للمرجع فقط — لا تنسخه حرفياً في الوصف النهائي): {raw_competitor_product_name}\n\n"
            "أعد JSON فقط (بدون أي نص خارج JSON) بمفاتيح عربية حرفياً:\n"
            "{"
            "\"type\":\"\","
            "\"clean_name\":\"\","
            "\"brand\":\"(Arabic | English)\","
            "\"concentration\":\"\","
            "\"size\":\"\","
            "\"top_notes\":\"\","
            "\"heart_notes\":\"\","
            "\"base_notes\":\"\""
            "}\n\n"
            "قواعد إلزامية:\n"
            f"1) type واحد فقط من: {allowed_types}\n"
            "2) clean_name: اسم الرائحة/الخط فقط — بدون ماركة، بدون حجم، بدون تركيز، بدون نوع المنتج في الاسم. "
            "احذف تماماً عبارات مثل: «بدون غطاء»، «بدون كرتون»، «تستر»، «تيستر»، «tester» إن وُجدت في الاسم.\n"
            "3) لا تكرّر نوع المنتج: إن كان type «طقم» فلا تبدأ clean_name بـ «عطر» أو «طقم» مكرر. "
            "إن كان type «عطر» فلا تكرّر كلمة «عطر» داخل clean_name.\n"
            "4) brand بصيغة \"عربي | English\" إن أمكن؛ وإلا أعد \"\".\n"
            "5) concentration بصيغة عربية صحيحة (مثل أو دو بارفيوم، أو دو تواليت، بارفيوم).\n"
            "6) size مثل «100 مل» إن وُجد في الاسم؛ وإلا استنتج المعتاد للخط أو أعد \"\".\n"
            "7) top_notes / heart_notes / base_notes: مكونات عربية مفصولة بفواصل. "
            "استخدم معرفتك بالعطر إن أمكن؛ وإن لم تتأكد استخدم مكونات منطقية للعائلة العطرية للماركة. "
            "ممنوع تماماً إخراج «غير متوفر» أو ترك الحقول فارغة — املأها دائماً بمكونات معقولة.\n"
            "8) لا تُدرج أي سعر أو رمز عملة أو رقم يشبه السعر في أي حقل.\n"
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=600,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="_ai_enrich_product_row")
        if not data:
            return empty

        ai_type = str(data.get("type", "") or "").strip()
        if ai_type not in ["عطر", "تستر", "طقم", "مزيل عرق", "معطر شعر", "معطر جسم"]:
            ai_type = "عطر"

        clean_name = str(data.get("clean_name", "") or "").strip()
        clean_name = _strip_junk_phrases_from_clean_name(clean_name)
        clean_name = _strip_type_words_from_clean_name_for_display(clean_name, ai_type)
        brand_raw = str(data.get("brand", "") or "").strip()
        conc = str(data.get("concentration", "") or "").strip()
        size = str(data.get("size", "") or "").strip()
        top = str(data.get("top_notes", "") or "").strip()
        heart = str(data.get("heart_notes", "") or "").strip()
        base = str(data.get("base_notes", "") or "").strip()

        if not clean_name or not brand_raw or not conc or not size:
            return empty

        brand_clean = _clean_brand_value_for_salla_output(brand_raw)
        if not brand_clean:
            return empty

        clean_name = _strip_brand_tokens_from_clean_name(clean_name, brand_clean)
        size = _normalize_product_size_ml(size)
        if not clean_name or not size:
            return empty

        top, heart, base = _normalize_ai_notes_triplet(top, heart, base)

        formatted_name = f"{ai_type} {clean_name} {brand_clean} {conc} {size}".strip()
        formatted_name = re.sub(r"\s+", " ", formatted_name).strip()

        bmatch = match_brand(brand_clean) if st.session_state.get("brands_df", None) is not None else {"name": brand_clean, "page_url": ""}
        brand_page_url = str(bmatch.get("page_url", "") or "").strip()
        if not brand_page_url:
            parts = [p.strip() for p in str(brand_clean).split("|", 1)]
            en_part = parts[1] if len(parts) == 2 else parts[0]
            brand_page_url = to_slug(en_part)

        gender_txt = _infer_gender_from_text(formatted_name + " " + raw_competitor_product_name)

        html = _build_html_description_from_product_ai_parts(
            formatted_name=formatted_name,
            product_type=ai_type,
            clean_name=clean_name,
            brand_name=brand_clean,
            brand_page_url=brand_page_url,
            concentration=conc,
            size=size,
            top_notes=top,
            heart_notes=heart,
            base_notes=base,
            gender=gender_txt,
        )

        out = {"formatted_name": formatted_name, "html_description": html, "brand": brand_clean}
        cache[cache_key] = out
        return out
    except Exception:
        return empty


def ai_generate(name: str, tester: bool, brand: dict,
                size: str, gender: str, conc: str) -> str:
    """توليد الوصف: AI يجلب المكونات فقط (300 token) والكود يولد HTML الكامل (مجاني)."""
    key = st.session_state.api_key
    if not key:
        return "<p>أضف مفتاح Anthropic API في الإعدادات أولاً</p>"
    # استدعاء AI صغير لجلب المكونات فقط
    notes = _ai_fetch_notes_only(name, brand.get("name", ""), key)
    # الكود يولد HTML الكامل مجاناً
    return _build_html_description(name, tester, brand, size, gender, conc, notes)


def build_empty_salla_row() -> dict:
    r = {c: "" for c in SALLA_COLS}
    r["النوع "]                    = "منتج"
    r["نوع المنتج"]               = "منتج جاهز"
    r["هل يتطلب شحن؟"]           = "نعم"
    r["خاضع للضريبة ؟"]          = "نعم"
    r["الكمية المتوفرة"]          = "0"
    r["الوزن"]                    = "0.2"
    r["وحدة الوزن"]               = "kg"
    r["حالة المنتج"]              = "مرئي"
    r["اقصي كمية لكل عميل"]      = "0"
    r["إخفاء خيار تحديد الكمية"] = "0"
    r["اضافة صورة عند الطلب"]    = "0"
    return r


def fill_row(name, price="", sku="", image="", desc="",
             brand=None, category="", seo=None, no="",
             weight="0.2", weight_unit="kg", size="") -> dict:
    if brand is None:
        brand = {}
    if seo is None:
        seo = {}
    r = build_empty_salla_row()
    r["No."]             = str(no)
    r["أسم المنتج"]      = str(name)
    r["سعر المنتج"]      = normalize_price_digits(price)
    r["رمز المنتج sku"]  = str(sku)
    r["صورة المنتج"]     = str(image)
    r["وصف صورة المنتج"] = seo.get("alt", "")
    r["الوصف"]           = compact_html_desc(str(desc))
    r["الماركة"]         = _clean_brand_value_for_salla_output(brand.get("name", ""))
    r["تصنيف المنتج"]    = str(category)
    r["الوزن"]           = str(weight) if weight else "0.2"
    r["وحدة الوزن"]      = str(weight_unit) if weight_unit else "kg"
    # If no price → set quantity to 0
    if not str(price).strip() or str(price).strip() in ("0", "nan", "None"):
        r["اقصي كمية لكل عميل"] = "0"
    return r

# ╔══════════════════════════════════════════════════════════════════╗
# ║  EXPORT FUNCTIONS                                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def _style_header_row(ws, row_num: int, cols: list,
                      bg: str = "0F0E0D", fg: str = "B8933A"):
    for i, col in enumerate(cols, 1):
        c = ws.cell(row_num, i, col)
        c.font      = Font(bold=True, color="FFFFFF" if bg != "E8D5B7" else "0F0E0D",
                           name="Cairo", size=9)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True, readingOrder=2)
        c.border    = Border(bottom=Side(style="thin", color=fg))
    ws.row_dimensions[row_num].height = 30


def get_latest_export_data() -> Optional[pd.DataFrame]:
    """آخر جدول منتجات جاهز للتصدير من المسار الآلي (متزامن مع المحرر والإثراء)."""
    if not hasattr(st, "session_state"):
        return None
    df = st.session_state.get("pipe_export_df")
    if df is None:
        df = st.session_state.get("pipe_approved")
    return df


def _prepare_salla_product_df_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """دمج أعمدة سلة، قيم افتراضية إلزامية، وتنظيف أسعار قبل CSV/XLSX."""
    df = dedupe_products_df(df.copy())
    for col in SALLA_COLS:
        if col not in df.columns:
            df[col] = ""
    df = df.reindex(columns=list(SALLA_COLS))
    df["النوع "] = "منتج"
    df["نوع المنتج"] = "منتج جاهز"
    df["حالة المنتج"] = "مرئي"
    df["خاضع للضريبة ؟"] = "نعم"
    df["الكمية المتوفرة"] = df["الكمية المتوفرة"].apply(
        lambda x: "0" if not str(x or "").strip() or str(x).lower() in ("nan", "none") else str(x)
    )
    df["سعر المنتج"] = df["سعر المنتج"].apply(sanitize_salla_price_for_export)
    df["السعر المخفض"] = df["السعر المخفض"].apply(sanitize_salla_price_for_export)

    def _fix_sale_price_vs_regular(row: pd.Series) -> str:
        disc_s = str(row.get("السعر المخفض", "") or "").strip()
        if not disc_s or disc_s.lower() in ("nan", "none"):
            return ""
        ok_r, p_reg = parse_price_numeric(row.get("سعر المنتج", ""))
        ok_d, p_disc = parse_price_numeric(row.get("السعر المخفض", ""))
        if ok_r and ok_d and p_disc >= p_reg:
            return ""
        return disc_s

    df["السعر المخفض"] = df.apply(_fix_sale_price_vs_regular, axis=1)

    def _first_product_image_url(v) -> str:
        s = str(v or "").strip()
        if not s or s.lower() in ("nan", "none"):
            return ""
        if "," in s:
            return s.split(",")[0].strip()
        return s

    df["صورة المنتج"] = df["صورة المنتج"].apply(_first_product_image_url)
    df["الماركة"] = df["الماركة"].apply(_clean_brand_value_for_salla_output)
    # SKU: بادئة V- لتقليل التصادم مع رموز المتجر (#33)
    if "رمز المنتج sku" in df.columns:

        def _export_sku_row(row) -> str:
            s = str(row.get("رمز المنتج sku", "") or "").strip()
            if s and s.lower() not in ("nan", "none"):
                if re.match(r"^v-", s, re.IGNORECASE):
                    rest = s[2:].strip()
                    return f"V-{rest}" if rest else ""
                return f"V-{s}"
            no = str(row.get("No.", "") or "").strip()
            if no and no.lower() not in ("nan", "none"):
                return f"V-{no}"
            return ""

        df["رمز المنتج sku"] = df.apply(_export_sku_row, axis=1)
    # وزن حسب نوع المنتج في الاسم (#34)
    if "أسم المنتج" in df.columns:
        df["الوزن"] = df["أسم المنتج"].apply(
            lambda n: "0.5" if "طقم" in str(n) else "0.2"
        )
    df["وحدة الوزن"] = "kg"
    # تصنيف كامل من الاسم المعياري (#40)
    if "أسم المنتج" in df.columns:
        df["تصنيف المنتج"] = df["أسم المنتج"].apply(
            lambda n: match_category(str(n), "")
        )

    def _export_cost_row(row) -> str:
        raw = row.get("سعر التكلفة", "")
        s = sanitize_salla_price_for_export(raw)
        if s:
            return s
        ok, price = parse_price_numeric(row.get("سعر المنتج", ""))
        if ok and price > 0:
            d = round(price * 0.70, 2)
            if abs(d - int(d)) < 1e-9:
                return str(int(d))
            ds = f"{d:.2f}".rstrip("0").rstrip(".")
            return ds
        return ""

    df["سعر التكلفة"] = df.apply(_export_cost_row, axis=1)

    def _promo_row(row) -> str:
        v = str(row.get("العنوان الترويجي", "") or "").strip()
        if v and v.lower() not in ("nan", "none"):
            return v
        name = str(row.get("أسم المنتج", "") or "")
        ptype = str(row.get("نوع المنتج", "") or "")
        if "تستر" in name or "تستر" in ptype:
            return "إصدار حصري"
        if "طقم" in name or "مجموعة" in name:
            return "الأكثر مبيعاً"
        return "عطر فاخر"

    df["العنوان الترويجي"] = df.apply(_promo_row, axis=1)

    def _disc_date(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none", "nat"):
            return ""
        return format_salla_date_yyyy_mm_dd(v) or ""

    df["تاريخ بداية التخفيض"] = df["تاريخ بداية التخفيض"].apply(_disc_date)
    df["تاريخ نهاية التخفيض"] = df["تاريخ نهاية التخفيض"].apply(_disc_date)
    return df


def export_product_xlsx(df: pd.DataFrame) -> bytes:
    if df is None:
        df = pd.DataFrame(columns=SALLA_COLS)
    elif not df.empty:
        df = _prepare_salla_product_df_for_export(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Salla Products Template Sheet"

    # Row 1 — merged section header
    ws.cell(1, 1, "بيانات المنتج")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(SALLA_COLS))
    c = ws.cell(1, 1)
    c.font      = Font(bold=True, color="FFFFFF", name="Cairo", size=12)
    c.fill      = PatternFill("solid", fgColor="0F0E0D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2 — column names
    _style_header_row(ws, 2, SALLA_COLS, bg="E8D5B7", fg="B8933A")

    # Data rows from row 3
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(SALLA_COLS, 1):
            v = str(row.get(col, "") if pd.notna(row.get(col, "")) else "")
            c = ws.cell(ri, ci, v)
            c.alignment = Alignment(horizontal="right", vertical="top",
                                    wrap_text=(col == "الوصف"),
                                    readingOrder=2)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FAFAF8")
        ws.row_dimensions[ri].height = 18

    # Column widths
    W = {
        "أسم المنتج": 45, "الوصف": 55, "تصنيف المنتج": 38,
        "صورة المنتج": 46, "الماركة": 24, "No.": 13,
        "وصف صورة المنتج": 36,
    }
    for i, col in enumerate(SALLA_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = W.get(col, 14)
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_product_csv(df: pd.DataFrame) -> bytes:
    if df is None:
        df = pd.DataFrame(columns=SALLA_COLS)
    elif not df.empty:
        df = _prepare_salla_product_df_for_export(df)
    out = io.StringIO()
    # Row 1
    out.write("بيانات المنتج" + "," * (len(SALLA_COLS) - 1) + "\n")
    # Row 2
    out.write(",".join(SALLA_COLS) + "\n")
    for _, row in df.iterrows():
        vals = []
        for c in SALLA_COLS:
            v = str(row.get(c, "") if pd.notna(row.get(c, "")) else "")
            if c == "الوصف":
                v = '"' + v.replace('"', '""') + '"'
            elif any(x in v for x in [",", "\n", '"']):
                v = '"' + v.replace('"', '""') + '"'
            vals.append(v)
        out.write(",".join(vals) + "\n")
    return out.getvalue().encode("utf-8-sig")


def _prepare_salla_seo_df_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """قص حقول SEO لحدود سلة قبل التصدير."""
    df = df.copy()
    tc = "عنوان صفحة المنتج (SEO Page Title)"
    dc = "وصف صفحة المنتج (SEO Page Description)"
    if tc in df.columns:
        df[tc] = df[tc].apply(
            lambda x: (str(x) if pd.notna(x) else "")[:60]
        )
    if dc in df.columns:
        df[dc] = df[dc].apply(
            lambda x: (str(x) if pd.notna(x) else "")[:160]
        )
    return df


def export_seo_xlsx(df: pd.DataFrame) -> bytes:
    if df is not None and not df.empty:
        df = _prepare_salla_seo_df_for_export(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Salla Product Seo Sheet"
    _style_header_row(ws, 1, SALLA_SEO_COLS, bg="1A1510", fg="B8933A")
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(SALLA_SEO_COLS, 1):
            v = str(row.get(col, "") if pd.notna(row.get(col, "")) else "")
            c = ws.cell(ri, ci, v)
            c.alignment = Alignment(horizontal="right", vertical="top",
                                    wrap_text=True, readingOrder=2)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="FFF8E1")
        ws.row_dimensions[ri].height = 18
    W2 = {"اسم المنتج (غير قابل للتعديل)": 50,
          "وصف صفحة المنتج (SEO Page Description)": 65,
          "عنوان صفحة المنتج (SEO Page Title)": 52}
    for i, col in enumerate(SALLA_SEO_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = W2.get(col, 22)
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def export_seo_csv(df: pd.DataFrame) -> bytes:
    if df is not None and not df.empty:
        df = _prepare_salla_seo_df_for_export(df)
    out = io.StringIO()
    out.write(",".join(SALLA_SEO_COLS) + "\n")
    for _, row in df.iterrows():
        vals = []
        for c in SALLA_SEO_COLS:
            v = str(row.get(c, "") if pd.notna(row.get(c, "")) else "")
            if any(x in v for x in [",", "\n"]):
                v = f'"{v}"'
            vals.append(v)
        out.write(",".join(vals) + "\n")
    return out.getvalue().encode("utf-8-sig")

def generate_seo_data_ai(product_name: str, missing_fields: list[str]) -> dict:
    """
    توليد حقول SEO المطلوبة من Claude.
    - missing_fields: قائمة من ["url","title","desc"]
    - يعيد dict مفاتيحه: url_slug / page_title / meta_description (حسب المتوفر)
    """
    empty_out = {}
    api_key = getattr(st.session_state, "api_key", None)
    if not api_key or not HAS_ANTHROPIC:
        return empty_out
    try:
        product_name = str(product_name or "").strip()
        if not product_name:
            return empty_out

        miss = set(missing_fields or [])
        miss = {x for x in miss if x in ("url", "title", "desc")}
        if not miss:
            return empty_out

        rules = []
        if "url" in miss:
            rules.append(
                "SEO Page URL: English lowercase, hyphen-separated, بدون مسافات، بدون سلاش/نقاط."
            )
        if "title" in miss:
            rules.append(
                "SEO Page Title: Arabic, جذاب ويحتوي اسم المنتج + كلمة 'مهووس' أو 'mahwous'، حد أقصى 60 حرفاً."
            )
        if "desc" in miss:
            rules.append(
                "SEO Page Description: Arabic، حد أقصى 160 حرفاً، CTA لشراء المنتج."
            )

        prompt = (
            "أنت خبير SEO لمتجر عطور مهووس.\n"
            "المطلوب: يولّد JSON نقي فقط يحتوي فقط على المفاتيح المطلوبة وفق missing_fields.\n\n"
            f"اسم المنتج: {product_name}\n"
            f"missing_fields: {sorted(list(miss))}\n\n"
            "القواعد:\n"
            + "\n".join(f"- {r}" for r in rules)
            + "\n\n"
            "المفاتيح المتوقعة داخل JSON (استخدم المفاتيح فقط المطابقة لـ missing_fields):\n"
            "- إذا url: url_slug\n"
            "- إذا title: page_title\n"
            "- إذا desc: meta_description\n"
            "\nأعد JSON فقط بدون أي نص إضافي."
        )

        client = anthropic.Anthropic(api_key=api_key)
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=500,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="generate_seo_data_ai")
        if not data:
            return empty_out
        out = {}
        if "url_slug" in data:
            out["url_slug"] = str(data.get("url_slug", "")).strip()
        if "page_title" in data:
            out["page_title"] = str(data.get("page_title", "")).strip()
        if "meta_description" in data:
            out["meta_description"] = str(data.get("meta_description", "")).strip()
        if out:
            t_lim, d_lim = _enforce_salla_product_seo_limits(
                out.get("page_title", ""), out.get("meta_description", "")
            )
            if "page_title" in out:
                out["page_title"] = t_lim
            if "meta_description" in out:
                out["meta_description"] = d_lim
        return out
    except Exception:
        try:
            st.warning(f"تعذر توليد SEO للمنتج: {product_name}")
        except Exception:
            pass
        return empty_out


def render_seo_processor_tab():
    """واجهة معالج SEO — ملتزمة تماماً بملف Salla SEO (5 أعمدة فقط)."""
    st.markdown("## معالج الـ SEO (Salla)")
    st.markdown("""<div class="al-info">
    ارفع ملف منتجات سلة كامل (Excel أو CSV). يكتشف النظام المنتجات التي تفتقد
    <strong>رابط SEO</strong> أو <strong>عنوان الصفحة</strong> أو <strong>وصف الصفحة</strong>،
    ثم يولّدها بالذكاء الاصطناعي بنفس تنسيق ملف <strong>سلة SEO</strong>.
    </div>""", unsafe_allow_html=True)

    up_seo = st.file_uploader(
        "ارفع ملف منتجات سلة SEO",
        type=["csv", "xlsx", "xls", "xlsm"],
        key="seo_proc_tab_uploader",
    )

    if up_seo:
        df_seo = read_file(up_seo, salla_2row=True)
        if df_seo.empty:
            df_seo = read_file(up_seo, salla_2row=False)
        if not df_seo.empty:
            st.session_state.seo_proc_tab_input_df = df_seo
            st.success(f"✅ تم تحميل {len(df_seo):,} صف")
            st.rerun()

    input_df = getattr(st.session_state, "seo_proc_tab_input_df", None)
    if input_df is None:
        return

    sdf = input_df
    name_col = "اسم المنتج (غير قابل للتعديل)"
    url_col = "رابط مخصص للمنتج (SEO Page URL)"
    title_col = "عنوان صفحة المنتج (SEO Page Title)"
    desc_col = "وصف صفحة المنتج (SEO Page Description)"

    for col in (name_col, url_col, title_col, desc_col):
        if col not in sdf.columns:
            st.error(f"الملف لا يحتوي العمود المطلوب: {col}")
            return

    with st.expander("👀 معاينة الملف", expanded=False):
        st.dataframe(sdf.head(12), use_container_width=True)

    if st.button("🚀 بدء المعالجة", type="primary", key="seo_proc_tab_run", use_container_width=True):
        if not st.session_state.api_key:
            st.error("أضف مفتاح Anthropic API من صفحة الإعدادات.")
            st.stop()

        work = sdf.copy()
        def cell_is_empty(v) -> bool:
            s = str(v or "").strip()
            if not s:
                return True
            return s.lower() in ("nan", "none")

        prog = st.progress(0, text="جاري معالجة الصفوف...")
        total = max(len(work), 1)

        for i in range(len(work)):
            row_dict = work.iloc[i].to_dict()
            prod_name = str(row_dict.get(name_col, "") or "").strip()
            if not prod_name or prod_name.lower() in ("nan", "none"):
                continue

            missing_fields: list[str] = []
            if cell_is_empty(row_dict.get(url_col, "")):
                missing_fields.append("url")
            if cell_is_empty(row_dict.get(title_col, "")):
                missing_fields.append("title")
            if cell_is_empty(row_dict.get(desc_col, "")):
                missing_fields.append("desc")

            if missing_fields:
                out = generate_seo_data_ai(prod_name, missing_fields)
                if "url_slug" in out and "url" in missing_fields:
                    work.at[i, url_col] = out["url_slug"]
                if "page_title" in out and "title" in missing_fields:
                    work.at[i, title_col] = out["page_title"]
                if "meta_description" in out and "desc" in missing_fields:
                    work.at[i, desc_col] = out["meta_description"]

            prog.progress(int((i + 1) / total * 100), text=f"معالجة: {i+1}/{total}")

        st.session_state.seo_proc_tab_output_df = work[SALLA_SEO_COLS].copy()
        st.success("✅ تم تحديث ملف SEO بنجاح")
        st.rerun()

    output_df = getattr(st.session_state, "seo_proc_tab_output_df", None)
    if output_df is not None and not output_df.empty:
        st.markdown("""<div class="sec-title"><div class="bar"></div>
        <h3>نتيجة التوليد (تنسيق سلة SEO)</h3></div>""", unsafe_allow_html=True)
        edited_df = st.data_editor(
            output_df.fillna(""),
            use_container_width=True,
            num_rows="dynamic",
            key="seo_proc_tab_editor",
        )
        st.session_state.seo_proc_tab_output_df = edited_df

        csv_str = edited_df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            "📥 تصدير SEO — CSV (محدث)",
            csv_str,
            "مهووس_SEO_محدث.csv",
            "text/csv",
            use_container_width=True,
            key="seo_proc_tab_dl",
        )


def export_price_xlsx(df: pd.DataFrame) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Price Update"
    ws.cell(1, 1, "بيانات المنتج")
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(SALLA_PRICE_COLS))
    c = ws.cell(1, 1)
    c.font      = Font(bold=True, color="FFFFFF", name="Cairo")
    c.fill      = PatternFill("solid", fgColor="0F0E0D")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    _style_header_row(ws, 2, SALLA_PRICE_COLS, bg="E8D5B7", fg="B8933A")
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, col in enumerate(SALLA_PRICE_COLS, 1):
            ws.cell(ri, ci, str(row.get(col, "") or ""))
        ws.row_dimensions[ri].height = 18
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def export_price_csv(df: pd.DataFrame) -> bytes:
    out = io.StringIO()
    out.write("بيانات المنتج" + "," * (len(SALLA_PRICE_COLS) - 1) + "\n")
    out.write(",".join(SALLA_PRICE_COLS) + "\n")
    for _, row in df.iterrows():
        out.write(",".join([f'"{str(row.get(c,"") or "")}"'
                            for c in SALLA_PRICE_COLS]) + "\n")
    return out.getvalue().encode("utf-8-sig")


def export_brands_xlsx(brands_list: list) -> bytes:
    """Export new brands in Salla brands file format."""
    wb = Workbook(); ws = wb.active; ws.title = "New Brands"
    _style_header_row(ws, 1, SALLA_BRANDS_COLS, bg="0F0E0D", fg="B8933A")
    for ri, brand in enumerate(brands_list, 2):
        for ci, col in enumerate(SALLA_BRANDS_COLS, 1):
            ws.cell(ri, ci, str(brand.get(col, "") or ""))
        ws.row_dimensions[ri].height = 18
    for i, col in enumerate(SALLA_BRANDS_COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = 40 if i == 3 else 28
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  SIDEBAR NAVIGATION                                             ║
# ╚══════════════════════════════════════════════════════════════════╝
with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:18px 0 10px">
      <div style="font-size:2.4rem">🌸</div>
      <div style="color:#b8933a;font-size:1.25rem;font-weight:900;margin:4px 0">مهووس</div>
      <div style="color:rgba(255,255,255,0.3);font-size:0.7rem">مركز التحكم الشامل v11.0</div>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    PAGES = [
        ("🚀", "المسار الآلي",           "pipeline"),
        ("🔀", "المقارنة",              "compare"),
        ("🏪", "مدقق ملف المتجر",       "store_audit"),
        ("🔍", "معالج الـ SEO",          "seo_processor"),
        ("➕", "منتج سريع",              "quickadd"),
        ("⚙️", "الإعدادات",             "settings"),
    ]
    for icon, label, key in PAGES:
        active = st.session_state.page == key
        if st.button(f"{icon}  {label}", width='stretch',
                     type="primary" if active else "secondary",
                     key=f"nav_{key}"):
            st.session_state.page = key
            st.rerun()

    st.divider()
    # Status
    bok = st.session_state.brands_df is not None
    cok = st.session_state.categories_df is not None
    aok = bool(st.session_state.api_key)
    gok = bool(st.session_state.google_api and st.session_state.google_cse)

    status_html = "".join([
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if bok else "❌"} '
        f'الماركات: {len(st.session_state.brands_df) if bok else "غير محملة"}</div>',
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if cok else "❌"} '
        f'التصنيفات: {len(st.session_state.categories_df) if cok else "غير محملة"}</div>',
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if aok else "❌"} '
        f'Claude API: {"متصل" if aok else "غير مضبوط"}</div>',
        f'<div style="font-size:0.77rem;padding:3px 0">{"✅" if gok else "—"} '
        f'Google CSE: {"متصل" if gok else "غير مضبوط"}</div>',
    ])
    st.markdown(status_html, unsafe_allow_html=True)

    st.divider()
    st.markdown("**💾 حماية البيانات (Railway Backup)**")
    bkp_col1, bkp_col2 = st.columns(2)
    backup_path = os.path.join(DATA_DIR, "session_backup.pkl")
    with bkp_col1:
        if st.button("💾 حفظ الجلسة", use_container_width=True, key="bkp_save"):
            try:
                save_dict = {k: v for k, v in st.session_state.items()
                             if k in ["up_df","up_seo","qa_rows","new_brands",
                                      "up_filename","pipe_store_df","cmp_new_df"]}
                import pickle as _pickle
                os.makedirs(DATA_DIR, exist_ok=True)
                with open(backup_path, "wb") as _f:
                    _pickle.dump(save_dict, _f)
                st.toast("✅ تم الحفظ بنجاح")
            except Exception as _e:
                st.toast(f"❌ فشل الحفظ: {_e}")
    with bkp_col2:
        if st.button("📂 استعادة", use_container_width=True, key="bkp_load"):
            try:
                import pickle as _pickle
                if os.path.exists(backup_path):
                    with open(backup_path, "rb") as _f:
                        loaded = _pickle.load(_f)
                    for k, v in loaded.items():
                        st.session_state[k] = v
                    st.toast("✅ تمت الاستعادة")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.toast("⚠️ لا توجد جلسة محفوظة")
            except Exception:
                st.toast("❌ فشل الاستعادة")

    # New brands indicator
    if st.session_state.new_brands:
        st.divider()
        nb = len(st.session_state.new_brands)
        st.markdown(f'<div style="font-size:0.77rem;padding:3px 0;color:#f9a825">🆕 {nb} ماركة جديدة بانتظار التصدير</div>',
                    unsafe_allow_html=True)

    # Active file info
    if st.session_state.up_df is not None:
        st.divider()
        fname = st.session_state.up_filename
        nrows = len(st.session_state.up_df)
        st.markdown(f"""
        <div style="background:rgba(184,147,58,0.1);border-radius:8px;padding:10px;
                    font-size:0.78rem;border:1px solid rgba(184,147,58,0.25)">
          <div style="font-weight:800;margin-bottom:4px">📄 الملف النشط</div>
          <div style="color:#b8933a">{fname[:30]}</div>
          <div style="color:rgba(255,255,255,0.4)">{nrows} صف</div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("🗑️ إغلاق الملف", width='stretch'):
            st.session_state.up_raw     = None
            st.session_state.up_df      = None
            st.session_state.up_seo     = None
            st.session_state.up_filename = ""
            st.session_state.up_mapped  = False
            st.rerun()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE HEADER                                                    ║
# ╚══════════════════════════════════════════════════════════════════╝
TITLES = {
    "pipeline":      ("🚀 المسار الآلي",            "مقارنة → فلترة AI → معالجة → جدول تفاعلي → تصدير منتج جديد.csv"),
    "seo_processor": ("🔍 معالج الـ SEO",           "توليد روابط وعناوين وأوصاف SEO بتنسيق سلة — ذكاء اصطناعي"),
    "compare":       ("🔀 المقارنة",                "مقارنة المنافسين مع المتجر واعتماد النتائج بصرياً"),
    "store_audit":   ("🏪 مدقق ملف المتجر",         "فحص ملف المتجر — اكتشاف النواقص — إصلاح وتصدير بتنسيق سلة"),
    "quickadd":      ("➕ منتج سريع",              "أدخل رابط منتج أو ارفع صورة وسيكمل النظام الباقي"),
    "settings":      ("⚙️ الإعدادات",             "مفاتيح API وقواعد البيانات المرجعية"),
}
ttl, sub = TITLES.get(st.session_state.page, ("مهووس", ""))
st.markdown(f"""
<div class="mhw-header">
  <div class="emblem">م</div>
  <div><h1>{ttl}</h1><p>{sub}</p></div>
</div>
""", unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE — COMPARE & AUDIT (helpers)                               ║
# ╚══════════════════════════════════════════════════════════════════╝
def _pick_first_non_http_text_col(df: pd.DataFrame) -> str:
    """Fallback: pick first text column that is not mostly URLs."""
    txt_cols = [c for c in df.columns if pd.api.types.is_object_dtype(df[c]) or pd.api.types.is_string_dtype(df[c])]
    for c in txt_cols:
        s = df[c].fillna("").astype(str).str.strip()
        nz = s[s != ""]
        if nz.empty:
            continue
        http_ratio = nz.str.contains(r"https?://", case=False, regex=True).mean()
        if http_ratio < 0.5:
            return c
    return txt_cols[0] if txt_cols else ""


def _guess_competitor_name_col(df: pd.DataFrame) -> str:
    """Smart guessing for raw scraped competitor files."""
    if "أسم المنتج" in df.columns:
        return "أسم المنتج"
    g = auto_guess_col(
        df.columns,
        ["أسم المنتج", "اسم المنتج", "name", "title", "product", "styles_productcard", "styles_productCard"],
        df,
    )
    if g and g != "— لا يوجد —":
        return g
    # fallback: first text-like col not dominated by URLs
    return _pick_first_non_http_text_col(df)


def _guess_competitor_price_col(df: pd.DataFrame) -> str:
    if "سعر المنتج" in df.columns:
        return "سعر المنتج"
    g = auto_guess_col(
        df.columns,
        ["سعر المنتج", "السعر", "price", "text-sm", "amount", "value"],
        df,
    )
    if g and g != "— لا يوجد —":
        return g
    return ""


def _read_competitor_file(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Normalize competitor file columns into app-friendly names.
    - Map guessed name col -> 'أسم المنتج'
    - Map guessed price col -> 'سعر المنتج'
    """
    if df_raw is None or df_raw.empty:
        return pd.DataFrame(), {"name_col": "", "price_col": "", "rows_before": 0, "rows_after": 0}

    df = df_raw.copy()
    name_col = _guess_competitor_name_col(df)
    price_col = _guess_competitor_price_col(df)

    # rename safely without overriding an existing canonical column
    if name_col and name_col != "أسم المنتج" and "أسم المنتج" not in df.columns:
        df = df.rename(columns={name_col: "أسم المنتج"})
    if price_col and price_col != "سعر المنتج" and "سعر المنتج" not in df.columns:
        df = df.rename(columns={price_col: "سعر المنتج"})

    # last-resort name col
    if "أسم المنتج" not in df.columns:
        fb = _pick_first_non_http_text_col(df)
        if fb:
            df = df.rename(columns={fb: "أسم المنتج"})
            name_col = fb

    # remove fully-empty rows
    df = df.dropna(how="all").reset_index(drop=True)
    return df, {
        "name_col": "أسم المنتج" if "أسم المنتج" in df.columns else name_col,
        "price_col": "سعر المنتج" if "سعر المنتج" in df.columns else price_col,
        "rows_before": len(df_raw),
        "rows_after": len(df),
    }


def apply_competitor_exclusions(
    df: pd.DataFrame,
    name_col: str = "أسم المنتج",
    *,
    exclude_samples: bool = True,
    exclude_makeup: bool = True,
    exclude_accessories: bool = True,
    exclude_missing_sizes: bool = True,
) -> tuple[pd.DataFrame, dict]:
    """Apply exclusions ONLY on competitor data before comparison."""
    if df is None or df.empty or name_col not in df.columns:
        return df, {"input_rows": 0, "output_rows": 0}

    work = df.copy()
    names = work[name_col].fillna("").astype(str)
    nl = names.str.lower()
    is_tester = nl.str.contains(r"(?:\btester\b|تستر|تيستر)", regex=True, na=False)

    sample_kw = r"(?:عينة|سمبل|فايال|\bvial\b|\bsample\b)"
    sample_size = r"\b[1-8]\s*(?:مل|ml)\b"
    makeup_kw = r"(?:مكياج|ارواج|روج|ماسكارا|فاونديشن|ظل|بلشر|كونسيلر)"
    acc_kw = r"(?:شنطة|كيس|تغليف|كرتون|مبخرة|فحم|ميدالية)"
    has_size = nl.str.contains(r"(?:مل|ml|oz|غرام|جرام)", regex=True, na=False)

    drop_sample = (nl.str.contains(sample_kw, regex=True, na=False) | nl.str.contains(sample_size, regex=True, na=False)) & (~is_tester)
    drop_makeup = nl.str.contains(makeup_kw, regex=True, na=False)
    drop_acc = nl.str.contains(acc_kw, regex=True, na=False)
    drop_no_size = ~has_size

    mask = pd.Series(False, index=work.index)
    stats = {"input_rows": len(work), "dropped_samples": 0, "dropped_makeup": 0, "dropped_accessories": 0, "dropped_missing_sizes": 0}

    if exclude_samples:
        mask |= drop_sample
        stats["dropped_samples"] = int(drop_sample.sum())
    if exclude_makeup:
        mask |= drop_makeup
        stats["dropped_makeup"] = int(drop_makeup.sum())
    if exclude_accessories:
        mask |= drop_acc
        stats["dropped_accessories"] = int(drop_acc.sum())
    if exclude_missing_sizes:
        mask |= drop_no_size
        stats["dropped_missing_sizes"] = int(drop_no_size.sum())

    out = work.loc[~mask].copy()
    stats["output_rows"] = len(out)
    return out, stats


def render_compare_tab():

        st.markdown("""<div class="al-info">
        قارن ملف المنتجات الجديدة بملف المتجر الأساسي. تُعرض المنتجات <strong>المشبوهة</strong> في كروت
        بصورتين للمراجعة السريعة. يمكنك <b>اعتماد</b> المنتج كجديد، أو <b>تجاهله</b> كمكرر، أو <b>تعديل</b> الاسم يدوياً.
        </div>""", unsafe_allow_html=True)

        if st.session_state.get("cmp_from_pipe") and st.session_state.cmp_new_df is not None:
            st.markdown('<div class="al-ok">📎 تم تحميل بيانات المسار الآلي (منتجات معتمدة + ملف المتجر).</div>',
                        unsafe_allow_html=True)

        c_up1, c_up2 = st.columns(2)
        with c_up1:
            st.markdown("**ملف المنتجات الجديدة**")
            up_n = st.file_uploader("CSV / Excel", type=["csv", "xlsx", "xls"], key="cmp_up_new", label_visibility="collapsed")
            if up_n:
                dfn = read_file(up_n, salla_2row=True)
                if dfn.empty:
                    dfn = read_file(up_n, salla_2row=False)
                if not dfn.empty:
                    dfn_norm, map_info = _read_competitor_file(dfn)
                    st.session_state.cmp_new_df = dfn_norm
                    st.session_state.cmp_from_pipe = False
                    st.success(f"✅ {len(dfn_norm)} منتج")
                    if map_info.get("name_col") != "أسم المنتج":
                        st.caption(f"تم تعيين عمود الاسم تلقائياً من: `{map_info.get('name_col', '')}`")
                    if map_info.get("price_col") and map_info.get("price_col") != "سعر المنتج":
                        st.caption(f"تم تعيين عمود السعر تلقائياً من: `{map_info.get('price_col', '')}`")
        with c_up2:
            st.markdown("**ملف المتجر الأساسي**")
            up_s = st.file_uploader("CSV / Excel", type=["csv", "xlsx", "xls"], key="cmp_up_store", label_visibility="collapsed")
            if up_s:
                dfs = read_file(up_s, salla_2row=True)
                if dfs.empty:
                    dfs = read_file(up_s, salla_2row=False)
                if not dfs.empty:
                    st.session_state.cmp_store_df = dfs
                    st.success(f"✅ {len(dfs)} منتج في المتجر")

        if st.session_state.cmp_new_df is not None and st.session_state.cmp_store_df is not None:
            new_df = st.session_state.cmp_new_df
            store_df = st.session_state.cmp_store_df
            NONE_C = "— لا يوجد —"
            n_opts = [NONE_C] + list(new_df.columns)
            s_opts = [NONE_C] + list(store_df.columns)

            def _gi(cols, kws, df_, opts):
                g = auto_guess_col(cols, kws, df_)
                return opts.index(g) if g in opts else 0

            r1, r2, r3, r4 = st.columns(4)
            with r1:
                new_nm = st.selectbox("عمود الاسم (جديد):", n_opts,
                                      index=_gi(new_df.columns, ["اسم", "name", "منتج"], new_df, n_opts), key="cmp_nn")
            with r2:
                new_sk = st.selectbox("عمود SKU (جديد):", n_opts,
                                        index=_gi(new_df.columns, ["sku", "رمز"], new_df, n_opts), key="cmp_nsk")
            with r3:
                st_nm = st.selectbox("عمود الاسم (متجر):", s_opts,
                                     index=_gi(store_df.columns, ["اسم", "name", "منتج"], store_df, s_opts), key="cmp_sn")
            with r4:
                st_sk = st.selectbox("عمود SKU (متجر):", s_opts,
                                     index=_gi(store_df.columns, ["sku", "رمز"], store_df, s_opts), key="cmp_ssk")

            new_img_g = auto_guess_col(new_df.columns, ["صورة", "image", "src"], new_df)
            new_img_g = None if new_img_g == "— لا يوجد —" else new_img_g
            sim_thr = st.slider("عتبة التشابه للمشبوه (%):", 50, 95, 75, key="cmp_sim")
            st.markdown("**فلاتر استبعاد المنافسين (تُطبق على ملف المنافس فقط):**")
            fx1, fx2, fx3, fx4 = st.columns(4)
            with fx1:
                cmp_fx_samples = st.checkbox("استبعاد العينات", value=True, key="cmp_fx_samples")
            with fx2:
                cmp_fx_makeup = st.checkbox("استبعاد المكياج", value=True, key="cmp_fx_makeup")
            with fx3:
                cmp_fx_accessories = st.checkbox("استبعاد الكماليات", value=True, key="cmp_fx_accessories")
            with fx4:
                cmp_fx_missing_size = st.checkbox("استبعاد منتجات بدون حجم", value=True, key="cmp_fx_missing_size")

            if st.button("🔍 تشغيل المقارنة والعرض المرئي", type="primary", key="cmp_run", use_container_width=True):
                if new_nm == NONE_C or st_nm == NONE_C:
                    st.error("حدد عمود الاسم للملفين.")
                else:
                    comp_for_compare = new_df.copy()
                    if "أسم المنتج" not in comp_for_compare.columns and new_nm in comp_for_compare.columns:
                        comp_for_compare = comp_for_compare.rename(columns={new_nm: "أسم المنتج"})
                        new_nm = "أسم المنتج"
                    comp_for_compare, ex_stats = apply_competitor_exclusions(
                        comp_for_compare,
                        name_col="أسم المنتج" if "أسم المنتج" in comp_for_compare.columns else new_nm,
                        exclude_samples=cmp_fx_samples,
                        exclude_makeup=cmp_fx_makeup,
                        exclude_accessories=cmp_fx_accessories,
                        exclude_missing_sizes=cmp_fx_missing_size,
                    )
                    if comp_for_compare.empty:
                        st.error("لا توجد منتجات صالحة بعد تطبيق فلاتر الاستبعاد على ملف المنافس.")
                        st.stop()
                    brands_l = []
                    if st.session_state.brands_df is not None:
                        brands_l = (st.session_state.brands_df[st.session_state.brands_df.columns[0]]
                                    .dropna().astype(str).str.strip().tolist())
                    with st.spinner("جاري المقارنة..."):
                        res_df = run_smart_comparison(
                            new_df=comp_for_compare,
                            store_df=store_df,
                            new_name_col=new_nm,
                            store_name_col=st_nm,
                            new_sku_col=new_sk if new_sk != NONE_C else None,
                            store_sku_col=st_sk if st_sk != NONE_C else None,
                            new_img_col=new_img_g,
                            t_dup=88, t_near=sim_thr, t_review=50,
                            brands_list=brands_l,
                        )
                        st.session_state.cmp_results = res_df
                        st.session_state.cmp_cfg = {
                            "new_nm": new_nm, "st_nm": st_nm, "new_img": new_img_g,
                            "cmp_ex_stats": ex_stats,
                        }
                        sus = res_df[res_df["الحالة"] == "مشبوه"]
                        st.session_state.cmp_approved = {
                            int(r["_idx"]): True for _, r in sus.iterrows()
                        }
                    st.rerun()

        if st.session_state.cmp_results is not None:
            res = st.session_state.cmp_results
            new_df = st.session_state.cmp_new_df
            store_df = st.session_state.cmp_store_df
            cfg = st.session_state.get("cmp_cfg", {})
            st_nm = cfg.get("st_nm", "أسم المنتج" if "أسم المنتج" in store_df.columns else store_df.columns[0])
            new_img_col = cfg.get("new_img")
            s_img_col = auto_guess_col(store_df.columns, ["صورة", "image", "src"], store_df)
            if s_img_col == "— لا يوجد —":
                s_img_col = None

            exact_dup = res[res["الحالة"].astype(str).str.contains("مكرر", na=False)]
            suspect = res[res["الحالة"] == "مشبوه"]
            new_clean = res[res["الحالة"] == "جديد"]

            st.markdown(f"""
            <div class="stats-bar">
              <div class="stat-box"><div class="n">{len(res)}</div><div class="lb">إجمالي</div></div>
              <div class="stat-box"><div class="n" style="color:#e53935">{len(exact_dup)}</div><div class="lb">مكرر</div></div>
              <div class="stat-box"><div class="n" style="color:#f9a825">{len(suspect)}</div><div class="lb">مشبوه</div></div>
              <div class="stat-box"><div class="n" style="color:#43a047">{len(new_clean)}</div><div class="lb">جديد</div></div>
            </div>
            """, unsafe_allow_html=True)
            exs = cfg.get("cmp_ex_stats", {})
            if exs:
                st.caption(
                    f"فلترة المنافسين قبل المقارنة: دخل {exs.get('input_rows', 0)} → خرج {exs.get('output_rows', 0)} | "
                    f"عينات: {exs.get('dropped_samples', 0)}، مكياج: {exs.get('dropped_makeup', 0)}، "
                    f"كماليات: {exs.get('dropped_accessories', 0)}، بدون حجم: {exs.get('dropped_missing_sizes', 0)}"
                )

            if not suspect.empty:
                st.markdown("""<div class="sec-title"><div class="bar"></div>
                <h3>منتجات مشبوهة — مقارنة بصرية</h3></div>""", unsafe_allow_html=True)
                for _, srow in suspect.iterrows():
                    idx = int(srow["_idx"])
                    new_img_u = str(srow.get("_img", "") or "").split(",")[0].strip().replace(" ", "%20")
                    if not new_img_u.startswith("http"):
                        new_img_u = ""
                    store_match = str(srow.get("أقرب تطابق في المتجر", "") or "")
                    store_img_u = ""
                    if store_match and st_nm in store_df.columns:
                        try:
                            sm = store_df[store_df[st_nm].astype(str) == store_match]
                            if not sm.empty and s_img_col and s_img_col in store_df.columns:
                                store_img_u = str(sm.iloc[0].get(s_img_col, "") or "").split(",")[0].strip().replace(" ", "%20")
                                if not store_img_u.startswith("http"):
                                    store_img_u = ""
                        except Exception:
                            pass

                    ph = "width:120px;height:120px;background:#eee;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:2rem"
                    img_new = f'<img src="{new_img_u}" style="width:120px;height:120px;object-fit:cover;border-radius:10px" onerror="this.style.display=\'none\'">' if new_img_u else f'<div style="{ph}">🖼</div>'
                    img_st = f'<img src="{store_img_u}" style="width:120px;height:120px;object-fit:cover;border-radius:10px" onerror="this.style.display=\'none\'">' if store_img_u else f'<div style="{ph}">🏪</div>'

                    st.markdown(f'<div class="cmp-card suspect">', unsafe_allow_html=True)
                    h1, h2, h3 = st.columns([2, 2, 1])
                    with h1:
                        st.markdown("**🆕 جديد**", unsafe_allow_html=True)
                        st.markdown(img_new, unsafe_allow_html=True)
                        st.caption(srow.get("الاسم الجديد", ""))
                    with h2:
                        st.markdown("**🏪 المتجر**", unsafe_allow_html=True)
                        st.markdown(img_st, unsafe_allow_html=True)
                        st.caption(store_match or "—")
                    with h3:
                        st.metric("تشابه", f"{srow.get('نسبة التشابه', 0)}%")

                    _ek = f"cmp_edit_{idx}"
                    if _ek not in st.session_state:
                        st.session_state[_ek] = str(
                            st.session_state.cmp_edit_name.get(idx, srow.get("الاسم الجديد", ""))
                        )
                    st.text_input("تعديل الاسم إن لزم", key=_ek)
                    st.session_state.cmp_edit_name[idx] = st.session_state[_ek]

                    b1, b2, b3 = st.columns(3)
                    with b1:
                        if st.button("✅ اعتماد كجديد", key=f"cmp_ok_{idx}", use_container_width=True):
                            st.session_state.cmp_approved[idx] = True
                            st.rerun()
                    with b2:
                        if st.button("⛔ تجاهل (مكرر)", key=f"cmp_no_{idx}", use_container_width=True):
                            st.session_state.cmp_approved[idx] = False
                            st.rerun()
                    with b3:
                        if st.button("💾 حفظ التعديل على الاسم", key=f"cmp_sv_{idx}", use_container_width=True):
                            st.toast("تم حفظ الاسم في المعاينة — اضغط اعتماد أو تجاهل")
                    st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
            <h3>تصدير القائمة بعد القرار</h3></div>""", unsafe_allow_html=True)
            if st.button("⚡ بناء ملف المنتجات النهائي", type="primary", key="cmp_build"):
                new_src = st.session_state.cmp_new_df
                rows_out = []
                for _, rrow in res.iterrows():
                    idx = int(rrow["_idx"])
                    stt = str(rrow["الحالة"])
                    if stt == "جديد":
                        if idx in new_src.index:
                            row_ser = new_src.loc[idx].copy()
                            if "الماركة" in rrow.index:
                                row_ser["الماركة"] = rrow.get("الماركة", "")
                            rows_out.append(row_ser)
                    elif stt == "مشبوه":
                        ap = st.session_state.cmp_approved.get(idx, True)
                        if ap and idx in new_src.index:
                            row_ser = new_src.loc[idx].copy()
                            if idx in st.session_state.cmp_edit_name:
                                if "أسم المنتج" in row_ser.index:
                                    row_ser["أسم المنتج"] = st.session_state.cmp_edit_name[idx]
                            if "الماركة" in rrow.index:
                                row_ser["الماركة"] = rrow.get("الماركة", "")
                            rows_out.append(row_ser)
                if rows_out:
                    final_cmp = pd.DataFrame(rows_out)
                    for col in SALLA_COLS:
                        if col not in final_cmp.columns:
                            final_cmp[col] = ""
                    final_cmp = final_cmp[[c for c in SALLA_COLS if c in final_cmp.columns]]
                    # Salla: النظامية (System fields) يجب أن تكون حرفياً لكل صف
                    if "النوع " in final_cmp.columns:
                        final_cmp["النوع "] = "منتج"
                    if "نوع المنتج" in final_cmp.columns:
                        final_cmp["نوع المنتج"] = "منتج جاهز"
                    if "الكمية المتوفرة" in final_cmp.columns:
                        final_cmp["الكمية المتوفرة"] = "0"
                    if "الماركة" in final_cmp.columns:
                        final_cmp["الماركة"] = final_cmp["الماركة"].apply(_clean_brand_value_for_salla_output)
                    st.session_state.cmp_export_df = final_cmp
                    st.success(f"✅ {len(final_cmp)} منتج في القائمة النهائية")
                else:
                    st.warning("لا توجد صفوف معتمدة.")

            if getattr(st.session_state, "cmp_export_df", None) is not None:
                fe = datetime.now().strftime("%Y%m%d_%H%M")
                st.download_button(
                    "📥 تنزيل Excel",
                    export_product_xlsx(st.session_state.cmp_export_df),
                    f"mahwous_after_compare_{fe}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="cmp_dl_x",
                )

            if st.button("🔄 إعادة ضبط المقارنة", key="cmp_reset"):
                st.session_state.cmp_results = None
                st.session_state.cmp_approved = {}
                st.session_state.cmp_edit_name = {}
                if hasattr(st.session_state, "cmp_export_df"):
                    del st.session_state["cmp_export_df"]
                for _k in list(st.session_state.keys()):
                    if isinstance(_k, str) and _k.startswith("cmp_edit_"):
                        del st.session_state[_k]
                st.rerun()

        elif st.session_state.cmp_new_df is None or st.session_state.cmp_store_df is None:
            st.markdown("""<div class="upload-zone"><div class="uz-icon">🔀</div>
            <div class="uz-title">ارفع ملف المنتجات الجديدة وملف المتجر</div>
            <div class="uz-sub">أو أكمل المسار الآلي واضغط زر الانتقال للمقارنة المرئي</div>
            </div>""", unsafe_allow_html=True)
def _normalize_simple_brand_token(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


_BRAND_STOP_WORDS = [
    "عطر", "تستر", "تيستر", "طقم", "مجموعة", "او دو بارفيوم", "او دو تواليت",
    "أو دو بارفيوم", "أو دو تواليت",
    "مل", "بديل", "تعبئة", "عينة", "tester", "parfum", "edt", "edp",
    "eau", "de", "parfum", "toilette", "ml", "oz",
]


def _clean_product_name_for_brand_search(name: str) -> str:
    """تنظيف اسم المنتج من كلمات/أنماط غير مفيدة قبل استخراج/مطابقة الماركة."""
    if not name:
        return ""
    txt = str(name).lower()
    txt = strip_trailing_discount_label(txt)
    txt = re.sub(r"\d+(?:[.,]\d+)?\s*(?:مل|ml|oz|غرام|جرام)\b", " ", txt, flags=re.IGNORECASE)
    # نزيل كلمات التوقّف بدون الاعتماد على `\b` لأن حدود الكلمات العربية عبر صياغات/علامات الترقيم قد تفشل،
    # وهذا يؤدي لظهور "تستر ..." أو "عطر ..." كجزء من اسم الماركة.
    for sw in _BRAND_STOP_WORDS:
        sw_l = str(sw).lower().strip()
        if not sw_l:
            continue
        txt = re.sub(re.escape(sw_l), " ", txt, flags=re.IGNORECASE)
    txt = re.sub(r"[^\w\u0600-\u06FF\s]", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _strip_brand_entity_stopwords(text: str) -> str:
    """إزالة كلمات التوقف قبل استخراج/عرض اسم العلامة (تستر، عطر، بخاخ، …)."""
    t = str(text or "").strip()
    if not t:
        return ""
    for sw in _BRAND_STOP_WORDS:
        sw_l = str(sw).lower().strip()
        if sw_l:
            t = re.sub(re.escape(sw_l), " ", t, flags=re.IGNORECASE)
    for sw in ("بخاخ", "بخاخات", "مجموعه", "معطر جسم", "معطر شعر"):
        t = re.sub(re.escape(sw), " ", t, flags=re.IGNORECASE)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _extract_brand_entity_with_ai(product_name: str, api_key: str) -> str:
    """استخراج اسم العلامة فقط عبر Claude؛ يعيد الاسم أو 'Unknown' عند عدم اليقين."""
    if not product_name or not api_key or not HAS_ANTHROPIC:
        return ""
    try:
        cache = st.session_state.setdefault("_brand_entity_ai_cache", {})
        ck = str(product_name).strip()[:2000]
        if ck in cache:
            return cache[ck]
        pre = _strip_brand_entity_stopwords(
            _clean_product_name_for_brand_search(product_name)
        )
        if not pre.strip():
            cache[ck] = "Unknown"
            return "Unknown"
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            "You extract ONLY the official fragrance house / brand name from a product title.\n"
            "Return a single JSON object only: {\"brand\":\"...\"}\n"
            "Rules:\n"
            "- brand = the house name (use established English spelling when the house is global).\n"
            "- Never return generic product words (perfume, tester, set, spray, deodorant, body mist).\n"
            "- Do NOT mistranslate English brand names into Arabic "
            "(e.g. 'The Merchant of Venice' must stay English — never return 'تاجر' alone).\n"
            "- If you cannot name the brand confidently, return exactly: Unknown\n\n"
            f"Product title:\n{pre}\n"
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=120,
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="extract_brand_entity")
        b = str(data.get("brand", "") or "").strip()
        if not b or b.lower() == "unknown":
            cache[ck] = "Unknown"
            return "Unknown"
        b = _strip_brand_entity_stopwords(b)
        b = _normalize_simple_brand_token(b)
        if not b:
            cache[ck] = "Unknown"
            return "Unknown"
        cache[ck] = b
        return b
    except Exception:
        return "Unknown"


def _fuzzy_ratio_local(a: str, b: str) -> float:
    a = _normalize_simple_brand_token(a).lower()
    b = _normalize_simple_brand_token(b).lower()
    if not a or not b:
        return 0.0
    if HAS_RAPIDFUZZ:
        try:
            return float(rf_fuzz.token_sort_ratio(a, b))
        except Exception:
            pass
    aw = set(a.split())
    bw = set(b.split())
    if not aw or not bw:
        return 0.0
    return 100.0 * len(aw & bw) / len(aw | bw)


def _collect_known_brand_names(pipe_store_df: pd.DataFrame, brands_df: pd.DataFrame) -> list:
    names = []
    def _push_variants(raw: str):
        raw = str(raw or "").strip()
        if not raw:
            return
        # split bilingual / aliased strings: "جيفنشي | Givenchy", "A / B"
        parts = [p.strip() for p in re.split(r"\s*[|/]\s*", raw) if p.strip()]
        if not parts:
            parts = [raw]
        for p in parts:
            p = _normalize_simple_brand_token(p)
            if p:
                names.append(p)

    if brands_df is not None and not brands_df.empty:
        bcol = brands_df.columns[0]
        for v in brands_df[bcol].dropna().astype(str):
            _push_variants(v)
    if pipe_store_df is not None and not pipe_store_df.empty:
        # detect brand column even if app guessed a different header name
        store_brand_col = (
            "الماركة" if "الماركة" in pipe_store_df.columns else
            auto_guess_col(pipe_store_df.columns, ["الماركة", "ماركة", "brand", "علامة"], pipe_store_df)
        )
        if store_brand_col and store_brand_col != "— لا يوجد —" and store_brand_col in pipe_store_df.columns:
            for v in pipe_store_df[store_brand_col].dropna().astype(str):
                _push_variants(v)
    out = []
    seen = set()
    for n in names:
        k = n.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(n)
    return out


def _build_missing_brands_df_from_competitors(
    comp_df: pd.DataFrame,
    name_col: str,
    known_brands: list,
    threshold: float = 85.0,
    api_key: str = "",
) -> pd.DataFrame:
    if comp_df is None or comp_df.empty or name_col not in comp_df.columns:
        return pd.DataFrame(columns=SALLA_BRANDS_COLS)
    if not api_key or not HAS_ANTHROPIC:
        APP_LOG.warning("missing brands: AI brand extraction skipped (no API key)")
        return pd.DataFrame(columns=SALLA_BRANDS_COLS)
    rows = []
    seen = set()
    known_sorted = sorted(
        [k for k in known_brands if str(k).strip()],
        key=lambda x: len(str(x)),
        reverse=True,
    )
    for nm in comp_df[name_col].fillna("").astype(str):
        cleaned = _clean_product_name_for_brand_search(nm)
        if not cleaned:
            continue
        cleaned_n = normalize_brand_name_v2(cleaned)

        matched_known = False
        for kb in known_sorted:
            kb_n = normalize_brand_name_v2(str(kb))
            if len(kb_n) < 2:
                continue
            if kb_n in cleaned_n or cleaned_n in kb_n:
                matched_known = True
                break
            kb_raw = _normalize_simple_brand_token(str(kb)).lower()
            if kb_raw and len(kb_raw) >= 3 and re.search(
                rf"(?<!\w){re.escape(kb_raw)}(?!\w)", cleaned, flags=re.IGNORECASE
            ):
                matched_known = True
                break
        if matched_known:
            continue

        cand = _extract_brand_entity_with_ai(nm, api_key)
        if not cand or cand == "Unknown":
            continue
        cand = _strip_brand_entity_stopwords(cand)
        cand = _normalize_simple_brand_token(cand)
        if not cand:
            continue

        best_score = 0.0
        for kb in known_brands:
            sc = _fuzzy_ratio_local(
                normalize_brand_name_v2(cand),
                normalize_brand_name_v2(str(kb)),
            )
            if sc > best_score:
                best_score = sc
        if best_score >= threshold:
            continue
        key = normalize_brand_name_v2(cand)
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append({
            "اسم الماركة": cand,
            "وصف مختصر عن الماركة": "",
            "صورة شعار الماركة": "",
            "(إختياري) صورة البانر": "",
            "(Page Title) عنوان صفحة العلامة التجارية": f"عطور {cand} الأصلية | مهووس",
            "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(cand),
            "(Page Description) وصف صفحة العلامة التجارية": "",
        })
    return pd.DataFrame(rows, columns=SALLA_BRANDS_COLS)


def _finalize_mahwous_brand_page_url(raw_from_ai: str, english_slug_source: str) -> str:
    """
    رابط سلة العلامة: أحرف لاتينية صغيرة + شرطات سفلية فقط،
    ويجب أن ينتهي بـ _mahwous_store (يُكمّل بايثون إن نسي النموذج).
    """
    raw = str(raw_from_ai or "").strip().lower()
    raw = re.sub(r"^https?://", "", raw).split("/")[0].split("?")[0]
    raw = re.sub(r"[^a-z0-9_-]", "", raw).replace("-", "_")
    raw = re.sub(r"_+", "_", raw).strip("_")
    base = raw if raw and re.search(r"[a-z]", raw) else ""
    if not base:
        ep = str(english_slug_source or "").strip().lower()
        base = re.sub(r"[^a-z0-9]+", "_", ep).strip("_")
    if not base:
        return ""
    base = base.replace("_mahwous_store", "").strip("_")
    if not base:
        return ""
    return f"{base}_mahwous_store"


def _ai_enrich_brand_row_with_domain(brand_name: str, api_key: str) -> dict:
    empty = {
        "brand_name": "",
        "description": "",
        "page_title": "",
        "page_url": "",
        "seo_description": "",
        "domain": "",
        "logo_clearbit_url": "",
    }
    if not brand_name or not api_key or not HAS_ANTHROPIC:
        return empty
    try:
        client = anthropic.Anthropic(api_key=api_key)
        prompt = (
            "أنت خبير عالمي في ماركات العطور.\n"
            f"الماركة: {brand_name}\n\n"
            "أعد JSON فقط بدون أي نص خارج JSON بالمفاتيح التالية حرفياً:\n"
            '{"brand_name":"Arabic_Name | English_Name", "description":"...", "page_title":"...", "page_url":"...", "seo_description":"...", "domain":"..."}\n'
            "الشروط الصارمة:\n"
            "- brand_name: \"Arabic_Name | English_Name\" (مسافة حول |)، حد أقصى 30 حرفاً مع المسافات.\n"
            "- description: وصف عربي فريد واحترافي (حد أقصى 255 حرفاً)، بدون نسخ قوالب جاهزة.\n"
            "- page_title: عنوان SEO (حد أقصى 70 حرفاً).\n"
            "- page_url: الاسم الإنجليزي للماركة فقط بأحرف لاتينية صغيرة؛ المسافات والشرطات تُستبدل بشرطة سفلية؛ "
            "يجب أن ينتهي بـ _mahwous_store. ممنوع أي حرف عربي أو رموز داخل page_url.\n"
            "- seo_description: وصف عربي SEO (حد أقصى 155 حرفاً).\n"
            "- domain: الدومين الرسمي للماركة فقط (مثل chanel.com) بدون http/https أو www أو مسار.\n"
            "- إذا تعذر تحديد أي حقل بدقة اعده كسلسلة فارغة فقط."
        )
        msg = anthropic_messages_create(
            client,
            model="claude-3-haiku-20240307",
            max_tokens=350,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = msg.content[0].text.strip()
        data = _parse_json_object_from_llm_text(raw, context="_ai_enrich_brand_row_with_domain")
        if not data:
            return empty

        brand_name_raw = str(data.get("brand_name", "") or "").strip()
        desc = str(data.get("description", "") or "").strip()
        page_title = str(data.get("page_title", "") or "").strip()
        seo_desc = str(data.get("seo_description", "") or "").strip()
        dom = str(data.get("domain", "") or "").strip().lower()
        dom = re.sub(r"^https?://", "", dom).split("/")[0].strip()
        dom = re.sub(r"^www\.", "", dom)
        if dom and "." not in dom:
            dom = ""

        # enforce stop-word stripping inside Arabic part of brand_name
        # (حتى لا ينتهي بنا الحال بعناوين مثل: "تستر فرزاتشي").
        def _strip_brand_stop_words_from_ar_part(bn: str) -> str:
            b = str(bn or "").strip()
            if "|" in b:
                ar_part, en_part = b.split("|", 1)
                ar_part = ar_part.strip()
                ar_part = re.sub(r"(?:تستر|عطر|طقم|مجموعة)", " ", ar_part, flags=re.IGNORECASE)
                ar_part = re.sub(r"مزيل\s+عرق", " ", ar_part, flags=re.IGNORECASE)
                ar_part = re.sub(r"\s+", " ", ar_part).strip()
                return f"{ar_part} | {en_part.strip()}"
            # if format isn't correct, just strip and keep
            b = re.sub(r"(?:تستر|عطر|طقم|مجموعة)", " ", b, flags=re.IGNORECASE)
            b = re.sub(r"مزيل\s+عرق", " ", b, flags=re.IGNORECASE)
            b = re.sub(r"\s+", " ", b).strip()
            return b

        brand_name_raw = _strip_brand_stop_words_from_ar_part(brand_name_raw)

        # enforce Salla limits in code (حتى لو Claude تجاوزها).
        desc = desc[:255].strip()
        seo_desc = seo_desc[:155].strip()
        page_title = page_title[:70].strip()

        # enforce brand_name format "Arabic | English" and max length <= 30.
        if "|" in brand_name_raw:
            parts = [p.strip() for p in brand_name_raw.split("|", 1)]
            bn_final = f"{parts[0]} | {parts[1]}".strip()
        else:
            bn_final = f"{brand_name_raw} | {brand_name_raw}".strip()
        if len(bn_final) > 30:
            # نحتفظ بـ "| " قدر الإمكان عند التقصير.
            if "|" in bn_final:
                ar_part, en_part = bn_final.split("|", 1)
                ar_part = ar_part.strip()
                en_part = en_part.strip()
                sep = " | "
                available = 30 - len(sep)
                if available > 0:
                    ar_len_max = min(len(ar_part), available // 2 + available % 2)
                    ar2 = ar_part[:ar_len_max].rstrip()
                    en_len_max = max(0, available - len(ar2))
                    en2 = en_part[:en_len_max].rstrip()
                    bn_final = f"{ar2} | {en2}".strip()
                else:
                    bn_final = bn_final[:30].rstrip()
            else:
                bn_final = bn_final[:30].rstrip()

        en_part = ""
        if "|" in bn_final:
            en_part = bn_final.split("|", 1)[1].strip()
        en_slug = en_part.lower()
        en_slug = re.sub(r"[^a-z0-9\s_]", "", en_slug)
        en_slug = re.sub(r"\s+", "_", en_slug).strip("_")
        cl_url = str(data.get("page_url", "") or "").strip()
        page_url = _finalize_mahwous_brand_page_url(cl_url, en_slug)

        return {
            "brand_name": bn_final,
            "description": desc,
            "page_title": page_title,
            "page_url": page_url,
            "seo_description": seo_desc,
            "domain": dom,
            "logo_clearbit_url": (f"https://logo.clearbit.com/{dom}" if dom else ""),
        }
    except Exception:
        return empty


def render_store_audit_tab():

        st.markdown("""<div class="al-info">
        ارفع ملف المتجر الأساسي (بتنسيق سلة). سيقوم النظام بفحصه واكتشاف المنتجات
        التي تحتاج معالجة (بدون صورة، بدون تصنيف، بدون ماركة، بدون وصف، بدون سعر).
        ثم يستخرج هذه المنتجات في ملف بتنسيق "ملف تحديث أو تعديل منتجات سلة" جاهز للرفع.
        </div>""", unsafe_allow_html=True)

        # ── رفع الملف ────────────────────────────────────────────────
        st.markdown("""<div class="sec-title"><div class="bar"></div><h3>رفع ملف المتجر</h3></div>""",
                    unsafe_allow_html=True)

        up_audit = st.file_uploader("ارفع ملف المتجر الأساسي (CSV / Excel)",
                                     type=["csv","xlsx","xls"], key="sa_audit_up")
        if up_audit:
            df_audit_raw = read_file(up_audit, salla_2row=True)
            if df_audit_raw.empty:
                df_audit_raw = read_file(up_audit, salla_2row=False)
            if not df_audit_raw.empty:
                st.session_state.audit_df = df_audit_raw
                st.success(f"✅ {len(df_audit_raw):,} منتج في الملف")
                st.rerun()

        if st.session_state.audit_df is not None:
            audit_df = st.session_state.audit_df

            # ── تعيين الأعمدة ────────────────────────────────────────
            st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
            <h3>تعيين الأعمدة</h3></div>""", unsafe_allow_html=True)

            NONE_A = "— لا يوجد —"
            a_opts = [NONE_A] + list(audit_df.columns)
            def agi(kws):
                g = auto_guess_col(audit_df.columns, kws, audit_df)
                return a_opts.index(g) if g in a_opts else 0

            a1, a2, a3, a4, a5, a6 = st.columns(6)
            with a1: a_no   = st.selectbox("No.", a_opts, index=agi(["no.","no","رقم","id"]), key="sa_a_no")
            with a2: a_nm   = st.selectbox("اسم المنتج", a_opts, index=agi(["اسم","name","منتج"]), key="sa_a_nm")
            with a3: a_img  = st.selectbox("الصورة", a_opts, index=agi(["صورة","image","img"]), key="sa_a_img")
            with a4: a_cat  = st.selectbox("التصنيف", a_opts, index=agi(["تصنيف","category","قسم"]), key="sa_a_cat")
            with a5: a_br   = st.selectbox("الماركة", a_opts, index=agi(["ماركة","brand","علامة"]), key="sa_a_br")
            with a6: a_desc = st.selectbox("الوصف", a_opts, index=agi(["وصف","description","desc"]), key="sa_a_desc")

            a7, a8, a9 = st.columns(3)
            with a7: a_pr   = st.selectbox("السعر", a_opts, index=agi(["سعر","price"]), key="sa_a_pr")
            with a8: a_sku  = st.selectbox("SKU", a_opts, index=agi(["sku","رمز","barcode"]), key="sa_a_sku")
            with a9: a_stat = st.selectbox(
                "حالة المنتج", a_opts, index=agi(["حالة المنتج","حالة","status","مرئي","متاح"]), key="sa_a_stat")

            a10, a11, a12 = st.columns(3)
            with a10: a_cost = st.selectbox(
                "سعر التكلفة", a_opts,
                index=agi(["سعر التكلفة", "تكلفة", "cost"]),
                key="sa_a_cost",
            )
            with a11: a_sale = st.selectbox(
                "السعر المخفض", a_opts,
                index=agi(["السعر المخفض", "مخفض", "sale", "discount"]),
                key="sa_a_sale",
            )
            with a12: a_tax = st.selectbox(
                "خاضع للضريبة ؟", a_opts,
                index=agi(["خاضع للضريبة ؟", "خاضع للضريبة", "ضريبة", "tax"]),
                key="sa_a_tax",
            )

            min_margin_pct = st.number_input(
                "الحد الأدنى للهامش الربحي (%)",
                min_value=0.0, max_value=90.0, value=15.0, step=0.5,
                key="sa_min_margin_pct",
            )

            # ── تشغيل الفحص ──────────────────────────────────────────
            if st.button("🔍 فحص الملف الآن", type="primary", key="sa_run_audit"):
                issues = []
                prog_bar = st.progress(0, text="جاري فحص المنتجات...")
                if a_pr == NONE_A or a_cost == NONE_A:
                    st.error("حدد عمود 'السعر' وعمود 'سعر التكلفة' لتدقيق الهامش.")
                    st.stop()
                total = len(audit_df)
                for i, row in audit_df.iterrows():
                    if i % 10 == 0:
                        prog_bar.progress(int((i / max(total, 1)) * 100), text=f"فحص: {i}/{total}")
                    row_issues = []
                    name = str(row.get(a_nm, "") or "").strip() if a_nm != NONE_A else ""
                    if not name or name == "nan":
                        continue

                    if a_img != NONE_A and not str(row.get(a_img, "") or "").strip():
                        row_issues.append("بدون صورة")
                    if a_cat != NONE_A and not str(row.get(a_cat, "") or "").strip():
                        row_issues.append("بدون تصنيف")
                    if a_br != NONE_A and not str(row.get(a_br, "") or "").strip():
                        row_issues.append("بدون ماركة")

                    desc_val = str(row.get(a_desc, "") or "").strip() if a_desc != NONE_A else ""
                    if not desc_val or desc_val == "nan" or len(desc_val) < 20:
                        row_issues.append("بدون وصف")
                    elif ("تستر" in name.lower() or "tester" in name.lower()) and                         "تستر" not in desc_val and "tester" not in desc_val.lower():
                        row_issues.append("وصف التستر غير صحيح")

                    if a_pr != NONE_A:
                        pr_raw = str(row.get(a_pr, "") or "").strip()
                        if pr_raw in ["0", "nan", ""]:
                            row_issues.append("بدون سعر")
                        else:
                            _ok_pr, _pv = parse_price_numeric(pr_raw)
                            if not _ok_pr or _pv <= 0:
                                row_issues.append("بدون سعر")

                    # ── Price audit: مقارنة سعر المنتج مقابل سعر التكلفة ──
                    ok_pr = False; pv = 0.0
                    ok_cost = False; cost_v = 0.0
                    if a_pr != NONE_A:
                        pr_raw = str(row.get(a_pr, "") or "").strip()
                        ok_pr, pv = parse_price_numeric(pr_raw)
                    if a_cost != NONE_A:
                        cost_raw = str(row.get(a_cost, "") or "").strip()
                        ok_cost, cost_v = parse_price_numeric(cost_raw)

                    if a_cost != NONE_A and (not ok_cost or cost_v <= 0):
                        row_issues.append("بدون سعر تكلفة")

                    if a_pr != NONE_A and a_cost != NONE_A and ok_pr and ok_cost and pv > 0 and cost_v > 0:
                        margin_pct = (pv - cost_v) / cost_v * 100.0
                        if pv < cost_v:
                            row_issues.append("السعر أقل من التكلفة")
                        elif margin_pct < float(min_margin_pct):
                            row_issues.append(f"هامش منخفض (<{min_margin_pct}%)")

                    if row_issues:
                        no_cell = ""
                        sku_cell = ""
                        if a_no != NONE_A:
                            no_cell = str(row.get(a_no, "") or "").strip()
                        if a_sku != NONE_A:
                            sku_cell = str(row.get(a_sku, "") or "").strip()
                        stat_cell = "مرئي"
                        if a_stat != NONE_A:
                            stat_cell = str(row.get(a_stat, "") or "").strip() or "مرئي"

                        tax_cell = "نعم"
                        if a_tax != NONE_A:
                            tax_cell = str(row.get(a_tax, "") or "").strip()
                            if not tax_cell or tax_cell.lower() in ("nan", "none"):
                                tax_cell = "نعم"
                            else:
                                tl = tax_cell.lower()
                                if tl in ("نعم", "yes", "true", "1"):
                                    tax_cell = "نعم"
                                elif tl in ("لا", "no", "false", "0"):
                                    tax_cell = "لا"
                                else:
                                    tax_cell = str(row.get(a_tax, "") or "").strip()

                        issues.append({
                            "No.":               no_cell,
                            "النوع ":            "منتج",
                            "أسم المنتج":        name,
                            "الماركة":           str(row.get(a_br, "") or "") if a_br != NONE_A else "",
                            "تصنيف المنتج":      str(row.get(a_cat, "") or "") if a_cat != NONE_A else "",
                            "صورة المنتج":       str(row.get(a_img, "") or "") if a_img != NONE_A else "",
                            "وصف صورة المنتج":   name,
                            "نوع المنتج":        "منتج جاهز",
                            "سعر المنتج":        normalize_price_digits(row.get(a_pr, "")) if a_pr != NONE_A else "",
                            "سعر التكلفة":       normalize_price_digits(row.get(a_cost, "")) if a_cost != NONE_A else "",
                            "السعر المخفض":      normalize_price_digits(row.get(a_sale, "")) if a_sale != NONE_A else "",
                            "الوصف":             desc_val,
                            "هل يتطلب شحن؟":    "نعم",
                            "رمز المنتج sku":    sku_cell,
                            "الوزن":             "0.2",
                            "وحدة الوزن":        "kg",
                            "حالة المنتج":       stat_cell,
                            "خاضع للضريبة ؟":   tax_cell,
                            "اقصي كمية لكل عميل": "0",
                            "تثبيت المنتج":      "لا",
                            "اضافة صورة عند الطلب": "لا",
                            "_issues":           " | ".join(row_issues),
                            "_idx":              i,
                        })
                prog_bar.progress(100, text="اكتمل الفحص!")
                st.session_state.audit_results = pd.DataFrame(issues) if issues else pd.DataFrame()
                st.rerun()

            # ── عرض نتائج الفحص ──────────────────────────────────────
            if st.session_state.audit_results is not None:
                audit_res = st.session_state.audit_results

                if audit_res.empty:
                    st.success("✅ الملف مكتمل — لا توجد منتجات تحتاج معالجة!")
                else:
                    # إحصائيات
                    no_img  = int(audit_res["_issues"].str.contains("بدون صورة").sum())
                    no_cat  = int(audit_res["_issues"].str.contains("بدون تصنيف").sum())
                    no_br   = int(audit_res["_issues"].str.contains("بدون ماركة").sum())
                    no_desc = int(audit_res["_issues"].str.contains("بدون وصف").sum())
                    no_pr   = int(audit_res["_issues"].str.contains("بدون سعر").sum())
                    no_cost = int(audit_res["_issues"].str.contains("بدون سعر تكلفة").sum())
                    bad_pr  = int(audit_res["_issues"].str.contains("السعر أقل من التكلفة").sum())
                    low_m   = int(audit_res["_issues"].str.contains("هامش منخفض").sum())

                    st.markdown(f"""
                    <div class="stats-bar">
                      <div class="stat-box"><div class="n" style="color:#e53935">{len(audit_res):,}</div><div class="lb">تحتاج معالجة</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{no_img:,}</div><div class="lb">بدون صورة</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{no_cat:,}</div><div class="lb">بدون تصنيف</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{no_br:,}</div><div class="lb">بدون ماركة</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{no_desc:,}</div><div class="lb">بدون وصف</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{no_pr:,}</div><div class="lb">بدون سعر</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{no_cost:,}</div><div class="lb">بدون سعر تكلفة</div></div>
                      <div class="stat-box"><div class="n" style="color:#e53935">{bad_pr:,}</div><div class="lb">سعر أقل من تكلفة</div></div>
                      <div class="stat-box"><div class="n" style="color:#f9a825">{low_m:,}</div><div class="lb">هامش منخفض</div></div>
                    </div>
                    """, unsafe_allow_html=True)

                    # فلتر حسب نوع المشكلة
                    filter_opts = ["الكل", "بدون صورة", "بدون تصنيف", "بدون ماركة", "بدون وصف", "بدون سعر",
                                    "بدون سعر تكلفة", "السعر أقل من التكلفة", "هامش منخفض"]
                    audit_filter = st.selectbox("فلتر حسب المشكلة:", filter_opts, key="sa_audit_filter")

                    if audit_filter == "الكل":
                        filtered_audit = audit_res
                    else:
                        filtered_audit = audit_res[audit_res["_issues"].str.contains(audit_filter)]

                    # عرض الجدول
                    display_cols = ["No.", "أسم المنتج", "الماركة", "تصنيف المنتج", "_issues"]
                    st.dataframe(
                        filtered_audit[[c for c in display_cols if c in filtered_audit.columns]],
                        use_container_width=True
                    )

                    # ── تصدير تدقيق الأسعار بنفس قالب سلة (42 عمود) ──
                    date_str = datetime.now().strftime("%Y%m%d_%H%M")
                    st.download_button(
                        "📥 تصدير تدقيق الأسعار — CSV (قالب سلة)",
                        export_product_csv(filtered_audit),
                        f"price_audit_{date_str}.csv",
                        "text/csv",
                        width='stretch',
                        key="sa_price_audit_dl_csv",
                    )

                    # ── الإصلاح التلقائي والتصدير ──────────────────────────────────
                    st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
                    <h3>🛠️ الإصلاح التلقائي وتصدير الملف</h3></div>""", unsafe_allow_html=True)
                    st.info(f"سيقوم النظام بإصلاح {len(filtered_audit):,} منتج تلقائياً (جلب صور، توليد أوصاف، تعيين ماركات وتصنيفات).")

                    if st.button("🚀 بدء إصلاح النواقص أوتوماتيكياً", type="primary",
                                 key="sa_start_auto_fix", use_container_width=True):
                        if not st.session_state.api_key:
                            st.error("⚠️ يرجى إضافة مفتاح Claude API لتوليد الأوصاف.")
                        else:
                            fix_prog = st.progress(0); fix_stat = st.empty()
                            fixed_rows = []
                            total_fix = len(filtered_audit)

                            for fix_i, (_, f_row) in enumerate(filtered_audit.iterrows()):
                                fix_prog.progress(int((fix_i / max(total_fix, 1)) * 100))
                                pname = f_row["أسم المنتج"]
                                iss   = f_row["_issues"]
                                fix_stat.markdown(
                                    f'<div class="prog-run">جاري إصلاح: {pname[:50]}...</div>',
                                    unsafe_allow_html=True)

                                attrs  = extract_product_attrs(pname)
                                size_f = attrs.get("size") or 0
                                if not size_f:
                                    size = "100 مل"
                                else:
                                    size = f"{int(size_f) if size_f == int(size_f) else size_f} مل"
                                size = _normalize_product_size_ml(size) or size
                                conc   = attrs.get("concentration") or "EDP"
                                conc_ar = {"EDP": "أو دو بارفيوم", "EDT": "أو دو تواليت",
                                           "EDC": "أو دو كولون", "PARFUM": "بارفيوم",
                                           "UNKNOWN": "أو دو بارفيوم"}.get(conc, "أو دو بارفيوم")
                                is_t   = "تستر" in attrs.get("type", "")
                                gender = ("للنساء" if any(w in pname.lower() for w in ["نساء","women","نسائ"])
                                          else "للرجال" if any(w in pname.lower() for w in ["رجال","men","رجالي"])
                                          else "للجنسين")

                                brand_dict = match_brand(pname)
                                if not brand_dict.get("name") and f_row["الماركة"]:
                                    brand_dict = match_brand(f_row["الماركة"])
                                if "بدون ماركة" in iss or not brand_dict.get("name"):
                                    if not brand_dict.get("name"):
                                        extracted_b = clean_brand_name(pname.split()[0] if pname.split() else "")
                                        if extracted_b:
                                            brand_dict = generate_new_brand(extracted_b)
                                f_row["الماركة"] = brand_dict.get("name", f_row["الماركة"])

                                f_row["أسم المنتج"] = standardize_product_name(pname, f_row["الماركة"])
                                pname = f_row["أسم المنتج"]

                                if "بدون تصنيف" in iss or not f_row["تصنيف المنتج"]:
                                    f_row["تصنيف المنتج"] = ("العطور > تستر" if is_t
                                                              else match_category(pname, gender))

                                if "بدون صورة" in iss or not f_row["صورة المنتج"]:
                                    f_row["صورة المنتج"] = fetch_image(pname, is_t)

                                if "بدون وصف" in iss or "وصف التستر غير صحيح" in iss:
                                    f_row["الوصف"] = ai_generate(
                                        pname, is_t, brand_dict, size, gender, conc_ar)

                                no_seo = str(f_row.get("No.", "") or fix_i + 1).strip()
                                seo_data = gen_seo(
                                    pname, brand_dict, size, is_t, gender,
                                    sku_suffix=f"V-{no_seo}",
                                    type_hint=str(attrs.get("type", "") or ""),
                                )
                                f_row["وصف صورة المنتج"] = seo_data["alt"]

                                final_dict = {col: "" for col in SALLA_COLS}
                                for col in SALLA_COLS:
                                    if col in f_row:
                                        final_dict[col] = str(f_row[col])
                                final_dict["النوع "] = "منتج"
                                fixed_rows.append(final_dict)

                            fix_prog.progress(100)
                            fix_stat.markdown(
                                '<div class="prog-ok">✅ اكتمل الإصلاح التلقائي!</div>',
                                unsafe_allow_html=True)
                            st.session_state.audit_fixed_df = pd.DataFrame(fixed_rows, columns=SALLA_COLS)

                    if "audit_fixed_df" in st.session_state and st.session_state.audit_fixed_df is not None:
                        date_str = datetime.now().strftime("%Y%m%d_%H%M")
                        aud_e1, aud_e2, aud_e3 = st.columns(3)
                        with aud_e1:
                            st.download_button(
                                f"📥 الملف المُصلح — Excel ({len(st.session_state.audit_fixed_df)})",
                                export_product_xlsx(st.session_state.audit_fixed_df),
                                f"تحديث_منتجات_{date_str}.xlsx",
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                width='stretch', key="sa_dl_audit_fix_x")
                        with aud_e2:
                            st.download_button(
                                f"📥 الملف المُصلح — CSV ({len(st.session_state.audit_fixed_df)})",
                                export_product_csv(st.session_state.audit_fixed_df),
                                f"تحديث_منتجات_{date_str}.csv",
                                "text/csv", width='stretch', key="sa_dl_audit_fix_c")
                        with aud_e3:
                            if st.button("🛠️ نقل للمسار الآلي لمراجعة إضافية", key="sa_audit_to_proc",
                                         width='stretch'):
                                st.session_state.pipe_approved = st.session_state.audit_fixed_df
                                st.session_state.pipe_export_df = st.session_state.audit_fixed_df.copy()
                                st.session_state.pipe_step = 5
                                st.session_state.pipe_results = None
                                st.session_state.page = "pipeline"
                                st.rerun()

                    if st.button("🔄 إعادة الفحص", key="sa_reset_audit"):
                        st.session_state.audit_results = None
                        if "audit_fixed_df" in st.session_state:
                            del st.session_state["audit_fixed_df"]
                        st.rerun()

        else:
            st.markdown("""
            <div class="upload-zone">
              <div class="uz-icon">🏪</div>
              <div class="uz-title">ارفع ملف المتجر الأساسي للبدء</div>
              <div class="uz-sub">سيكتشف النظام المنتجات الناقصة ويجهّز ملف التحديث تلقائياً</div>
            </div>
            """, unsafe_allow_html=True)




# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 1 — AUTO PIPELINE (المسار الآلي)                         ║
# ╚══════════════════════════════════════════════════════════════════╝
if st.session_state.page == "pipeline":

    # ── وصف الصفحة ──────────────────────────────────────────────────
    st.markdown("""<div class="al-info">
    <b>المسار الآلي الكامل:</b> ارفع ملف المتجر وملفات المنافسين ← ضغطة زر واحدة ←
    المحرك الذكي v10.4 يقارن ويصفي ← AI يراجع المشبوه ← الوصف والـ SEO يتولدان تلقائياً
    ← ملف سلة جاهز للرفع + ماركات جديدة جاهزة.<br>
    <span style="opacity:0.9">قبل التصدير يُستبعد تلقائياً أي منتج يطابق متجرك في <b>أسم المنتج</b> (fuzzy 90٪)
    أو في <b>رمز المنتج sku</b> إن وُجد.</span>
    </div>""", unsafe_allow_html=True)

    # ── STEP INDICATOR ───────────────────────────────────────────────
    step = st.session_state.pipe_step
    steps_html = "".join([
        f'<div style="display:inline-flex;align-items:center;gap:6px;margin-left:14px">'
        f'<div style="width:28px;height:28px;border-radius:50%;background:'
        f'{"#b8933a" if step>=i else "rgba(184,147,58,0.2)"};'
        f'color:{"#0f0e0d" if step>=i else "#9a8e80"};display:flex;align-items:center;justify-content:center;'
        f'font-size:0.8rem;font-weight:900">{i}</div>'
        f'<span style="font-size:0.78rem;color:{"#b8933a" if step>=i else "#9a8e80"};font-weight:{"800" if step==i else "400"}">{lbl}</span>'
        f'</div>'
        for i, lbl in enumerate(
            ["رفع الملفات","مقارنة","فلترة AI","معالجة","التصدير"], 1)
    ])
    st.markdown(
        f'<div style="background:white;border:1px solid rgba(184,147,58,0.2);'
        f'border-radius:12px;padding:14px 20px;margin-bottom:16px;display:flex;'
        f'align-items:center;gap:4px;flex-wrap:wrap;">{steps_html}</div>',
        unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════════
    # STEP 1 — رفع الملفات
    # ════════════════════════════════════════════════════════════════
    st.markdown("""<div class="sec-title"><div class="bar"></div><h3>رفع الملفات</h3></div>""",
                unsafe_allow_html=True)
    pc1, pc2 = st.columns(2)
    with pc1:
        st.markdown("**ملف متجرنا (مهووس)** — ملف سلة كامل")
        if st.session_state.pipe_store_df is not None:
            st.markdown(f'<div class="al-ok">محمّل: {len(st.session_state.pipe_store_df):,} منتج</div>',
                        unsafe_allow_html=True)
        up_ps = st.file_uploader("ارفع ملف المتجر", type=["csv","xlsx","xls"],
                                  key="pipe_store_up", label_visibility="collapsed")
        if up_ps:
            df_ps = read_file(up_ps, salla_2row=True)
            if df_ps.empty:
                df_ps = read_file(up_ps, salla_2row=False)
            if not df_ps.empty:
                ok_v, err_v = validate_input_dataframe(df_ps, "ملف المتجر")
                if not ok_v:
                    st.error(err_v[0])
                    APP_LOG.warning("store upload validation: %s", err_v)
                else:
                    st.session_state.pipe_store_df = df_ps
                    st.session_state.pipe_step = max(st.session_state.pipe_step, 1)
                    st.success(f"✅ {len(df_ps):,} منتج في المتجر")

    with pc2:
        st.markdown("**ملفات المنافسين** — يمكن رفع أكثر من ملف")
        if st.session_state.pipe_comp_dfs:
            tot = sum(len(d) for d in st.session_state.pipe_comp_dfs)
            st.markdown(f'<div class="al-ok">محمّل: {tot:,} منتج من {len(st.session_state.pipe_comp_dfs)} ملف</div>',
                        unsafe_allow_html=True)
        up_pc = st.file_uploader("ارفع ملفات المنافسين", type=["csv","xlsx","xls"],
                                  key="pipe_comp_up", accept_multiple_files=True,
                                  label_visibility="collapsed")
        if up_pc:
            new_dfs = []
            for f in up_pc:
                df_c = read_file(f)
                if not df_c.empty:
                    df_c["_source"] = f.name
                    new_dfs.append(df_c)
            if new_dfs:
                merged_c = pd.concat(new_dfs, ignore_index=True)
                ok_v, err_v = validate_input_dataframe(merged_c, "ملفات المنافسين")
                if not ok_v:
                    st.error(err_v[0])
                    APP_LOG.warning("competitor upload validation: %s", err_v)
                else:
                    st.session_state.pipe_comp_dfs = new_dfs
                    tot2 = sum(len(d) for d in new_dfs)
                    st.session_state.pipe_step = max(st.session_state.pipe_step, 1)
                    st.success(f"✅ {tot2:,} منتج من {len(new_dfs)} ملف")

    # ── إعدادات المحرك (خفية بشكل افتراضي) ──────────────────────────
    with st.expander("⚙️ إعدادات خوارزمية المقارنة (اختياري — القيم الافتراضية مثالية)", expanded=False):
        pe1, pe2, pe3 = st.columns(3)
        with pe1:
            pipe_t_dup  = st.slider("عتبة المكرر (%)", 80, 99, 98, key="pipe_tdup")
        with pe2:
            pipe_t_near = st.slider("عتبة المراجعة (%)", 40, 85, 70, key="pipe_tnear")
        with pe3:
            pipe_suspect_pct = st.slider("نسبة الاختلاف المشبوه (%) للأسعار", 10, 50, 20, key="pipe_suspect")
    pipe_t_dup       = st.session_state.get("pipe_tdup", 98)
    pipe_t_near      = st.session_state.get("pipe_tnear", 70)
    pipe_suspect_pct = st.session_state.get("pipe_suspect", 20)

    # ── فلاتر الاستبعاد الصارمة (قبل المقارنة) ─────────────────────
    st.markdown("""<div class="sec-title"><div class="bar"></div>
    <h3>شروط الاستبعاد قبل المقارنة</h3></div>""", unsafe_allow_html=True)
    st.caption(
        "تُطبَّق هذه الشروط على ملفات المنافسين فقط — ملف المتجر يبقى كاملاً للمقارنة. "
        "«العينات» هنا: الحجم الصغير (مثلاً 8 مل وأقل) أو سمبل/فايال؛ عطور التستر لا تُستبعد بهذا الخيار."
    )
    fx1, fx2 = st.columns(2)
    with fx1:
        st.checkbox(
            "استبعاد العينات (حجم صغير / سمبل — لا يشمل عطور التستر)",
            key="pipe_fx_samples",
        )
        st.checkbox(
            "استبعاد الكماليات والمنتجات الثانوية (حقائب، أدوات تطبيق، …)",
            key="pipe_fx_accessories",
        )
    with fx2:
        st.checkbox(
            "الاحتفاظ بمنتجات الماركات العالمية فقط (حسب قائمة الماركات المعتمدة)",
            key="pipe_fx_brand",
        )
        st.checkbox(
            "استبعاد المنتجات التي لا يظهر في اسمها أو وصفها حجم/سعة (مل، أونصة، …)",
            key="pipe_fx_volume",
        )

    # ════════════════════════════════════════════════════════════════
    # MAIN ACTION BUTTON — بدء المسار الآلي
    # ════════════════════════════════════════════════════════════════
    has_store_p = st.session_state.pipe_store_df is not None
    has_comp_p  = bool(st.session_state.pipe_comp_dfs)

    if not (has_store_p and has_comp_p):
        st.markdown("""<div class="upload-zone"><div class="uz-icon">🚀</div>
        <div class="uz-title">ارفع ملف المتجر وملفات المنافسين لتفعيل المسار</div>
        <div class="uz-sub">بعد الرفع، اضغط "بدء المسار الآلي" — كل شيء يعمل تلقائياً</div>
        </div>""", unsafe_allow_html=True)
    else:
        aok_pipe = bool(st.session_state.api_key)
        if not aok_pipe:
            st.markdown(
                '<div class="al-warn">⚠️ لم يُضبط مفتاح Claude API — المسار يعمل بدون تحقق AI للمشبوه '
                "(يُعامل كـ «جديد» لعدم فقدان البيانات)، والأوصاف التلقائية تبقى فارغة.</div>",
                unsafe_allow_html=True,
            )

        if st.button("🚀 بدء المسار الآلي الكامل", type="primary",
                     key="pipe_run", use_container_width=True):
            st.session_state.pipe_results    = None
            st.session_state.pipe_approved   = None
            st.session_state.pipe_export_df  = None
            st.session_state.pipe_new_brands = []
            st.session_state.pipe_session_brands = []
            st.session_state.pipe_seo_df     = None
            st.session_state.pipe_store_dedup_dropped = 0
            st.session_state.pipe_missing_brands_df = None
            if "_brand_entity_ai_cache" in st.session_state:
                del st.session_state["_brand_entity_ai_cache"]
            st.session_state.pipe_running    = True
            st.session_state.pipe_step       = 2

            store_df_p  = st.session_state.pipe_store_df
            comp_merged = pd.concat(st.session_state.pipe_comp_dfs, ignore_index=True)

            ok_s, err_s = validate_input_dataframe(store_df_p, "ملف المتجر")
            ok_c, err_c = validate_input_dataframe(comp_merged, "ملفات المنافسين")
            if not ok_s or not ok_c:
                for msg in err_s + err_c:
                    st.error(msg)
                    APP_LOG.warning("pipeline input validation: %s", msg)
                st.session_state.pipe_running = False
                st.session_state.pipe_step = 1
                st.stop()

            # ── تحديد أعمدة الاسم تلقائياً ────────────────────────
            NONE_P = "— لا يوجد —"
            store_nm = ("أسم المنتج" if "أسم المنتج" in store_df_p.columns
                        else auto_guess_col(store_df_p.columns,
                                            ["أسم المنتج","اسم","name","منتج"], store_df_p))
            store_sk = (
                "رمز المنتج sku" if "رمز المنتج sku" in store_df_p.columns
                else auto_guess_col(store_df_p.columns, ["رمز المنتج sku", "sku", "رمز", "barcode"], store_df_p)
            )
            store_br = auto_guess_col(store_df_p.columns, ["ماركة","brand"], store_df_p)
            _cn = auto_guess_col(comp_merged.columns,
                                 ["أسم المنتج","اسم","name","منتج"], comp_merged)
            comp_nm = (_cn if _cn != NONE_P else comp_merged.columns[0])
            comp_br_guess = auto_guess_col(comp_merged.columns, ["ماركة", "brand", "label"], comp_merged)
            comp_img = auto_guess_col(comp_merged.columns, ["صورة","image","src","img","w-full src"], comp_merged)
            comp_pr  = auto_guess_col(comp_merged.columns, ["سعر","price","text-sm-2","text-sm","amount"], comp_merged)
            store_sk  = store_sk if store_sk != NONE_P else None
            store_br  = store_br if store_br != NONE_P else None
            comp_img  = comp_img if comp_img != NONE_P else None
            comp_pr   = comp_pr  if comp_pr  != NONE_P else None

            comp_desc_g = auto_guess_col(comp_merged.columns,
                                         ["وصف","description","desc","detail"], comp_merged)
            comp_desc_g = None if comp_desc_g == NONE_P else comp_desc_g

            # استخراج قائمة الماركات
            brands_p = []
            bdf_p = st.session_state.brands_df
            if bdf_p is not None:
                col0_p = bdf_p.columns[0]
                brands_p = bdf_p[col0_p].dropna().astype(str).str.strip().tolist()

            # عند غياب عمود ماركة المنافس: استنتج الماركة من أول كلمات الاسم
            # ثم طابقها fuzzy مع ماركات ملف المتجر/brands.csv. غير المتطابق يُسجل كمفقود.
            comp_br_missing = (not comp_br_guess) or (comp_br_guess == NONE_P) or (comp_br_guess not in comp_merged.columns)
            if comp_br_missing:
                known_brands = _collect_known_brand_names(store_df_p, bdf_p)
                mb_df = _build_missing_brands_df_from_competitors(
                    comp_merged,
                    name_col=comp_nm,
                    known_brands=known_brands,
                    threshold=80.0,
                    api_key=st.session_state.get("api_key") or "",
                )
                if not mb_df.empty:
                    st.session_state.pipe_missing_brands_df = mb_df

            opts = StrictFilterOptions(
                exclude_samples_testers=st.session_state.get("pipe_fx_samples", False),
                exclude_accessories=st.session_state.get("pipe_fx_accessories", False),
                exclude_non_global_brands=st.session_state.get("pipe_fx_brand", False),
                exclude_without_volume=st.session_state.get("pipe_fx_volume", False),
            )
            try:
                comp_merged, cp_stats = apply_strict_pipeline_filters(
                    comp_merged, comp_nm, comp_desc_g, brands_p, opts, label="المنافسين")
            except Exception as _fxe:
                APP_LOG.exception("strict filters failed: %s", _fxe)
                st.error(f"فشل تطبيق الفلاتر: {_fxe}")
                st.session_state.pipe_running = False
                st.session_state.pipe_step = 1
                st.stop()

            st.session_state.pipe_filter_stats = {"competitor": cp_stats}

            if comp_merged.empty:
                st.error(
                    "بعد تطبيق الفلاتر لم يبق أي منتج في ملف المنافسين — "
                    "خفّف شروط الاستبعاد أو غيّر الملفات.")
                APP_LOG.warning("pipeline aborted: competitor empty after strict filters")
                st.session_state.pipe_running = False
                st.session_state.pipe_step = 1
                st.stop()

            # ══ STEP 2: المقارنة ══════════════════════════════════
            status_ph = st.empty()
            status_ph.markdown('<div class="prog-run">⚙️ الخطوة 2: تشغيل محرك المقارنة v10.4...</div>',
                               unsafe_allow_html=True)
            prog_bar = st.progress(10)

            results_p = run_smart_comparison(
                new_df=comp_merged, store_df=store_df_p,
                new_name_col=comp_nm, store_name_col=store_nm,
                new_sku_col=None,
                store_sku_col=store_sk,
                new_img_col=comp_img,
                t_dup=pipe_t_dup, t_near=pipe_t_near,
                t_review=40, brands_list=brands_p,
                store_brand_col=store_br,
            )

            if results_p.empty or "_idx" not in results_p.columns:
                st.error(
                    "محرك المقارنة لم يُرجع أي منتج صالح — غالباً **عمود الاسم** غير مُعرَّف أو كل القيم "
                    "فارغة/غير مقروءة بعد التنظيف. راجع تطابق أعمدة الملف أو ارفع نسخة CSV/Excel سليمة."
                )
                APP_LOG.warning("pipeline: run_smart_comparison empty or missing _idx (rows=%s)", len(comp_merged))
                st.session_state.pipe_running = False
                st.session_state.pipe_step = 1
                st.stop()

            # إضافة عمود السعر (مطابقة مؤشر الصف مع _idx)
            if comp_pr and comp_pr in comp_merged.columns:
                pm = comp_merged[comp_pr]
                results_p["سعر المنافس"] = results_p["_idx"].map(
                    lambda ix: str(pm.loc[ix]) if ix in pm.index else "")

            st.session_state.pipe_results = results_p
            prog_bar.progress(30)

            # تصنيف النتائج
            new_confirmed   = results_p[results_p["الحالة"] == "جديد"].copy()
            suspects_p      = results_p[results_p["الحالة"] == "مشبوه"].copy()

            # ══ STEP 3: الفلترة النهائية الذكية ══════════════════
            st.session_state.pipe_step = 3
            status_ph.markdown(
                f'<div class="prog-run">🤖 الخطوة 3: فلترة {len(suspects_p)} منتج مشبوه بالذكاء الاصطناعي...</div>',
                unsafe_allow_html=True)
            prog_bar.progress(40)

            store_names_p = [str(r.get(store_nm, "")) for _, r in store_df_p.iterrows()]

            _pipe_resolve_key = _effective_anthropic_api_key()
            if _pipe_resolve_key and not str(st.session_state.get("api_key", "") or "").strip():
                st.session_state.api_key = _pipe_resolve_key

            if _pipe_resolve_key and HAS_ANTHROPIC and not suspects_p.empty:
                # تحقق عميق صف-بصف عبر Claude (YES => مكرر، NO => جديد)
                verified_approved_rows = []
                verified_rejected_rows = []
                cache = {}  # (competitor, store) -> YES/NO
                prog_ai = st.progress(0, text="🔍 جاري التحقق العميق بالذكاء الاصطناعي للمنتجات المشتبه بها...")
                for j, (orig_idx, srow) in enumerate(suspects_p.iterrows()):
                    if j % 1 == 0:
                        prog_ai.progress(int((j + 1) / max(len(suspects_p), 1) * 100))
                    comp_nm = str(srow.get("الاسم الجديد", "") or "").strip()
                    closest_nm = str(srow.get("أقرب تطابق في المتجر", "") or "").strip()
                    if not comp_nm or not closest_nm:
                        # بدون اسم متجر كافٍ: افتراض NO لحفظ البيانات
                        verified_approved_rows.append(srow)
                        continue
                    key = (comp_nm, closest_nm)
                    if key in cache:
                        verdict = cache[key]
                    else:
                        verdict = _resolve_suspicious_with_ai(comp_nm, closest_nm)
                        cache[key] = verdict

                    # Golden rule: عند أي خطأ/غموض (""), افترض أنه جديد (NO) لحفظ البيانات
                    if verdict == "YES":
                        verified_rejected_rows.append(srow)
                    else:
                        verified_approved_rows.append(srow)

                prog_ai.progress(100, text="✅ اكتمل التحقق العميق")
                ai_approved = pd.DataFrame(verified_approved_rows) if verified_approved_rows else pd.DataFrame()
                ai_rejected = pd.DataFrame(verified_rejected_rows) if verified_rejected_rows else pd.DataFrame()
            else:
                # بدون AI: اعتمد كل المشبوه كمنتجات جديدة (لا تُسقَط من المسار الآلي)
                ai_approved = suspects_p
                ai_rejected = pd.DataFrame()

            # دمج المؤكد + ما اعتمده AI
            frames = [new_confirmed]
            if not ai_approved.empty:
                frames.append(ai_approved)
            approved_all = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
            prog_bar.progress(55)

            # ══ STEP 4: المُعالج التلقائي ════════════════════════
            st.session_state.pipe_step = 4
            status_ph.markdown(
                f'<div class="prog-run">🛠️ الخطوة 4: معالجة {len(approved_all)} منتج معتمد (ماركات، SEO، وصف)...</div>',
                unsafe_allow_html=True)

            final_rows = []
            seo_rows   = []
            new_brands_found = []
            known_brand_names = set()
            for b in st.session_state.new_brands:
                nm = b.get("اسم الماركة", "") or ""
                if str(nm).strip():
                    known_brand_names.add(normalize_brand_name_v2(nm))

            approved_rows = list(approved_all.iterrows())
            total_ap = len(approved_rows)
            CHUNK_SIZE = 10
            total_chunks = (total_ap + CHUNK_SIZE - 1) // CHUNK_SIZE if total_ap else 0
            pi_idx = 0
            fallback_count = 0
            _pipe_ai_k = _effective_anthropic_api_key()

            for chunk_idx, batch_start in enumerate(range(0, total_ap, CHUNK_SIZE)):
                batch_slice = approved_rows[batch_start:batch_start + CHUNK_SIZE]
                st.info(
                    f"⏳ جاري معالجة الدفعة {chunk_idx + 1} من {total_chunks}..."
                )
                status_ph.markdown(
                    f'<div class="prog-run">🛠️ معالجة الدفعة {chunk_idx + 1} '
                    f'({len(batch_slice)} منتج)...</div>',
                    unsafe_allow_html=True,
                )
                for _, prow in batch_slice:
                    prog_bar.progress(
                        55 + int((pi_idx + 1) / max(total_ap, 1) * 34))

                    pname = strip_trailing_discount_label(str(prow.get("الاسم الجديد", "")))
                    if not pname.strip():
                        continue
                    status_ph.markdown(
                        f'<div class="prog-run">🛠️ معالجة الدفعة {chunk_idx + 1} '
                        f'({len(batch_slice)} منتج) — جاري: '
                        f'<span style="color:#b8933a;font-weight:700">{pname[:72]}</span>'
                        f'{"…" if len(pname) > 72 else ""}</div>'
                        f'<div style="font-size:0.76rem;color:#7a6e60;margin-top:4px">'
                        f'المنتج {pi_idx + 1} من {total_ap} — إثراء المنتج عبر Claude وجلب صورة Google عند توفر المفاتيح؛ '
                        f'كل استدعاء شبكة يضيف ثوانٍ.</div>',
                        unsafe_allow_html=True,
                    )

                    orig_idx = prow.get("_idx", 0)
                    pimg = prow.get("_img", "")
                    if not pimg and comp_img and comp_img in comp_merged.columns:
                        try:
                            pimg = str(comp_merged.iloc[int(orig_idx)].get(comp_img, "") or "")
                        except Exception:
                            pimg = ""
                    pprice = prow.get("سعر المنافس", "")

                    ai_out = {}
                    if _pipe_ai_k and HAS_ANTHROPIC:
                        ai_out = _ai_enrich_product_row(pname, _pipe_ai_k) or {}
                        time.sleep(1.5)
                    if ai_out:
                        pname = ai_out.get("formatted_name") or pname
                        desc = ai_out.get("html_description") or ""
                    else:
                        desc = ""

                    row_br_raw = clean_brand_name(str(prow.get("الماركة", "") or ""))
                    _nk_fb = (
                        normalize_brand_name_v2(row_br_raw)
                        if str(row_br_raw).strip()
                        else ""
                    )
                    prow_brand_fb = (
                        str(row_br_raw).strip()
                        if (_nk_fb and _nk_fb in known_brand_names)
                        else "غير محدد"
                    )
                    clean_pname = _clean_pname_for_fallback(pname)
                    if (not ai_out) or (not str(desc or "").strip()):
                        desc = _generate_fallback_html(clean_pname, prow_brand_fb)
                        fallback_count += 1

                    attrs  = extract_product_attrs(pname)
                    size_s = attrs.get("size") or 0
                    if not size_s:
                        size = "100 مل"
                    else:
                        size = f"{int(size_s) if size_s == int(size_s) else size_s} مل"
                    size = _normalize_product_size_ml(size) or size
                    is_t   = "تستر" in attrs.get("type", "")
                    gender_kw = "للجنسين"
                    nl = pname.lower()
                    if any(w in nl for w in ["رجال","للرجال","men","homme"]): gender_kw = "للرجال"
                    elif any(w in nl for w in ["نساء","للنساء","women","femme"]): gender_kw = "للنساء"

                    brand_d    = match_brand(pname)
                    prow_brand = clean_brand_name(str(prow.get("الماركة","") or ""))
                    if ai_out.get("brand"):
                        prow_brand = str(ai_out["brand"]).strip() or prow_brand

                    if not brand_d.get("name") and prow_brand:
                        brand_d = match_brand(prow_brand)

                    is_new_generated = False

                    if not brand_d.get("name") and not prow_brand:
                        api_k = st.session_state.get("api_key") or _effective_anthropic_api_key()
                        if api_k and HAS_ANTHROPIC:
                            ext = _extract_brand_entity_with_ai(pname, api_k)
                            if ext and ext != "Unknown":
                                ext = _strip_brand_entity_stopwords(ext)
                                brand_d = match_brand(ext)
                                if not brand_d.get("name"):
                                    prow_brand = ext

                    if not brand_d.get("name") and prow_brand:
                        brand_d = generate_new_brand(prow_brand)
                        is_new_generated = True

                    if is_discount_like_brand(str(brand_d.get("name") or "")):
                        brand_d = {"name": "غير محدد", "page_url": ""}
                        is_new_generated = False
                    elif not str(brand_d.get("name") or "").strip():
                        brand_d = {"name": "غير محدد", "page_url": ""}
                        is_new_generated = False

                    if (
                        brand_d.get("name")
                        and str(brand_d["name"]).strip() != "غير محدد"
                        and not brand_exists_in_catalog(brand_d["name"])
                    ):
                        bn_key = normalize_brand_name_v2(brand_d["name"])
                        if bn_key and bn_key not in known_brand_names:
                            if is_new_generated:
                                new_brands_found.append(brand_d.copy())
                            else:
                                new_brands_found.append({
                                    "name": brand_d["name"],
                                    "page_url": brand_d.get("page_url", to_slug(brand_d["name"])),
                                    "اسم الماركة":                                   brand_d["name"],
                                    "وصف مختصر عن الماركة":                         f"علامة تجارية متخصصة في العطور الفاخرة — {brand_d['name']}",
                                    "صورة شعار الماركة":                            "",
                                    "(إختياري) صورة البانر":                        "",
                                    "(Page Title) عنوان صفحة العلامة التجارية":     f"عطور {brand_d['name']} الأصلية | مهووس",
                                    "(SEO Page URL) رابط صفحة العلامة التجارية":    brand_d.get("page_url", to_slug(brand_d["name"])),
                                    "(Page Description) وصف صفحة العلامة التجارية": f"تسوّق أحدث عطور {brand_d['name']} الأصلية بأسعار حصرية من متجر مهووس.",
                                })
                                register_pipe_session_brand(
                                    brand_d["name"],
                                    brand_d.get("page_url", to_slug(brand_d["name"])),
                                )
                            known_brand_names.add(bn_key)

                    cat    = match_category(pname, gender_kw)

                    pi_idx += 1
                    sku_slug = f"V-{pi_idx}"
                    th_pipe = str(attrs.get("type", "") or "")
                    seo    = gen_seo(
                        pname, brand_d, str(size), is_t, gender_kw,
                        sku_suffix=sku_slug,
                        type_hint=th_pipe,
                    )

                    if not str(pimg).strip():
                        pimg = fetch_image(pname, is_t)

                    r = fill_row(
                        name=pname, price=str(pprice), sku="",
                        image=pimg, desc=desc, brand=brand_d,
                        category=cat, seo=seo, no=str(pi_idx),
                        weight="0.2", weight_unit="kg", size=str(size)
                    )
                    final_rows.append(r)
                    seo_rows.append({
                        "No. (غير قابل للتعديل)":                str(pi_idx),
                        "اسم المنتج (غير قابل للتعديل)":         pname,
                        "رابط مخصص للمنتج (SEO Page URL)":       seo.get("url",""),
                        "عنوان صفحة المنتج (SEO Page Title)":    seo.get("title",""),
                        "وصف صفحة المنتج (SEO Page Description)":seo.get("desc",""),
                    })

                prog_bar.progress(55 + int(min(pi_idx, total_ap) / max(total_ap, 1) * 34))
                if _pipe_ai_k and HAS_ANTHROPIC:
                    time.sleep(6)

            if fallback_count > 0:
                st.warning(
                    f"تم استخدام الوصف الذكي البديل لـ {fallback_count} منتج بسبب ضغط سيرفرات الذكاء الاصطناعي."
                )

            prog_bar.progress(90)

            # استبعاد صفوف موجودة مسبقاً في ملف المتجر (أسم المنتج fuzzy + SKU)
            if final_rows and store_df_p is not None and not store_df_p.empty:
                try:
                    tmp_df = pd.DataFrame(final_rows)
                    out_df, kept_pos = filter_new_products_against_store(
                        tmp_df,
                        store_df_p,
                        similarity_threshold=90,
                        use_sku_exact=True,
                        return_positions=True,
                    )
                    n_drop = len(tmp_df) - len(kept_pos)
                    st.session_state.pipe_store_dedup_dropped = int(n_drop)
                    if n_drop:
                        APP_LOG.info(
                            "pipeline store dedup: dropped %s rows already in store file", n_drop
                        )
                    final_rows = out_df.to_dict("records")
                    if len(seo_rows) == len(tmp_df):
                        seo_rows = [seo_rows[i] for i in kept_pos]
                    else:
                        APP_LOG.warning(
                            "pipeline: seo_rows length mismatch (%s vs %s), skip SEO align",
                            len(seo_rows), len(tmp_df),
                        )
                except Exception as _ded_e:
                    APP_LOG.warning("pipeline store dedup failed: %s", _ded_e)
                    st.session_state.pipe_store_dedup_dropped = 0

            final_rows, seo_rows = dedupe_final_rows_and_seo(final_rows, seo_rows)
            new_brands_found = dedupe_brand_entries(new_brands_found)

            # حفظ النتائج
            st.session_state.pipe_approved   = pd.DataFrame(final_rows) if final_rows else pd.DataFrame()
            st.session_state.pipe_export_df  = (
                st.session_state.pipe_approved.copy()
                if final_rows
                else pd.DataFrame()
            )
            st.session_state.pipe_seo_df     = pd.DataFrame(seo_rows)   if seo_rows  else pd.DataFrame()
            st.session_state.pipe_new_brands = new_brands_found
            st.session_state.pipe_step       = 5
            st.session_state.pipe_running    = False
            prog_bar.progress(100)
            status_ph.markdown('<div class="prog-ok">✅ المسار الآلي اكتمل!</div>',
                               unsafe_allow_html=True)
            st.rerun()

    # ════════════════════════════════════════════════════════════════
    # STEP 5 — عرض النتائج والتصدير
    # ════════════════════════════════════════════════════════════════
    if st.session_state.pipe_step >= 5 and st.session_state.pipe_approved is not None:
        approved_df  = st.session_state.pipe_approved
        if st.session_state.get("pipe_export_df") is None:
            st.session_state.pipe_export_df = (
                approved_df.copy() if approved_df is not None else None
            )
        raw_results  = st.session_state.pipe_results
        new_brs      = st.session_state.pipe_new_brands

        new_all    = raw_results[raw_results["الحالة"]=="جديد"]  if raw_results is not None else pd.DataFrame()
        dups_all   = raw_results[raw_results["الحالة"]=="مكرر"]  if raw_results is not None else pd.DataFrame()
        suspect_all= raw_results[raw_results["الحالة"]=="مشبوه"] if raw_results is not None else pd.DataFrame()

        st.markdown(f"""
        <div class="stats-bar">
          <div class="stat-box"><div class="n">{len(approved_df):,}</div><div class="lb">منتج معتمد نهائياً</div></div>
          <div class="stat-box"><div class="n" style="color:#43a047">{len(new_all):,}</div><div class="lb">جديد مؤكد</div></div>
          <div class="stat-box"><div class="n" style="color:#e53935">{len(dups_all):,}</div><div class="lb">مكرر محذوف</div></div>
          <div class="stat-box"><div class="n" style="color:#f9a825">{len(suspect_all):,}</div><div class="lb">مشبوه (راجعه AI)</div></div>
          <div class="stat-box"><div class="n" style="color:#7b1fa2">{len(new_brs):,}</div><div class="lb">ماركة جديدة</div></div>
        </div>
        """, unsafe_allow_html=True)

        _dd = int(st.session_state.get("pipe_store_dedup_dropped") or 0)
        if _dd > 0:
            st.caption(
                f"استُبعد تلقائياً {_dd:,} منتجاً لأنه يطابق منتجات موجودة في ملف المتجر "
                "(تطابق اسم fuzzy 90٪ أو تطابق SKU)."
            )

        pfs = st.session_state.get("pipe_filter_stats")
        if pfs:
            with st.expander("📊 إحصائيات فلاتر الاستبعاد (آخر مسار)", expanded=False):
                st.json(pfs)

        if not approved_df.empty:
            st.markdown("""<div class="sec-title"><div class="bar"></div>
            <h3>جدول المنتجات المعتمدة — تعديل مباشر</h3></div>""", unsafe_allow_html=True)
            pdf = approved_df.copy()
            # عمود مؤقت لتحديد الصفوف التي سيتم إثراؤها بالذكاء الاصطناعي
            if "_تحديد_للإثراء" not in pdf.columns:
                pdf["_تحديد_للإثراء"] = False
            all_pc = list(pdf.columns)
            show_default_p = [c for c in EDITOR_COLS if c in all_pc]
            if "_تحديد_للإثراء" in pdf.columns and "_تحديد_للإثراء" not in show_default_p:
                show_default_p = ["_تحديد_للإثراء"] + show_default_p
            show_cols_p = st.multiselect(
                "الأعمدة المعروضة:", options=all_pc, default=show_default_p or all_pc[:10],
                key="pipe_show_cols")
            if not show_cols_p:
                show_cols_p = show_default_p or all_pc[:10]
            if "_تحديد_للإثراء" not in show_cols_p and "_تحديد_للإثراء" in pdf.columns:
                show_cols_p = ["_تحديد_للإثراء"] + list(show_cols_p)

            grid_df = pdf[show_cols_p].copy()
            for c in show_cols_p:
                if c != "_تحديد_للإثراء":
                    grid_df[c] = grid_df[c].fillna("")
            edited_pipe = st.data_editor(
                grid_df,
                use_container_width=True,
                num_rows="dynamic",
                height=440,
                key="pipe_main_grid",
            )
            for c in show_cols_p:
                pdf[c] = edited_pipe[c]
            if "_تحديد_للإثراء" in pdf.columns:
                pdf["_تحديد_للإثراء"] = pdf["_تحديد_للإثراء"].fillna(False).astype(bool)
            st.session_state.pipe_approved = pdf
            st.session_state.pipe_export_df = pdf.copy()

            # ── بدء إثراء AI للمنتجات المحددة ────────────────────────────
            selected_idx = []
            if "_تحديد_للإثراء" in pdf.columns:
                try:
                    selected_idx = pdf.index[pdf["_تحديد_للإثراء"] == True].tolist()
                except Exception:
                    selected_idx = []

            if selected_idx:
                st.caption(f"تم تحديد {len(selected_idx)} منتج/منتجات لإثراء AI.")
            else:
                st.caption("اختر صفوفاً عبر `_تحديد_للإثراء` ثم اضغط زر الإثراء.")

            if st.button(
                "✨ بدء الإثراء بالذكاء الاصطناعي للمنتجات المحددة",
                key="pipe_start_product_ai_enrich",
                type="primary",
                use_container_width=True,
            ):
                if not st.session_state.api_key:
                    st.error("أضف مفتاح Anthropic API من صفحة الإعدادات أولاً.")
                    st.stop()
                if not selected_idx:
                    st.warning("لا توجد صفوف محددة لإثراء AI.")
                    st.stop()

                prog = st.progress(0)
                total = len(selected_idx)
                for n, ix in enumerate(selected_idx):
                    raw_nm = str(pdf.at[ix, "أسم المنتج"]) if "أسم المنتج" in pdf.columns else ""
                    with st.spinner(f"جاري إثراء المنتج ({n+1}/{total}): {raw_nm[:55]}..."):
                        ai_out = _ai_enrich_product_row(raw_nm, st.session_state.api_key)
                    if ai_out.get("formatted_name") and ai_out.get("html_description"):
                        fn = ai_out.get("formatted_name", raw_nm)
                        pdf.at[ix, "أسم المنتج"] = fn
                        pdf.at[ix, "الوصف"] = ai_out.get("html_description", pdf.at[ix, "الوصف"])
                        if "وصف صورة المنتج" in pdf.columns:
                            pdf.at[ix, "وصف صورة المنتج"] = f"زجاجة {fn} الأصلية"
                        if "الماركة" in pdf.columns and ai_out.get("brand"):
                            pdf.at[ix, "الماركة"] = _clean_brand_value_for_salla_output(ai_out.get("brand", ""))
                    prog.progress(int((n + 1) / max(total, 1) * 100))

                st.session_state.pipe_approved = pdf
                st.session_state.pipe_export_df = pdf.copy()
                st.success("✅ اكتمل إثراء AI للمنتجات المحددة.")
                st.rerun()

        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>⬇️ التصدير النهائي</h3></div>""", unsafe_allow_html=True)
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        approved_df = get_latest_export_data()

        if approved_df is not None and not approved_df.empty:
            pv_ok, pv_issues = validate_export_product_dataframe(approved_df)
            if not pv_ok:
                st.markdown(
                    '<div class="al-warn">⚠️ تحقق التصدير — راجع النقاط التالية قبل الرفع:<br>'
                    + "<br>".join(pv_issues[:18]) + "</div>",
                    unsafe_allow_html=True,
                )
                APP_LOG.warning("pipeline export validation: %s", pv_issues[:25])
        if new_brs:
            bv_ok, bv_issues = validate_export_brands_list(new_brs)
            if not bv_ok:
                st.markdown(
                    '<div class="al-warn">⚠️ تحقق الماركات الجديدة:<br>'
                    + "<br>".join(bv_issues[:8]) + "</div>",
                    unsafe_allow_html=True,
                )

        ex1, ex2, ex3, ex4 = st.columns(4)
        with ex1:
            if approved_df is not None and not approved_df.empty:
                st.download_button(
                    "📥 منتج جديد.csv — التصدير الرئيسي",
                    export_product_csv(get_latest_export_data()),
                    "منتج جديد.csv", "text/csv",
                    use_container_width=True, key="pipe_dl_csv_main"
                )
        with ex2:
            if approved_df is not None and not approved_df.empty:
                st.download_button(
                    f"📥 منتج جديد — Excel ({len(approved_df):,})",
                    export_product_xlsx(get_latest_export_data()),
                    f"منتج_جديد_{date_str}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="pipe_dl_xlsx"
                )
        with ex3:
            pass
        with ex4:
            if new_brs:
                st.download_button(
                    f"🏷️ ماركات جديدة ({len(new_brs)})",
                    export_brands_xlsx(new_brs),
                    f"ماركات_{date_str}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="pipe_dl_brands"
                )
            else:
                st.caption("لا ماركات جديدة")

        with st.expander("🛠️ أدوات إضافية على الجدول (توليد أوصاف، جلب صور، ماركات — دفعات 10)", expanded=False):
            pdf2 = st.session_state.pipe_approved
            if pdf2 is None or pdf2.empty:
                st.caption("لا بيانات.")
            else:
                t1, t2, t3, t4 = st.tabs(["🤖 أوصاف AI", "🖼 جلب صور", "🏷 ماركات", "⚡ مجمّع"])
                with t1:
                    if st.button("✨ توليد أوصاف (دفعات 10)", key="pipe_tab_ai", type="primary"):
                        if not st.session_state.api_key:
                            st.error("أضف مفتاح Claude في الإعدادات")
                        else:
                            dfw = pdf2.copy()
                            prog = st.progress(0)
                            idxs = [i for i in range(len(dfw)) if not str(dfw.iloc[i].get("الوصف", "")).strip()]
                            for n, i in enumerate(idxs):
                                row = dfw.iloc[i]
                                name = str(row.get("أسم المنتج", "")).strip()
                                if not name:
                                    continue
                                is_t = any(w in name.lower() for w in ["تستر", "tester"])
                                _mbr = _strip_brand_name_edges(row.get("الماركة", "") or "")
                                brand = {"name": _mbr,
                                         "page_url": to_slug(_mbr) if _mbr else ""}
                                size_m = re.search(r"\d+\s*(?:مل|ml)", name, re.I)
                                size = size_m.group() if size_m else "100 مل"
                                gender = ("للنساء" if any(w in name for w in ["نسائ", "women"]) else
                                          "للرجال" if any(w in name for w in ["رجال", "men"]) else "للجنسين")
                                dfw.at[dfw.index[i], "الوصف"] = ai_generate(
                                    name, is_t, brand, size, gender, "أو دو بارفيوم")
                                prog.progress(int((n + 1) / max(len(idxs), 1) * 100))
                            st.session_state.pipe_approved = dfw
                            st.session_state.pipe_export_df = dfw.copy()
                            st.success("✅ تم")
                            st.rerun()
                with t2:
                    if st.button("🖼 جلب صور ناقصة (دفعات 10)", key="pipe_tab_img", type="primary"):
                        dfw = pdf2.copy()
                        prog = st.progress(0)
                        idxs = [i for i in range(len(dfw)) if not str(dfw.iloc[i].get("صورة المنتج", "")).strip()]
                        for n, i in enumerate(idxs):
                            name = str(dfw.iloc[i].get("أسم المنتج", "")).strip()
                            if not name:
                                continue
                            is_t = any(w in name.lower() for w in ["تستر", "tester"])
                            u = fetch_image(name, is_t)
                            if u:
                                dfw.at[dfw.index[i], "صورة المنتج"] = u
                            prog.progress(int((n + 1) / max(len(idxs), 1) * 100))
                        st.session_state.pipe_approved = dfw
                        st.session_state.pipe_export_df = dfw.copy()
                        st.success("✅ تم")
                        st.rerun()
                with t3:
                    if st.button("🏷 تعيين ماركات ناقصة", key="pipe_tab_br", type="primary"):
                        dfw = pdf2.copy()
                        for i in range(len(dfw)):
                            if str(dfw.iloc[i].get("الماركة", "")).strip():
                                continue
                            name = str(dfw.iloc[i].get("أسم المنتج", "")).strip()
                            if not name:
                                continue
                            b = match_brand(name)
                            if b.get("name"):
                                dfw.at[dfw.index[i], "الماركة"] = b["name"]
                        st.session_state.pipe_approved = dfw
                        st.session_state.pipe_export_df = dfw.copy()
                        st.rerun()
                with t4:
                    st.info("SEO لا يتم تصديره من مسار المقارنة. استخدم تبويب `SEO Processor` بعد إنشاء المنتجات في سلة ومنحها `No.`.")

        # معاينة الماركات الجديدة
        if new_brs:
            with st.expander(f"🏷️ معاينة الماركات الجديدة ({len(new_brs)} ماركة)"):
                st.dataframe(pd.DataFrame(new_brs), use_container_width=True)

        # ماركات مفقودة مستخرجة من المنافس (عند غياب عمود الماركة) + إثراء AI + شعار تلقائي
        mb_df = st.session_state.get("pipe_missing_brands_df")
        if mb_df is not None and not mb_df.empty:
            st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
            <h3>ماركات مفقودة للرفع أولاً</h3></div>""", unsafe_allow_html=True)
            st.caption(
                "تم استخراج اسم العلامة عبر الذكاء الاصطناعي من عنوان المنتج (بدون افتراض «أول كلمتين») "
                "لأنها غير موجودة في مرجع الماركات (عتبة تطابق 80٪ بعد التوحيد)."
            )
            if st.button("✨ بدء الإثراء بالذكاء الاصطناعي", key="pipe_enrich_missing_brands", type="primary"):
                if not st.session_state.api_key:
                    st.error("أضف مفتاح Anthropic API من صفحة الإعدادات أولاً.")
                else:
                    w = mb_df.copy()
                    prog = st.progress(0)
                    for i in range(len(w)):
                        bname = str(w.iloc[i].get("اسم الماركة", "") or "").strip()
                        if not bname:
                            prog.progress(int((i + 1) / max(len(w), 1) * 100))
                            continue
                        ai_b = _ai_enrich_brand_row_with_domain(bname, st.session_state.api_key)
                        if ai_b.get("brand_name"):
                            w.at[w.index[i], "اسم الماركة"] = ai_b["brand_name"]
                        if ai_b.get("description"):
                            w.at[w.index[i], "وصف مختصر عن الماركة"] = ai_b["description"]
                        if ai_b.get("seo_description"):
                            w.at[w.index[i], "(Page Description) وصف صفحة العلامة التجارية"] = ai_b["seo_description"]
                        dom = str(ai_b.get("domain", "") or "").strip().lower()
                        dom = re.sub(r"^https?://", "", dom).split("/")[0].strip()
                        dom = re.sub(r"^www\.", "", dom)
                        logo_u = (ai_b.get("logo_clearbit_url") or "").strip()
                        if not logo_u and dom and "." in dom:
                            logo_u = f"https://logo.clearbit.com/{dom}"
                        w.at[w.index[i], "صورة شعار الماركة"] = logo_u
                        if ai_b.get("brand_name"):
                            register_pipe_session_brand(
                                ai_b["brand_name"],
                                ai_b.get("page_url", "") or "",
                            )
                        if ai_b.get("page_title"):
                            w.at[w.index[i], "(Page Title) عنوان صفحة العلامة التجارية"] = ai_b["page_title"]
                        else:
                            w.at[w.index[i], "(Page Title) عنوان صفحة العلامة التجارية"] = w.iloc[i].get(
                                "(Page Title) عنوان صفحة العلامة التجارية", ""
                            )
                        if ai_b.get("page_url"):
                            w.at[w.index[i], "(SEO Page URL) رابط صفحة العلامة التجارية"] = ai_b["page_url"]
                        else:
                            w.at[w.index[i], "(SEO Page URL) رابط صفحة العلامة التجارية"] = w.iloc[i].get(
                                "(SEO Page URL) رابط صفحة العلامة التجارية", to_slug(bname)
                            )
                        prog.progress(int((i + 1) / max(len(w), 1) * 100))
                    st.session_state.pipe_missing_brands_df = w[SALLA_BRANDS_COLS].copy()
                    st.success("✅ اكتمل إثراء الماركات المفقودة.")
                    st.rerun()

            edited_mb = st.data_editor(
                st.session_state.pipe_missing_brands_df[SALLA_BRANDS_COLS].fillna(""),
                num_rows="dynamic",
                use_container_width=True,
                key="pipe_missing_brands_editor",
            )
            st.session_state.pipe_missing_brands_df = edited_mb[SALLA_BRANDS_COLS].copy()
            st.download_button(
                "📥 تنزيل ماركات_جديدة_للرفع_أولا.csv",
                edited_mb[SALLA_BRANDS_COLS].to_csv(index=False, encoding="utf-8-sig"),
                "ماركات_جديدة_للرفع_أولا.csv",
                "text/csv",
                use_container_width=True,
                key="pipe_missing_brands_dl_csv",
            )

        # إعادة الضبط
        st.divider()
        if st.button("🔄 مسار جديد (إعادة ضبط)", key="pipe_reset"):
            st.session_state.pipe_store_df   = None
            st.session_state.pipe_comp_dfs   = []
            st.session_state.pipe_results    = None
            st.session_state.pipe_approved   = None
            st.session_state.pipe_export_df  = None
            st.session_state.pipe_new_brands = []
            st.session_state.pipe_session_brands = []
            st.session_state.pipe_missing_brands_df = None
            st.session_state.pipe_seo_df     = None
            st.session_state.pipe_step       = 0
            if "_brand_entity_ai_cache" in st.session_state:
                del st.session_state["_brand_entity_ai_cache"]
            st.rerun()

        st.markdown("""<div class="al-info" style="margin-top:14px">
        <b>خطوة اختيارية:</b> بعد التصدير يمكنك مراجعة المنتجات المشبوهة بصرياً مقابل المتجر قبل اعتماد القائمة النهائية.
        </div>""", unsafe_allow_html=True)
        if st.button("🔀 الانتقال للمقارنة والتدقيق المرئي", key="pipe_to_visual", use_container_width=True):
            if (st.session_state.pipe_store_df is not None
                    and st.session_state.pipe_approved is not None
                    and not st.session_state.pipe_approved.empty):
                st.session_state.cmp_new_df = st.session_state.pipe_approved.copy()
                st.session_state.cmp_store_df = st.session_state.pipe_store_df.copy()
                st.session_state.cmp_results = None
                st.session_state.cmp_approved = {}
                st.session_state.cmp_edit_name = {}
                st.session_state.cmp_from_pipe = True
                st.session_state.page = "compare"
                st.rerun()
            else:
                st.warning("يلزم إكمال المسار الآلي بملف متجر ومنتجات معتمدة أولاً.")


# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE — SEO PROCESSOR (معالج SEO مستقل)                        ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "seo_processor":
    render_seo_processor_tab()
    st.stop()

    st.markdown("""<div class="al-info">
    ارفع ملف منتجات سلة كامل (Excel أو CSV). يكتشف النظام المنتجات التي تفتقد
    <strong>رابط SEO</strong> أو <strong>عنوان الصفحة</strong> أو <strong>وصف الصفحة</strong>،
    ثم يولّدها بالذكاء الاصطناعي بنفس تنسيق ملف <strong>سلة SEO</strong>.
    </div>""", unsafe_allow_html=True)

    up_seo = st.file_uploader(
        "ارفع ملف المتجر أو المنتجات",
        type=["csv", "xlsx", "xls", "xlsm"],
        key="seo_proc_uploader",
    )
    if up_seo:
        df_seo = read_file(up_seo, salla_2row=True)
        if df_seo.empty:
            df_seo = read_file(up_seo, salla_2row=False)
        if not df_seo.empty:
            st.session_state.seo_proc_df = df_seo
            st.success(f"✅ تم تحميل {len(df_seo):,} صف")
            st.rerun()

    if st.session_state.seo_proc_df is not None:
        sdf = st.session_state.seo_proc_df
        with st.expander("👀 معاينة الملف", expanded=False):
            st.dataframe(sdf.head(12), use_container_width=True)

        if st.button("🤖 تحليل وإكمال SEO الناقص", type="primary", key="seo_proc_run", use_container_width=True):
            if not st.session_state.api_key:
                st.error("أضف مفتاح Anthropic API من صفحة الإعدادات.")
            else:
                prog = st.progress(0)
                st.session_state._seo_batch_prog = prog
                full_seo, gen_only = generate_seo_for_products_dataframe(sdf)
                st.session_state._seo_batch_prog = None
                st.session_state.seo_proc_full = full_seo
                st.session_state.seo_proc_gen = gen_only
                st.success(f"✅ جاهز: صفوف مُولّدة جديدة: {len(gen_only):,} من أصل {len(sdf):,}")
                st.rerun()

        if getattr(st.session_state, "seo_proc_gen", None) is not None:
            gen_df = st.session_state.seo_proc_gen
            if gen_df.empty:
                st.info("لا توجد صفوف تحتاج توليد SEO — جميع الحقول مكتملة.")
            else:
                st.markdown("""<div class="sec-title"><div class="bar"></div>
                <h3>نتيجة التوليد (تنسيق سلة SEO)</h3></div>""", unsafe_allow_html=True)
                st.data_editor(gen_df.fillna(""), use_container_width=True, num_rows="dynamic", key="seo_proc_editor")
                date_s = datetime.now().strftime("%Y-%m-%d_%H-%M")
                st.download_button(
                    "📥 تصدير SEO — Excel (المُولّد فقط)",
                    export_seo_xlsx(gen_df),
                    f"seo_mahwous_{date_s}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="seo_proc_dl",
                )
        if st.button("🗑️ مسح ملف المعالج", key="seo_proc_clear"):
            st.session_state.seo_proc_df = None
            st.session_state.seo_proc_gen = None
            if hasattr(st.session_state, "seo_proc_full"):
                del st.session_state["seo_proc_full"]
            st.rerun()



# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE — COMPARE (routed)                                        ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "compare":

    st.markdown("""<div class="al-info">
    <b>المقارنة:</b> راجع نتائج مقارنة المنافسين مع ملف المتجر، ثم اعتمد أو عدّل النتائج قبل التصدير.
    </div>""", unsafe_allow_html=True)
    render_compare_tab()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE — STORE AUDIT (routed)                                    ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "store_audit":
    render_store_audit_tab()

# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 3 — QUICK ADD                                             ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "quickadd":
    render_quick_add_tab()
    st.stop()

    st.markdown("""<div class="al-info">
    أضف منتجات جديدة بسرعة بطريقتين: (1) من خلال رابط منتج أو أكثر، أو (2) بإدخال يدوي سريع مع رفع صور.
    </div>""", unsafe_allow_html=True)

    tab1, tab2 = st.tabs(["🔗 سحب من رابط", "📝 إدخال يدوي ورفع صور"])

    # ── TAB 1: Fetch from URLs (multiple) ───────────────────────
    with tab1:
        st.markdown("### سحب بيانات المنتجات من روابط")
        st.caption("أدخل رابطاً واحداً أو أكثر — كل رابط في سطر منفصل")

        # Initialize URL list in session state
        if "qa_url_list" not in st.session_state:
            st.session_state.qa_url_list = [""]

        # Dynamic URL inputs
        urls_to_remove = []
        for idx, url_val in enumerate(st.session_state.qa_url_list):
            col_url, col_del = st.columns([10, 1])
            with col_url:
                new_val = st.text_input(
                    f"رابط المنتج {idx + 1}",
                    value=url_val,
                    placeholder="https://example.com/product/...",
                    key=f"qa_url_{idx}",
                    label_visibility="collapsed",
                )
                st.session_state.qa_url_list[idx] = new_val
            with col_del:
                if len(st.session_state.qa_url_list) > 1:
                    if st.button("✕", key=f"del_url_{idx}", help="حذف هذا الرابط"):
                        urls_to_remove.append(idx)

        for idx in reversed(urls_to_remove):
            st.session_state.qa_url_list.pop(idx)
            st.rerun()

        col_add, col_fetch = st.columns([1, 3])
        with col_add:
            if st.button("➕ إضافة رابط آخر", width='stretch'):
                st.session_state.qa_url_list.append("")
                st.rerun()
        with col_fetch:
            do_fetch = st.button("🔄 سحب البيانات من الروابط", type="primary",
                                 width='stretch', key="qa_fetch_urls")

        # Options for URL scraping
        uo1, uo2 = st.columns(2)
        with uo1:
            qa_url_gen_desc = st.checkbox("🤖 توليد وصف AI لكل منتج", value=True, key="qa_url_gen_desc")
        with uo2:
            qa_url_gen_seo = st.checkbox("🔍 توليد SEO", value=True, key="qa_url_gen_seo")

        if do_fetch:
            valid_urls = [u.strip() for u in st.session_state.qa_url_list if u.strip()]
            if not valid_urls:
                st.error("الرجاء إدخال رابط منتج واحد على الأقل")
            else:
                progress_bar = st.progress(0, text="جاري السحب...")
                success_count = 0
                for url_i, url_item in enumerate(valid_urls):
                    progress_bar.progress(
                        int((url_i / len(valid_urls)) * 100),
                        text=f"جاري سحب المنتج {url_i + 1} من {len(valid_urls)}: {url_item[:60]}..."
                    )
                    with st.spinner(f"سحب: {url_item[:80]}..."):
                        scraped = scrape_product_url(url_item)

                    if scraped.get("error"):
                        st.warning(f"⚠️ تعذّر سحب {url_item[:60]}: {scraped['error']}")
                        continue

                    ex_name  = scraped.get("name", "") or "منتج جديد"
                    ex_price = scraped.get("price", "") or ""
                    ex_img   = scraped.get("image", "") or ""
                    ex_imgs  = scraped.get("images", [])
                    ex_desc  = scraped.get("desc", "") or ""
                    ex_brand_hint = scraped.get("brand_hint", "") or ""

                    # Show preview card
                    with st.expander(f"📦 {ex_name[:80]}", expanded=True):
                        pc1, pc2 = st.columns([1, 3])
                        with pc1:
                            if ex_img:
                                st.image(ex_img, width=120, caption="الصورة الرئيسية")
                            if len(ex_imgs) > 1:
                                st.caption(f"📷 {len(ex_imgs)} صور متاحة")
                                thumb_cols = st.columns(min(len(ex_imgs), 4))
                                for ti, timg in enumerate(ex_imgs[:4]):
                                    with thumb_cols[ti]:
                                        try:
                                            st.image(timg, width=60)
                                        except Exception:
                                            pass
                        with pc2:
                            st.markdown(f"**الاسم:** {ex_name}")
                            st.markdown(f"**السعر:** {ex_price} ريال" if ex_price else "**السعر:** غير محدد")
                            st.markdown(f"**الماركة المكتشفة:** {ex_brand_hint}" if ex_brand_hint else "")
                            st.caption(f"**الوصف:** {ex_desc[:200]}..." if len(ex_desc) > 200 else f"**الوصف:** {ex_desc}")

                    # Match brand
                    if ex_brand_hint:
                        brand = match_brand(ex_brand_hint)
                        if not brand.get("name"):
                            brand = generate_new_brand(ex_brand_hint)
                            existing_b = [b.get("اسم العلامة التجارية", "") for b in st.session_state.new_brands]
                            if ex_brand_hint not in existing_b:
                                st.session_state.new_brands.append({
                                    "اسم العلامة التجارية": brand.get("name", ex_brand_hint),
                                    "(SEO Page URL) رابط صفحة العلامة التجارية": brand.get("page_url", to_slug(ex_brand_hint)),
                                    "وصف العلامة التجارية": brand.get("desc", ""),
                                    "صورة العلامة التجارية": "",
                                })
                    else:
                        brand = match_brand(ex_name)

                    ex_name = standardize_product_name(ex_name, brand.get("name", ""))
                    cat = match_category(ex_name, "للجنسين")

                    # Extract size from name
                    size_match = re.search(r'(\d+)\s*(?:ml|مل|ML)', ex_name, re.IGNORECASE)
                    ex_size = size_match.group(0) if size_match else "100 مل"
                    ex_size = _normalize_product_size_ml(ex_size) or ex_size

                    _qa_sku = f"V-{len(st.session_state.get('qa_rows', [])) + 1}"
                    seo = (
                        gen_seo(
                            ex_name, brand, ex_size, False, "للجنسين",
                            sku_suffix=_qa_sku,
                        )
                        if qa_url_gen_seo
                        else {"url": "", "title": "", "desc": ""}
                    )

                    # Generate AI description
                    if qa_url_gen_desc and st.session_state.api_key:
                        final_desc = ai_generate(ex_name, False, brand, ex_size, "للجنسين", "أو دو بارفيوم")
                    else:
                        final_desc = f"<p>{ex_desc}</p>" if ex_desc else f"<p>وصف مبدئي لـ {ex_name}</p>"

                    nr = fill_row(
                        name=ex_name, price=ex_price, image=ex_img,
                        desc=final_desc, brand=brand, category=cat,
                        seo=seo, weight="0.2"
                    )
                    st.session_state.qa_rows.append({
                        "product": nr,
                        "seo": {"url": seo["url"], "title": seo["title"], "desc": seo["desc"]},
                        "images": ex_imgs,
                    })
                    success_count += 1

                progress_bar.progress(100, text="اكتمل السحب!")
                if success_count:
                    st.success(f"✅ تم سحب {success_count} منتج بنجاح!")
                    # Reset URL list
                    st.session_state.qa_url_list = [""]
                    st.rerun()

    # ── TAB 2: Manual Entry with Image Upload ────────────────────
    with tab2:
        with st.form("qa_form", clear_on_submit=True):
            f1, f2, f3 = st.columns(3)
            with f1:
                qa_nm = st.text_input("اسم العطر ⭐", placeholder="مثال: شانيل بلو دو شانيل 100 مل للرجال")
                qa_pr = st.text_input("السعر ⭐", placeholder="299")
            with f2:
                qa_gn = st.selectbox("الجنس", ["للجنسين","للرجال","للنساء"])
                qa_sk = st.text_input("SKU", placeholder="اختياري")
            with f3:
                qa_sz = st.text_input("الحجم", "100 مل")
                qa_cn = st.selectbox("التركيز", ["أو دو بارفيوم","أو دو كولون","أو دو تواليت","بارفيوم"])

            f4, f5, f6 = st.columns(3)
            with f4: qa_tp   = st.selectbox("النوع", ["عطر عادي","تستر"])
            with f5: qa_wt   = st.text_input("الوزن (kg)", "0.2")
            with f6: qa_br   = st.text_input("الماركة (اختياري)")
            
            st.markdown("**صور المنتج**")
            qa_imgs = st.file_uploader("ارفع صورة أو صورتين", type=["png","jpg","jpeg","webp"], accept_multiple_files=True)

            o1, o2, o3 = st.columns(3)
            with o1: qa_do_d = st.checkbox("🤖 توليد وصف AI",   value=True)
            with o2: qa_do_i = st.checkbox("🖼 جلب صورة من جوجل (إذا لم ترفع)", value=False)
            with o3: qa_do_s = st.checkbox("🔍 توليد SEO",       value=True)

            sub = st.form_submit_button("➕ إضافة للقائمة وتجهيز الملف", type="primary", width='stretch')

        if sub:
            if not qa_nm.strip() or not qa_pr.strip():
                st.error("الاسم والسعر حقول إجبارية!")
            else:
                with st.spinner("جاري التجهيز..."):
                    is_t   = qa_tp == "تستر"
                    if qa_br.strip():
                        brand = {"name": qa_br.strip(), "page_url": to_slug(qa_br.strip())}
                        existing_brands = [b["اسم العلامة التجارية"] for b in st.session_state.new_brands]
                        if match_brand(qa_nm).get("name") == "" and qa_br not in existing_brands:
                            st.session_state.new_brands.append({
                                "اسم العلامة التجارية": qa_br.strip(),
                                "(SEO Page URL) رابط صفحة العلامة التجارية": to_slug(qa_br.strip()),
                                "وصف العلامة التجارية": "",
                                "صورة العلامة التجارية": "",
                            })
                    else:
                        brand  = match_brand(qa_nm)
                    
                    cat    = match_category(qa_nm, qa_gn)
                    qa_sku_slug = qa_sk.strip() if qa_sk.strip() else f"V-{len(st.session_state.get('qa_rows', [])) + 1}"
                    seo    = gen_seo(
                        qa_nm, brand, qa_sz, is_t, qa_gn,
                        sku_suffix=qa_sku_slug,
                    )
                    
                    # Handle Images
                    img_url = ""
                    if qa_imgs:
                        st.warning(
                            "الصورة المرفوعة محلياً لا تُرفع تلقائياً إلى سلة — "
                            "أضف رابط الصورة في المنتج بعد رفعها من لوحة سلة، أو استخدم «جلب صورة»."
                        )
                        img_url = ""
                    elif qa_do_i:
                        img_url = fetch_image(qa_nm, is_t)
                        
                    desc   = ai_generate(qa_nm, is_t, brand, qa_sz, qa_gn, qa_cn) if qa_do_d else ""
                    nr     = fill_row(name=qa_nm, price=qa_pr, sku=qa_sk, image=img_url,
                                      desc=desc, brand=brand, category=cat, seo=seo,
                                      weight=qa_wt)
                    st.session_state.qa_rows.append({
                        "product": nr,
                        "seo": {"url": seo["url"], "title": seo["title"], "desc": seo["desc"]},
                    })
                st.success(f"✅ تمت الإضافة: **{qa_nm}**")

    # ── Common List Display ───────────────────────────────────────
    if st.session_state.qa_rows:
        st.markdown(f"### القائمة ({len(st.session_state.qa_rows)} منتج)")
        prev = []
        for r in st.session_state.qa_rows:
            p = r["product"]
            prev.append({
                "الاسم":    p.get("أسم المنتج",""),
                "الماركة":  p.get("الماركة",""),
                "التصنيف":  p.get("تصنيف المنتج",""),
                "السعر":    p.get("سعر المنتج",""),
                "الوزن":    p.get("الوزن",""),
                "وصف ✓":   "✅" if str(p.get("الوصف","")).strip() else "—",
                "صورة ✓":  "✅" if str(p.get("صورة المنتج","")).startswith("http") else "—",
            })
        st.dataframe(pd.DataFrame(prev), use_container_width=True)

        prod_df_qa = pd.DataFrame([r["product"] for r in st.session_state.qa_rows])
        seo_df_qa  = pd.DataFrame([{
            "No. (غير قابل للتعديل)": "",
            "اسم المنتج (غير قابل للتعديل)": r["product"]["أسم المنتج"],
            "رابط مخصص للمنتج (SEO Page URL)": r["seo"]["url"],
            "عنوان صفحة المنتج (SEO Page Title)": r["seo"]["title"],
            "وصف صفحة المنتج (SEO Page Description)": r["seo"]["desc"],
        } for r in st.session_state.qa_rows])

        qe1, qe2, qe3, qe4, qe5 = st.columns(5)
        with qe1:
            st.download_button("📥 منتجات Excel",
                export_product_xlsx(prod_df_qa), "qa_products.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch')
        with qe2:
            st.download_button("📥 منتجات CSV",
                export_product_csv(prod_df_qa), "qa_products.csv", "text/csv",
                width='stretch')
        with qe3:
            st.download_button("📥 SEO Excel",
                export_seo_xlsx(seo_df_qa), "qa_seo.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch')
        with qe4:
            st.download_button("📥 SEO CSV",
                export_seo_csv(seo_df_qa), "qa_seo.csv", "text/csv",
                width='stretch')
        with qe5:
            if st.button("🔀 نقل للمسار الآلي", width='stretch', key="move_qa"):
                ex = st.session_state.up_df
                combined = pd.concat(
                    [ex, prod_df_qa], ignore_index=True) if ex is not None else prod_df_qa
                st.session_state.pipe_approved = combined
                st.session_state.pipe_export_df = combined.copy()
                st.session_state.pipe_seo_df = seo_df_qa
                st.session_state.pipe_step = 5
                st.session_state.pipe_results = None
                st.session_state.qa_rows = []
                st.session_state.page = "pipeline"
                st.rerun()

        if st.button("🗑️ مسح القائمة", key="clear_qa"):
            st.session_state.qa_rows = []
            st.rerun()


# ╔══════════════════════════════════════════════════════════════════╗
# ║  PAGE 6 — SETTINGS                                              ║
# ╚══════════════════════════════════════════════════════════════════╝
elif st.session_state.page == "settings":

    st.markdown("""<div class="sec-title"><div class="bar"></div>
    <h3>مفاتيح API</h3></div>""", unsafe_allow_html=True)

    s1, s2 = st.columns(2)
    with s1:
        new_key = st.text_input("🔑 Anthropic API Key (Claude):",
                                value=st.session_state.api_key,
                                type="password", key="set_api")
        if st.button("💾 حفظ مفتاح Claude", key="save_api"):
            st.session_state.api_key = new_key
            st.success("✅ تم حفظ المفتاح")
    with s2:
        new_gk = st.text_input("🔑 Google API Key:",
                               value=st.session_state.google_api,
                               type="password", key="set_gk")
        new_cx = st.text_input("🔍 Google CSE ID:",
                               value=st.session_state.google_cse,
                               type="password", key="set_cx")
        if st.button("💾 حفظ مفاتيح Google", key="save_gk"):
            st.session_state.google_api = new_gk
            st.session_state.google_cse = new_cx
            st.success("✅ تم حفظ المفاتيح")

    st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
    <h3>قواعد البيانات المرجعية</h3></div>""", unsafe_allow_html=True)

    db1, db2 = st.columns(2)
    with db1:
        st.markdown("**ملف الماركات**")
        bdf = st.session_state.brands_df
        if bdf is not None:
            st.markdown(f'<div class="al-ok">{len(bdf)} ماركة محملة</div>',
                        unsafe_allow_html=True)
            with st.expander("👀 معاينة"): st.dataframe(bdf.head(5), use_container_width=True)
        up_brands = st.file_uploader("تحديث ملف الماركات:", type=["csv","xlsx"],
                                      key="up_brands_db")
        if up_brands:
            df_b = read_file(up_brands)
            if not df_b.empty:
                st.session_state.brands_df = df_b
                os.makedirs(DATA_DIR, exist_ok=True)
                df_b.to_csv(os.path.join(DATA_DIR, "brands.csv"),
                            index=False, encoding="utf-8-sig")
                st.success(f"✅ تم تحديث {len(df_b)} ماركة")
                st.rerun()

    with db2:
        st.markdown("**ملف التصنيفات**")
        cdf = st.session_state.categories_df
        if cdf is not None:
            st.markdown(f'<div class="al-ok">{len(cdf)} تصنيف محمّل</div>',
                        unsafe_allow_html=True)
            with st.expander("👀 معاينة"): st.dataframe(cdf.head(5), use_container_width=True)
        up_cats = st.file_uploader("تحديث ملف التصنيفات:", type=["csv","xlsx"],
                                    key="up_cats_db")
        if up_cats:
            df_c = read_file(up_cats)
            if not df_c.empty:
                st.session_state.categories_df = df_c
                os.makedirs(DATA_DIR, exist_ok=True)
                df_c.to_csv(os.path.join(DATA_DIR, "categories.csv"),
                            index=False, encoding="utf-8-sig")
                st.success(f"✅ تم تحديث {len(df_c)} تصنيف")
                st.rerun()

    # New brands export section
    if st.session_state.new_brands:
        st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
        <h3>الماركات الجديدة المُولَّدة — جاهزة للتصدير</h3></div>""", unsafe_allow_html=True)
        st.markdown(f'<div class="al-warn">{len(st.session_state.new_brands)} ماركة جديدة اكتُشفت خلال المعالجة وتحتاج إلى إضافتها لمتجرك على سلة.</div>',
                    unsafe_allow_html=True)
        nb_df_s = pd.DataFrame(st.session_state.new_brands)
        st.dataframe(nb_df_s, use_container_width=True)
        sn1, sn2 = st.columns(2)
        with sn1:
            st.download_button("📥 تصدير الماركات الجديدة — Excel",
                export_brands_xlsx(st.session_state.new_brands),
                "new_brands_salla.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', key="exp_nb_set_x")
        with sn2:
            nb_csv_s = io.StringIO()
            nb_csv_s.write(",".join(SALLA_BRANDS_COLS) + "\n")
            for nb in st.session_state.new_brands:
                nb_csv_s.write(",".join([f'"{str(nb.get(c,"") or "")}"'
                                          for c in SALLA_BRANDS_COLS]) + "\n")
            st.download_button("📥 تصدير الماركات الجديدة — CSV",
                nb_csv_s.getvalue().encode("utf-8-sig"),
                "new_brands_salla.csv", "text/csv",
                width='stretch', key="exp_nb_set_c")

    import os
    import pandas as pd
    import asyncio
    from utils.async_scraper import run_scraper_engine

    st.markdown("---")
    st.subheader("🤖 تشغيل محرك الكشط وعرض النتائج")
    st.info("سيسحب هذا المحرك أحدث أسعار المنافسين بناءً على الروابط المدخلة ويعرضها فوراً.")

    # 1. The Scraper Button
    if st.button("🚀 بدء جلب بيانات المنافسين الآن", use_container_width=True):
        with st.spinner("جاري اختراق الموقع وسحب البيانات... يرجى الانتظار."):
            try:
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                loop.run_until_complete(run_scraper_engine())
                st.success("✅ تمت عملية الكشط بنجاح!")
                # CRITICAL: Force Streamlit to refresh and read the newly created file
                st.rerun()
            except Exception as e:
                st.error(f"❌ حدث خطأ أثناء الكشط: {str(e)}")

    # 2. Persistent Data Viewer (Always visible if the file exists)
    st.markdown("### 📊 البيانات المسحوبة من المنافسين")
    data_path = os.path.join(os.getcwd(), "data", "competitors_latest.csv")

    if os.path.exists(data_path):
        try:
            df_comp = pd.read_csv(data_path)
            if df_comp.empty:
                st.warning("⚠️ تمت عملية الكشط، ولكن الملف فارغ! تأكد أن رابط الـ Sitemap صحيح ويحتوي على منتجات.")
            else:
                st.success(f"✅ تم العثور على {len(df_comp)} منتج مسحوب وجاهز للمطابقة.")
                st.dataframe(df_comp, use_container_width=True, height=400)
        except Exception as e:
            st.error(f"❌ حدث خطأ في قراءة ملف البيانات: {str(e)}")
    else:
        st.warning("⚠️ لا توجد أي بيانات مسحوبة حالياً. اضغط على زر الجلب أعلاه للبدء.")

    st.markdown("""<hr class="gdiv"><div class="sec-title"><div class="bar"></div>
    <h3>معلومات النظام</h3></div>""", unsafe_allow_html=True)
    st.markdown(f"""
    <div style="direction:rtl;font-size:0.85rem;line-height:2">
      <b>الإصدار:</b> مهووس مركز التحكم الشامل v11.0<br>
      <b>لوحة التطبيق:</b> <a href="{PUBLIC_APP_URL}/" target="_blank" rel="noopener">{PUBLIC_APP_URL}</a><br>
      <b>أعمدة سلة المنتجات:</b> {len(SALLA_COLS)} عمود<br>
      <b>أعمدة سلة SEO:</b> {len(SALLA_SEO_COLS)} عمود<br>
      <b>أعمدة تحديث الأسعار:</b> {len(SALLA_PRICE_COLS)} عمود<br>
      <b>أعمدة ملف الماركات:</b> {len(SALLA_BRANDS_COLS)} عمود<br>
    </div>
    """, unsafe_allow_html=True)

# ╔══════════════════════════════════════════════════════════════════╗
# ║  FOOTER                                                         ║
# ╚══════════════════════════════════════════════════════════════════╝
st.markdown(f"""
<div class="mhw-footer">
  مهووس — مركز التحكم الشامل v11.0 &nbsp;|&nbsp;
  جميع الملفات المُصدَّرة متوافقة 100% مع منصة سلة &nbsp;|&nbsp;
  <a href="{PUBLIC_APP_URL}/" target="_blank" rel="noopener">لوحة التشغيل</a>
</div>
""", unsafe_allow_html=True)
